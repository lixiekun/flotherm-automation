#!/usr/bin/env python3
"""
Excel Batch Driver - Excel 驱动的批量处理

从 Excel 配置文件读取批量处理配置。

Usage:
    from pack_editor.batch import ExcelBatchDriver

    driver = ExcelBatchDriver("template.pack", "cases.xlsx")
    driver.run("./results/")
"""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Union

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .batch_editor import BatchEditor, BatchResult


@dataclass
class CaseConfig:
    """用例配置"""
    case_id: str
    changes: Dict[str, float]  # component_name -> power


class ExcelBatchDriver:
    """
    Excel 驱动的批量处理

    从 Excel/CSV 文件读取配置，批量修改 Pack 文件。
    """

    # 必需的列
    REQUIRED_COLUMNS = {"case_id"}

    def __init__(
        self,
        template_pack: Union[str, Path],
        config_file: Union[str, Path],
        calibration_rules: Optional[Dict[str, Union[str, Path]]] = None,
    ):
        """
        初始化 Excel 批量驱动

        Args:
            template_pack: 模板 Pack 文件
            config_file: 配置文件 (Excel 或 CSV)
            calibration_rules: 校准规则文件 {component_name: rule_path}
        """
        self.template_pack = Path(template_pack)
        self.config_file = Path(config_file)
        self.calibration_rules = calibration_rules or {}
        self.cases: List[CaseConfig] = []

        self._load_config()

    def _load_config(self) -> None:
        """加载配置文件"""
        suffix = self.config_file.suffix.lower()

        if suffix == ".csv":
            self._load_csv()
        elif suffix in {".xlsx", ".xlsm"}:
            self._load_excel()
        else:
            raise ValueError(f"Unsupported config file format: {suffix}")

    def _load_csv(self) -> None:
        """加载 CSV 配置"""
        with self.config_file.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        if not rows:
            raise ValueError("Config file is empty")

        # 验证必需列
        columns = set(rows[0].keys())
        missing = self.REQUIRED_COLUMNS - columns
        if missing:
            raise ValueError(f"Missing required columns: {missing}")

        # 解析用例
        for row in rows:
            case_id = row.get("case_id", "").strip()
            if not case_id:
                continue

            changes: Dict[str, float] = {}
            for col, value in row.items():
                if col == "case_id":
                    continue

                value = str(value).strip()
                if value:
                    try:
                        changes[col] = float(value)
                    except ValueError:
                        pass  # 忽略非数值

            if changes:
                self.cases.append(CaseConfig(case_id=case_id, changes=changes))

    def _load_excel(self) -> None:
        """加载 Excel 配置"""
        if not HAS_OPENPYXL:
            raise ImportError("Reading Excel requires openpyxl. Install it or use CSV.")

        workbook = openpyxl.load_workbook(self.config_file, read_only=True, data_only=True)
        sheet = workbook.active

        # 读取表头
        headers = []
        for cell in next(sheet.iter_rows(max_row=1)):
            value = cell.value
            headers.append(str(value).strip() if value else "")

        if not headers:
            raise ValueError("Excel file has no headers")

        # 验证必需列
        columns = set(headers)
        missing = self.REQUIRED_COLUMNS - columns
        if missing:
            raise ValueError(f"Missing required columns: {missing}")

        # 解析用例
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            case_id = ""
            changes: Dict[str, float] = {}

            for idx, value in enumerate(row):
                if idx >= len(headers):
                    break

                col = headers[idx]
                if not col:
                    continue

                if col == "case_id":
                    case_id = str(value).strip() if value else ""
                else:
                    if value is not None:
                        try:
                            changes[col] = float(value)
                        except (ValueError, TypeError):
                            pass

            if case_id and changes:
                self.cases.append(CaseConfig(case_id=case_id, changes=changes))

        workbook.close()

    def get_cases(self) -> List[CaseConfig]:
        """获取所有用例配置"""
        return self.cases.copy()

    def run(
        self,
        output_dir: Union[str, Path],
        sheet: Optional[str] = None,
        parallel: int = 1,
    ) -> Dict:
        """
        执行批量处理

        Args:
            output_dir: 输出目录
            sheet: Excel 工作表名 (可选)
            parallel: 并行数

        Returns:
            Dict: 运行结果
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        results: List[Dict] = []
        batch = BatchEditor()

        # 加载校准规则
        for component_name, rule_path in self.calibration_rules.items():
            batch.load_calibration_rule(component_name, rule_path)

        # 处理每个用例
        for case in self.cases:
            case_dir = output_dir / case.case_id
            case_dir.mkdir(parents=True, exist_ok=True)

            # 添加功耗修改
            for component_name, power in case.changes.items():
                batch.add_power_change(component_name, power)

            # 处理
            batch_results = batch.execute(
                output_dir=case_dir,
                naming="folder",
                parallel=parallel,
            )

            # 重命名输出文件
            for result in batch_results:
                if result.success and result.output_pack:
                    new_name = case_dir / f"{case.case_id}.pack"
                    result.output_pack.rename(new_name)

            results.append({
                "case_id": case.case_id,
                "changes": case.changes,
                "success": all(r.success for r in batch_results),
                "output_dir": str(case_dir),
            })

            # 清空功耗修改列表以准备下一个用例
            batch.power_changes.clear()

        # 保存运行清单
        manifest = {
            "template_pack": str(self.template_pack),
            "config_file": str(self.config_file),
            "total_cases": len(self.cases),
            "cases": results,
        }

        manifest_path = output_dir / "run_manifest.json"
        manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

        return manifest

    @classmethod
    def create_template_csv(cls, output_path: Union[str, Path]) -> Path:
        """
        创建 CSV 模板文件

        Args:
            output_path: 输出路径

        Returns:
            Path: 创建的文件路径
        """
        output_path = Path(output_path)

        template_content = """case_id,CPU,GPU,Memory,Ambient
case1,25.0,15.0,5.0,25.0
case2,30.0,20.0,6.0,35.0
case3,35.0,25.0,7.0,45.0
"""

        output_path.write_text(template_content, encoding="utf-8")
        return output_path

    @classmethod
    def create_template_excel(cls, output_path: Union[str, Path]) -> Path:
        """
        创建 Excel 模板文件

        Args:
            output_path: 输出路径

        Returns:
            Path: 创建的文件路径
        """
        if not HAS_OPENPYXL:
            raise ImportError("Creating Excel requires openpyxl")

        output_path = Path(output_path)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Batch Config"

        # 表头
        headers = ["case_id", "CPU", "GPU", "Memory", "Ambient"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)

        # 示例数据
        data = [
            ["case1", 25.0, 15.0, 5.0, 25.0],
            ["case2", 30.0, 20.0, 6.0, 35.0],
            ["case3", 35.0, 25.0, 7.0, 45.0],
        ]

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        workbook.save(output_path)
        return output_path
