#!/usr/bin/env python3
"""
Calibration-based pack patcher for Simcenter Flotherm 2504 project packs.

This tool treats ``PDProject/group`` and, when available, ``PDProject/group.bak``
as opaque private-format blobs. It does not try to fully parse them. Instead, it:

1. Compares a baseline pack against a calibration pack where exactly one power
   value changed.
2. Locates byte offsets whose encoded numeric value changed from the baseline
   power to the calibrated power.
3. Stores those offsets and encodings as a calibration rule.
4. Applies the rule to new copies of the same pack family.

The v1 scope is intentionally narrow:
    - Simcenter Flotherm 2504 project packs
    - one model family per calibration
    - component power only
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import shutil
import struct
import sys
import tempfile
import zipfile
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

try:
    import openpyxl  # type: ignore

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


ENTRY_GROUP = "PDProject/group"
ENTRY_GROUP_BAK = "PDProject/group.bak"


@dataclass
class EncodingSpec:
    name: str
    baseline_bytes: bytes
    calibrated_bytes: bytes
    kind: str
    precision: int


@dataclass
class RuleOffset:
    offset: int
    encoding: str
    baseline_hex: str
    calibrated_hex: str
    calibration_diff_region: Tuple[int, int]


@dataclass
class EntryRule:
    entry_suffix: str
    size: int
    sha256: str
    offsets: List[RuleOffset]
    anchor_hits: List[int]
    diff_regions: List[Tuple[int, int]]


@dataclass
class CalibrationRule:
    tool: str
    version: int
    flotherm_version: str
    component_name: str
    baseline_power: float
    calibrated_power: float
    pack_family_prefix: str
    entries: List[EntryRule]


class CalibrationError(RuntimeError):
    """Raised when a pack cannot be calibrated safely."""


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def normalize_suffix(member_name: str) -> str:
    if "/" not in member_name:
        return member_name
    return member_name.split("/", 1)[1]


def find_member_name(names: Sequence[str], suffix: str) -> str:
    for name in names:
        if normalize_suffix(name) == suffix:
            return name
    raise CalibrationError(f"Missing required entry: {suffix}")


def read_pack_entries(pack_path: Path) -> Dict[str, bytes]:
    if not pack_path.exists():
        raise FileNotFoundError(f"Pack file not found: {pack_path}")
    if not zipfile.is_zipfile(pack_path):
        raise CalibrationError(f"Not a valid .pack zip archive: {pack_path}")

    with zipfile.ZipFile(pack_path, "r") as archive:
        names = archive.namelist()
        entries: Dict[str, bytes] = {}
        for suffix in (ENTRY_GROUP, ENTRY_GROUP_BAK):
            for name in names:
                if normalize_suffix(name) == suffix:
                    entries[suffix] = archive.read(name)
                    break
        entries["_root_prefix"] = find_member_name(names, ENTRY_GROUP).rsplit("/", 2)[0].encode("utf-8")
        return entries


def pack_member_map(pack_path: Path) -> Dict[str, str]:
    with zipfile.ZipFile(pack_path, "r") as archive:
        return {normalize_suffix(name): name for name in archive.namelist()}


def compute_diff_regions(left: bytes, right: bytes) -> List[Tuple[int, int]]:
    if len(left) != len(right):
        raise CalibrationError("Calibration requires identical entry sizes")

    regions: List[Tuple[int, int]] = []
    start: Optional[int] = None
    for idx, (a, b) in enumerate(zip(left, right)):
        if a != b and start is None:
            start = idx
        elif a == b and start is not None:
            regions.append((start, idx))
            start = None
    if start is not None:
        regions.append((start, len(left)))
    return regions


def float_variants(value_text: str, value: float) -> List[str]:
    variants = []
    for candidate in (
        value_text,
        str(value),
        f"{value:g}",
        f"{value:.1f}",
        f"{value:.2f}",
        f"{value:.3f}",
        f"{value:.6f}",
    ):
        if candidate not in variants:
            variants.append(candidate)
    return variants


def candidate_encodings(
    baseline_power: float,
    calibrated_power: float,
    baseline_text: str,
    calibrated_text: str,
) -> List[EncodingSpec]:
    specs: List[EncodingSpec] = []
    binary_specs = [
        ("float64_le", "<d", "binary", 64),
        ("float64_be", ">d", "binary", 64),
        ("float32_le", "<f", "binary", 32),
        ("float32_be", ">f", "binary", 32),
    ]

    for name, fmt, kind, precision in binary_specs:
        try:
            specs.append(
                EncodingSpec(
                    name=name,
                    baseline_bytes=struct.pack(fmt, baseline_power),
                    calibrated_bytes=struct.pack(fmt, calibrated_power),
                    kind=kind,
                    precision=precision,
                )
            )
        except struct.error:
            pass

    if baseline_power.is_integer() and calibrated_power.is_integer():
        int_specs = [
            ("int32_le", "<i"),
            ("int32_be", ">i"),
            ("int64_le", "<q"),
            ("int64_be", ">q"),
        ]
        for name, fmt in int_specs:
            try:
                specs.append(
                    EncodingSpec(
                        name=name,
                        baseline_bytes=struct.pack(fmt, int(baseline_power)),
                        calibrated_bytes=struct.pack(fmt, int(calibrated_power)),
                        kind="binary",
                        precision=struct.calcsize(fmt) * 8,
                    )
                )
            except struct.error:
                pass

    for base_variant in float_variants(baseline_text, baseline_power):
        for cal_variant in float_variants(calibrated_text, calibrated_power):
            if base_variant == cal_variant:
                continue
            specs.append(
                EncodingSpec(
                    name=f"ascii:{base_variant}->{cal_variant}",
                    baseline_bytes=base_variant.encode("ascii"),
                    calibrated_bytes=cal_variant.encode("ascii"),
                    kind="ascii",
                    precision=len(base_variant),
                )
            )

    return specs


def find_anchor_hits(blob: bytes, component_name: str) -> List[int]:
    needle = component_name.encode("utf-8", errors="ignore")
    if not needle:
        return []
    hits: List[int] = []
    start = 0
    while True:
        idx = blob.find(needle, start)
        if idx == -1:
            break
        hits.append(idx)
        start = idx + 1
    return hits


def region_for_offset(regions: Sequence[Tuple[int, int]], offset: int, size: int) -> Optional[Tuple[int, int]]:
    end = offset + size
    for region_start, region_end in regions:
        if offset < region_end and end > region_start:
            return (region_start, region_end)
    return None


def score_candidate(
    offset: int,
    size: int,
    region: Tuple[int, int],
    anchor_hits: Sequence[int],
    spec_name: str,
) -> Tuple[int, int, int, int]:
    anchor_score = 0
    if anchor_hits:
        anchor_distance = min(abs(offset - hit) for hit in anchor_hits)
        if anchor_distance <= 256:
            anchor_score = 3
        elif anchor_distance <= 1024:
            anchor_score = 2
        elif anchor_distance <= 4096:
            anchor_score = 1

    exact_region = 1 if (region[1] - region[0]) == size else 0
    binary_priority = 1 if spec_name.startswith("float64_le") else 0
    compactness = -(region[1] - region[0])
    return (anchor_score, exact_region, binary_priority, compactness)


def locate_offsets_for_entry(
    baseline_blob: bytes,
    calibrated_blob: bytes,
    component_name: str,
    baseline_power: float,
    calibrated_power: float,
    baseline_text: str,
    calibrated_text: str,
) -> EntryRule:
    diff_regions = compute_diff_regions(baseline_blob, calibrated_blob)
    if not diff_regions:
        raise CalibrationError("Calibration pack has no changes in target entry")

    anchor_hits = find_anchor_hits(baseline_blob, component_name)
    specs = candidate_encodings(baseline_power, calibrated_power, baseline_text, calibrated_text)

    matches_by_spec: Dict[str, List[RuleOffset]] = {}
    scores_by_spec: Dict[str, Tuple[int, int, int, int]] = {}

    for spec in specs:
        hits: List[RuleOffset] = []
        start = 0
        while True:
            idx = baseline_blob.find(spec.baseline_bytes, start)
            if idx == -1:
                break
            region = region_for_offset(diff_regions, idx, len(spec.baseline_bytes))
            if region and calibrated_blob[idx : idx + len(spec.calibrated_bytes)] == spec.calibrated_bytes:
                hits.append(
                    RuleOffset(
                        offset=idx,
                        encoding=spec.name,
                        baseline_hex=spec.baseline_bytes.hex(),
                        calibrated_hex=spec.calibrated_bytes.hex(),
                        calibration_diff_region=region,
                    )
                )
            start = idx + 1

        if hits:
            matches_by_spec[spec.name] = hits
            spec_scores = [
                score_candidate(hit.offset, len(spec.baseline_bytes), hit.calibration_diff_region, anchor_hits, spec.name)
                for hit in hits
            ]
            scores_by_spec[spec.name] = max(spec_scores)

    if not matches_by_spec:
        raise CalibrationError(
            "Failed to locate any byte offset that changed from the baseline power to the calibrated power"
        )

    ranked_specs = sorted(
        matches_by_spec.keys(),
        key=lambda name: (scores_by_spec[name], -len(matches_by_spec[name]), name),
        reverse=True,
    )
    winner = ranked_specs[0]

    if len(ranked_specs) > 1 and scores_by_spec[ranked_specs[0]] == scores_by_spec[ranked_specs[1]]:
        raise CalibrationError(
            "Calibration is ambiguous; multiple encodings/offset groups match equally well. "
            "Use a cleaner calibration sample or a more distinctive component anchor."
        )

    return EntryRule(
        entry_suffix="",
        size=len(baseline_blob),
        sha256=sha256_bytes(baseline_blob),
        offsets=matches_by_spec[winner],
        anchor_hits=anchor_hits,
        diff_regions=diff_regions,
    )


def extract_flotherm_version(group_blob: bytes) -> str:
    first_line = group_blob.splitlines()[0].decode("latin-1", errors="ignore").strip()
    return first_line or "unknown"


def calibrate_pack(
    baseline_pack: Path,
    calibrated_pack: Path,
    component_name: str,
    baseline_power_text: str,
    calibrated_power_text: str,
) -> CalibrationRule:
    baseline_power = float(baseline_power_text)
    calibrated_power = float(calibrated_power_text)

    base_entries = read_pack_entries(baseline_pack)
    cal_entries = read_pack_entries(calibrated_pack)

    if base_entries["_root_prefix"] != cal_entries["_root_prefix"]:
        raise CalibrationError("Baseline and calibrated packs do not appear to belong to the same project family")

    entries: List[EntryRule] = []
    flotherm_version = extract_flotherm_version(base_entries[ENTRY_GROUP])

    for suffix in (ENTRY_GROUP, ENTRY_GROUP_BAK):
        base_blob = base_entries.get(suffix)
        cal_blob = cal_entries.get(suffix)
        if base_blob is None or cal_blob is None:
            continue

        entry_rule = locate_offsets_for_entry(
            baseline_blob=base_blob,
            calibrated_blob=cal_blob,
            component_name=component_name,
            baseline_power=float(baseline_power),
            calibrated_power=float(calibrated_power),
            baseline_text=baseline_power_text,
            calibrated_text=calibrated_power_text,
        )
        entry_rule.entry_suffix = suffix
        entries.append(entry_rule)

    if not entries:
        raise CalibrationError("Neither PDProject/group nor PDProject/group.bak could be calibrated")

    return CalibrationRule(
        tool="pack_group_power_tool",
        version=1,
        flotherm_version=flotherm_version,
        component_name=component_name,
        baseline_power=float(baseline_power),
        calibrated_power=float(calibrated_power),
        pack_family_prefix=base_entries["_root_prefix"].decode("utf-8"),
        entries=entries,
    )


def encoding_bytes(encoding_name: str, value: float) -> bytes:
    if encoding_name == "float64_le":
        return struct.pack("<d", value)
    if encoding_name == "float64_be":
        return struct.pack(">d", value)
    if encoding_name == "float32_le":
        return struct.pack("<f", value)
    if encoding_name == "float32_be":
        return struct.pack(">f", value)
    if encoding_name == "int32_le":
        return struct.pack("<i", int(value))
    if encoding_name == "int32_be":
        return struct.pack(">i", int(value))
    if encoding_name == "int64_le":
        return struct.pack("<q", int(value))
    if encoding_name == "int64_be":
        return struct.pack(">q", int(value))
    if encoding_name.startswith("ascii:"):
        _, mapping = encoding_name.split(":", 1)
        _, calibrated = mapping.split("->", 1)
        return calibrated.encode("ascii")
    raise CalibrationError(f"Unsupported encoding in rule: {encoding_name}")


def expected_baseline_bytes(rule_offset: RuleOffset) -> bytes:
    return bytes.fromhex(rule_offset.baseline_hex)


def apply_rule_to_blob(blob: bytes, entry_rule: EntryRule, new_power: float) -> bytes:
    patched = bytearray(blob)
    if sha256_bytes(blob) != entry_rule.sha256:
        raise CalibrationError(
            f"Entry {entry_rule.entry_suffix} does not match the calibrated baseline. "
            "Apply this rule only to copies of the original baseline pack."
        )

    for offset_rule in entry_rule.offsets:
        old_bytes = expected_baseline_bytes(offset_rule)
        new_bytes = encoding_bytes(offset_rule.encoding, new_power)
        current = bytes(patched[offset_rule.offset : offset_rule.offset + len(old_bytes)])
        if current != old_bytes:
            raise CalibrationError(
                f"Entry {entry_rule.entry_suffix} offset {offset_rule.offset} no longer matches the baseline bytes"
            )
        if len(new_bytes) != len(old_bytes):
            raise CalibrationError(
                f"Encoding {offset_rule.encoding} changed width for value {new_power}; refusing to corrupt the blob"
            )
        patched[offset_rule.offset : offset_rule.offset + len(new_bytes)] = new_bytes

    return bytes(patched)


def load_rule(rule_path: Path) -> CalibrationRule:
    payload = json.loads(rule_path.read_text(encoding="utf-8"))
    return CalibrationRule(
        tool=payload["tool"],
        version=payload["version"],
        flotherm_version=payload["flotherm_version"],
        component_name=payload["component_name"],
        baseline_power=float(payload["baseline_power"]),
        calibrated_power=float(payload["calibrated_power"]),
        pack_family_prefix=payload["pack_family_prefix"],
        entries=[
            EntryRule(
                entry_suffix=entry["entry_suffix"],
                size=entry["size"],
                sha256=entry["sha256"],
                offsets=[
                    RuleOffset(
                        offset=offset["offset"],
                        encoding=offset["encoding"],
                        baseline_hex=offset["baseline_hex"],
                        calibrated_hex=offset["calibrated_hex"],
                        calibration_diff_region=tuple(offset["calibration_diff_region"]),
                    )
                    for offset in entry["offsets"]
                ],
                anchor_hits=entry["anchor_hits"],
                diff_regions=[tuple(region) for region in entry["diff_regions"]],
            )
            for entry in payload["entries"]
        ],
    )


def save_rule(rule: CalibrationRule, output_path: Path) -> None:
    output_path.write_text(json.dumps(asdict(rule), indent=2), encoding="utf-8")


def patch_pack(input_pack: Path, rule: CalibrationRule, new_power: float, output_pack: Path) -> None:
    member_map = pack_member_map(input_pack)
    missing = [entry.entry_suffix for entry in rule.entries if entry.entry_suffix not in member_map]
    if missing:
        raise CalibrationError(f"Input pack is missing expected entries: {', '.join(missing)}")
    if input_pack.resolve() == output_pack.resolve():
        raise CalibrationError("Refusing to overwrite the input pack in place; choose a new output path")

    output_pack.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(input_pack, "r") as source, zipfile.ZipFile(output_pack, "w") as target:
        for info in source.infolist():
            data = source.read(info.filename)
            suffix = normalize_suffix(info.filename)
            entry_rule = next((entry for entry in rule.entries if entry.entry_suffix == suffix), None)
            if entry_rule:
                data = apply_rule_to_blob(data, entry_rule, new_power)
            cloned = zipfile.ZipInfo(info.filename, date_time=info.date_time)
            cloned.compress_type = info.compress_type
            cloned.comment = info.comment
            cloned.extra = info.extra
            cloned.internal_attr = info.internal_attr
            cloned.external_attr = info.external_attr
            cloned.create_system = info.create_system
            cloned.flag_bits = info.flag_bits
            target.writestr(cloned, data)


def load_cases_table(path: Path) -> List[Dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
    elif suffix in {".xlsx", ".xlsm"}:
        if not HAS_OPENPYXL:
            raise CalibrationError("Reading Excel requires openpyxl. Install it or use CSV.")
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(sheet.iter_rows(max_row=1))]
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(value is not None for value in row):
                continue
            rows.append(
                {
                    headers[idx]: "" if idx >= len(row) or row[idx] is None else str(row[idx]).strip()
                    for idx in range(len(headers))
                    if headers[idx]
                }
            )
        workbook.close()
    else:
        raise CalibrationError("Cases file must be .csv, .xlsx, or .xlsm")

    required = {"case_id", "component_name", "power"}
    if not rows:
        raise CalibrationError("Cases file is empty")
    missing = required - set(rows[0].keys())
    if missing:
        raise CalibrationError(f"Cases file is missing required columns: {', '.join(sorted(missing))}")
    return rows


def write_manifest(path: Path, manifest: Dict[str, object]) -> None:
    path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")


def run_batch(input_pack: Path, rule: CalibrationRule, cases_path: Path, output_dir: Path) -> Dict[str, object]:
    cases = load_cases_table(cases_path)
    output_dir.mkdir(parents=True, exist_ok=True)

    manifest_cases: List[Dict[str, object]] = []
    for row in cases:
        case_id = row["case_id"]
        component_name = row["component_name"]
        power = float(row["power"])

        if component_name != rule.component_name:
            raise CalibrationError(
                f"Case '{case_id}' references component '{component_name}', "
                f"but this rule only supports '{rule.component_name}'"
            )

        case_dir = output_dir / case_id
        case_dir.mkdir(parents=True, exist_ok=True)
        output_pack = case_dir / f"{case_id}.pack"
        patch_pack(input_pack=input_pack, rule=rule, new_power=power, output_pack=output_pack)

        manifest_cases.append(
            {
                "case_id": case_id,
                "component_name": component_name,
                "power": power,
                "output_pack": str(output_pack),
            }
        )

    manifest = {
        "input_pack": str(input_pack),
        "rule_component": rule.component_name,
        "rule_flotherm_version": rule.flotherm_version,
        "cases": manifest_cases,
    }
    write_manifest(output_dir / "run_manifest.json", manifest)
    return manifest


def inspect_pack(pack_path: Path, include_strings: bool = False, max_strings: int = 60) -> Dict[str, object]:
    member_map = pack_member_map(pack_path)
    result: Dict[str, object] = {
        "pack": str(pack_path),
        "entries": sorted(name for name in member_map.keys() if name),
    }

    with zipfile.ZipFile(pack_path, "r") as archive:
        if ENTRY_GROUP in member_map:
            group_blob = archive.read(member_map[ENTRY_GROUP])
            result["flotherm_version"] = extract_flotherm_version(group_blob)
            if include_strings:
                strings: List[str] = []
                current: List[int] = []
                for byte in group_blob:
                    if 32 <= byte < 127:
                        current.append(byte)
                    else:
                        if len(current) >= 4:
                            strings.append(bytes(current).decode("ascii"))
                        current = []
                if len(current) >= 4:
                    strings.append(bytes(current).decode("ascii"))
                result["group_strings"] = strings[:max_strings]
    return result


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Calibrate and patch Flotherm 2504 project packs")
    subparsers = parser.add_subparsers(dest="command", required=True)

    calibrate_parser = subparsers.add_parser("calibrate", help="Create a calibration rule from two packs")
    calibrate_parser.add_argument("baseline_pack", type=Path)
    calibrate_parser.add_argument("calibrated_pack", type=Path)
    calibrate_parser.add_argument("--component-name", required=True)
    calibrate_parser.add_argument("--baseline-power", required=True)
    calibrate_parser.add_argument("--calibrated-power", required=True)
    calibrate_parser.add_argument("-o", "--output", type=Path, required=True)

    patch_parser = subparsers.add_parser("patch", help="Apply a saved calibration rule to a pack")
    patch_parser.add_argument("input_pack", type=Path)
    patch_parser.add_argument("--rule", type=Path, required=True)
    patch_parser.add_argument("--component-name", required=False)
    patch_parser.add_argument("--new-power", required=True, type=float)
    patch_parser.add_argument("-o", "--output", type=Path, required=True)

    batch_parser = subparsers.add_parser("batch", help="Generate multiple patched packs from a cases table")
    batch_parser.add_argument("input_pack", type=Path)
    batch_parser.add_argument("--rule", type=Path, required=True)
    batch_parser.add_argument("--cases", type=Path, required=True)
    batch_parser.add_argument("-o", "--output-dir", type=Path, required=True)

    inspect_parser = subparsers.add_parser("inspect", help="Inspect pack entries and visible group strings")
    inspect_parser.add_argument("input_pack", type=Path)
    inspect_parser.add_argument("--strings", action="store_true", help="Include printable strings from PDProject/group")
    inspect_parser.add_argument("--max-strings", type=int, default=60)

    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    try:
        if args.command == "calibrate":
            rule = calibrate_pack(
                baseline_pack=args.baseline_pack,
                calibrated_pack=args.calibrated_pack,
                component_name=args.component_name,
                baseline_power_text=args.baseline_power,
                calibrated_power_text=args.calibrated_power,
            )
            save_rule(rule, args.output)
            print(f"[OK] Wrote calibration rule: {args.output}")
            print(f"[OK] Component: {rule.component_name}")
            for entry in rule.entries:
                offsets = ", ".join(f"{offset.offset}:{offset.encoding}" for offset in entry.offsets)
                print(f"[OK] {entry.entry_suffix} -> {offsets}")
            return 0

        if args.command == "patch":
            rule = load_rule(args.rule)
            if args.component_name and args.component_name != rule.component_name:
                raise CalibrationError(
                    f"Rule component mismatch: expected '{rule.component_name}', got '{args.component_name}'"
                )
            patch_pack(args.input_pack, rule, args.new_power, args.output)
            print(f"[OK] Wrote patched pack: {args.output}")
            return 0

        if args.command == "batch":
            rule = load_rule(args.rule)
            manifest = run_batch(args.input_pack, rule, args.cases, args.output_dir)
            print(f"[OK] Generated {len(manifest['cases'])} case(s) in {args.output_dir}")
            return 0

        if args.command == "inspect":
            info = inspect_pack(args.input_pack, include_strings=args.strings, max_strings=args.max_strings)
            print(json.dumps(info, indent=2))
            return 0

        parser.error("Unknown command")
        return 2
    except (CalibrationError, FileNotFoundError, ValueError) as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
