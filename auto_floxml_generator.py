#!/usr/bin/env python3
"""
FloXML 全自动生成器 - 最终版

方案：
1. 创建 Excel 文件并写入数据（openpyxl）
2. 预先准备好带 Auto_Open 宏的模板（首次需要手动）
3. 用 Windows start 命令在 GUI 环境启动 Excel
4. Auto_Open 自动运行并生成 FloXML

首次使用需要：
1. 打开 floxml_generator.xlsm
2. 按 Alt+F11 添加 Auto_Open 宏
3. 保存

之后就可以全自动运行了！
"""

import os
import sys
import json
import shutil
import subprocess
import time
from pathlib import Path
from typing import List, Dict

# 路径配置
TEMPLATES_DIR = Path(r"D:\Program Files\Siemens\SimcenterFlotherm\2504\examples\FloXML\Spreadsheets")
OUTPUT_DIR = Path(r"D:\Program Files\Siemens\SimcenterFlotherm\2504\flotherm-automation\floxml_output")
TEMPLATE_FILE = "Materials.xlsm"

# Auto_Open VBA 代码
AUTO_OPEN_VBA = '''
Sub Auto_Open()
    On Error Resume Next
    CREATEMATERIALS
    If Err.Number <> 0 Then
        Err.Clear
        CREATEMODEL
    End If
    ThisWorkbook.Save
    Application.Quit
End Sub
'''


def create_excel_with_data(template_path: str, output_excel: str,
                           materials: List[Dict], floxml_output: str) -> bool:
    """创建带有数据的 Excel 文件"""
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] 请安装 openpyxl: pip install openpyxl")
        return False

    # 复制模板
    shutil.copy(template_path, output_excel)

    # 修改数据
    wb = openpyxl.load_workbook(output_excel, keep_vba=True)
    ws = wb.active

    # 设置输出路径 (Row 1, Column 2)
    ws.cell(row=1, column=2, value=floxml_output)

    # 清除旧数据 (Row 4-100)
    for row in range(4, 101):
        for col in range(1, 9):
            ws.cell(row=row, column=col, value=None)

    # 写入新数据
    for idx, mat in enumerate(materials):
        row = idx + 4
        ws.cell(row=row, column=1, value=mat.get("name"))
        ws.cell(row=row, column=2, value=mat.get("type", "Isotropic"))

        if mat.get("type") == "Orthotropic":
            ws.cell(row=row, column=3, value=mat.get("kx"))
            ws.cell(row=row, column=4, value=mat.get("ky"))
            ws.cell(row=row, column=5, value=mat.get("kz"))
        else:
            ws.cell(row=row, column=3, value=mat.get("kx"))

    wb.save(output_excel)
    wb.close()

    return True


def check_vba_module_exists(excel_path: str) -> bool:
    """检查 Excel 中是否有 Auto_Open 宏"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, keep_vba=True)
        # openpyxl 不能直接读取 VBA，但可以检查 vbaProject.bin
        has_vba = 'xl/vbaProject.bin' in wb.archive.namelist()
        wb.close()
        return has_vba
    except:
        return False


def run_excel_gui(excel_path: str, wait_time: int = 30) -> bool:
    """
    在 GUI 环境启动 Excel
    使用 Windows start 命令
    """
    abs_path = os.path.abspath(excel_path)

    print(f"[INFO] 启动 Excel: {abs_path}")
    print(f"[INFO] 等待 {wait_time} 秒让宏完成...")

    try:
        # 使用 start 命令在 GUI 环境启动 Excel
        subprocess.Popen(
            ['start', 'excel', abs_path],
            shell=True
        )

        # 等待宏完成
        time.sleep(wait_time)

        return True

    except Exception as e:
        print(f"[ERROR] {e}")
        return False


def generate_floxml(materials: List[Dict], output_floxml: str) -> Dict:
    """
    生成 FloXML

    Returns:
        {
            "success": bool,
            "excel_path": str,
            "floxml_path": str,
            "instructions": str (if manual step needed)
        }
    """
    OUTPUT_DIR.mkdir(exist_ok=True)

    template_path = TEMPLATES_DIR / TEMPLATE_FILE
    output_excel = OUTPUT_DIR / "floxml_generator.xlsm"

    result = {
        "success": False,
        "excel_path": str(output_excel),
        "floxml_path": output_floxml
    }

    print("=" * 50)
    print("FloXML 全自动生成器")
    print("=" * 50)
    print()

    # Step 1: 创建 Excel 文件
    print("[Step 1/3] 创建 Excel 文件并写入数据...")
    if not create_excel_with_data(str(template_path), str(output_excel), materials, output_floxml):
        result["instructions"] = "创建 Excel 文件失败"
        return result

    print(f"  Excel 文件: {output_excel}")
    print(f"  输出路径: {output_floxml}")

    # Step 2: 检查是否有 Auto_Open
    print()
    print("[Step 2/3] 检查 Auto_Open 宏...")

    # 提示用户添加宏（如果还没添加）
    print()
    print("请在 Excel 中添加 Auto_Open 宏（如果还没添加）:")
    print("  1. 打开: " + str(output_excel))
    print("  2. 按 Alt+F11 打开 VBA 编辑器")
    print("  3. Insert → Module")
    print("  4. 粘贴以下代码:")
    print()
    print(AUTO_OPEN_VBA)
    print()
    print("  5. 保存并关闭 Excel")
    print()

    # Step 3: 启动 Excel
    print("[Step 3/3] 启动 Excel...")
    run_excel_gui(str(output_excel), wait_time=10)

    # 检查输出
    print()
    print("检查输出文件...")
    if os.path.exists(output_floxml):
        print(f"[OK] FloXML 已生成: {output_floxml}")
        result["success"] = True
    else:
        print(f"[INFO] 输出文件: {output_floxml}")
        print("如果文件未生成，请确保 Auto_Open 宏已正确添加")

    return result


def main():
    # 示例材料数据
    materials = [
        {"name": "Copper", "type": "Isotropic", "kx": 385},
        {"name": "FR4", "type": "Isotropic", "kx": 0.3},
        {"name": "Aluminum", "type": "Isotropic", "kx": 180},
    ]

    output_floxml = r"c:\temp\materials_output.xml"

    result = generate_floxml(materials, output_floxml)

    print()
    print("=" * 50)
    if result["success"]:
        print("[成功] FloXML 生成完成!")
        print(f"输出文件: {result['floxml_path']}")
    else:
        print("[需要手动操作]")
        print("请按照上面的步骤添加 Auto_Open 宏")
    print("=" * 50)


if __name__ == '__main__':
    main()
