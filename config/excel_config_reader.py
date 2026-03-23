"""
Excel 配置读取器

读取 Excel 文件中的仿真参数配置。

配置格式：
- 第一行为表头，必须包含：component, parameter, value
- 后续行为数据行

示例：
| component | parameter | value |
|-----------|-----------|-------|
| U1_CPU    | power     | 15.0  |
| U1_CPU    | x_size    | 0.02  |
| U2_GPU    | power     | 10.0  |
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Dict, List, Any, Set

from openpyxl import load_workbook, Workbook


class ExcelConfigReaderError(Exception):
    """Excel 配置读取器错误基类"""
    pass


class MissingColumnError(ExcelConfigReaderError):
    """缺少必要列错误"""
    pass


class ExcelConfigReader:
    """
    Excel 配置读取器

    读取 Excel 文件中的仿真参数配置。

    Usage:
        reader = ExcelConfigReader("config.xlsx")
        config = reader.read()
        # [{"component": "U1_CPU", "parameter": "power", "value": 15.0}, ...]

        # 获取组件列表
        components = reader.get_components()

        # 按组件筛选
        u1_config = reader.get_by_component("U1_CPU")

        # 转换为字典
        config_dict = reader.to_dict()
        # {"U1_CPU": {"power": 15.0, "x_size": 0.02}, ...}
    """

    # 必要的列
    REQUIRED_COLUMNS = {"component", "parameter", "value"}

    def __init__(self, file_path: str | Path):
        """
        初始化配置读取器

        Args:
            file_path: Excel 文件路径
        """
        self.file_path = Path(file_path)
        self._data: List[Dict[str, Any]] = []
        self._loaded = False

    def _validate_columns(self, headers: List[str]) -> None:
        """验证表头包含必要列"""
        header_set = {h.lower().strip() for h in headers}
        missing = self.REQUIRED_COLUMNS - header_set
        if missing:
            raise MissingColumnError(
                f"Missing required columns: {missing}. "
                f"Required: {self.REQUIRED_COLUMNS}, Found: {header_set}"
            )

    def _parse_value(self, value: Any) -> Any:
        """解析单元格值"""
        if value is None:
            return None
        # 保持原始类型（int, float, str）
        return value

    def read(self) -> List[Dict[str, Any]]:
        """
        读取 Excel 配置文件

        Returns:
            配置列表，每项包含 component, parameter, value

        Raises:
            FileNotFoundError: 文件不存在
            MissingColumnError: 缺少必要列
        """
        if not self.file_path.exists():
            raise FileNotFoundError(f"Config file not found: {self.file_path}")

        # 加载工作簿（不使用 read_only 以避免 Windows 文件锁定问题）
        try:
            wb: Workbook = load_workbook(self.file_path, data_only=True)
        except Exception as e:
            raise ExcelConfigReaderError(f"Failed to load Excel file: {e}")

        try:
            ws = wb.active

            # 读取所有行
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                self._data = []
                self._loaded = True
                return self._data

            # 第一行是表头
            headers = [str(h).lower().strip() if h else "" for h in rows[0]]
            self._validate_columns(headers)

            # 找到必要列的索引
            component_idx = headers.index("component")
            parameter_idx = headers.index("parameter")
            value_idx = headers.index("value")

            # 读取数据行
            self._data = []
            for row in rows[1:]:
                if len(row) <= max(component_idx, parameter_idx, value_idx):
                    continue

                component = row[component_idx]
                parameter = row[parameter_idx]
                value = self._parse_value(row[value_idx])

                # 跳过空行
                if component is None and parameter is None:
                    continue

                self._data.append({
                    "component": component,
                    "parameter": parameter,
                    "value": value
                })

            self._loaded = True
            return self._data
        finally:
            wb.close()

    def get_components(self) -> Set[str]:
        """
        获取所有组件名称

        Returns:
            组件名称集合
        """
        if not self._loaded:
            self.read()

        return {item["component"] for item in self._data if item["component"]}

    def get_by_component(self, component_name: str) -> List[Dict[str, Any]]:
        """
        获取指定组件的配置

        Args:
            component_name: 组件名称

        Returns:
            该组件的配置列表
        """
        if not self._loaded:
            self.read()

        return [
            item for item in self._data
            if item["component"] == component_name
        ]

    def to_dict(self) -> Dict[str, Dict[str, Any]]:
        """
        转换为字典格式

        Returns:
            {component: {parameter: value, ...}, ...}
        """
        if not self._loaded:
            self.read()

        result: Dict[str, Dict[str, Any]] = {}

        for item in self._data:
            component = item["component"]
            parameter = item["parameter"]
            value = item["value"]

            if component not in result:
                result[component] = {}

            result[component][parameter] = value

        return result

    def __repr__(self) -> str:
        return f"ExcelConfigReader({self.file_path})"
