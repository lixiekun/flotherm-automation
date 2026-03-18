#!/usr/bin/env python3
"""
FloXML generator driven by the official Excel templates.

This script automates the workflow that has been validated locally:
1. Copy an official `.xlsm` template or use an existing workbook.
2. Fill the worksheet with input data when using the `materials` template.
3. Detect the workbook button macro from the `.xlsm` package.
4. Launch Excel in an isolated COM instance and run the macro.
5. Wait for the target FloXML file to appear, then clean up the Excel instance.

The helper process pattern is used because some templates generate the XML
successfully but leave the Excel COM client hanging during shutdown.
"""

from __future__ import annotations

import argparse
import ctypes
import json
import shutil
import subprocess
import sys
import tempfile
import time
import zipfile
from pathlib import Path
from typing import Dict, Iterable, List, Optional
from xml.etree import ElementTree as ET


TEMPLATES_DIR = Path(
    r"D:\Program Files\Siemens\SimcenterFlotherm\2504\examples\FloXML\Spreadsheets"
)
OUTPUT_DIR = Path(
    r"D:\Program Files\Siemens\SimcenterFlotherm\2504\flotherm-automation\floxml_output"
)

TEMPLATE_FILES = {
    "materials": "Materials.xlsm",
    "advanced_resistance": "Advanced-Resistance.xlsm",
    "data_center_si": "Data_Center_SI_Units.xlsm",
    "data_center_us": "Data_Center_US_Units.xlsm",
    "detailed_substrate": "Detailed-Substrate.xlsm",
    "heatpipe": "Heatpipe-LShaped.xlsm",
    "igbt": "IGBT-Creator.xlsm",
    "package_on_package": "Package-on-Package.xlsm",
    "folded_fin": "Solve-Folded_Fin_Heat_Sink.xlsm",
    "windtunnel": "Windtunnel-AdvancedResistance.xlsm",
}

FALLBACK_MACROS = [
    "create_all_materials",
    "CREATEMODEL",
]


def _write_status(status_file: Path, payload: Dict[str, object]) -> None:
    status_file.write_text(json.dumps(payload, ensure_ascii=True), encoding="utf-8")


def _read_status(status_file: Path) -> Dict[str, object]:
    if not status_file.exists():
        return {}
    try:
        return json.loads(status_file.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def _get_pid_from_hwnd(hwnd: int) -> int:
    pid = ctypes.c_ulong()
    ctypes.windll.user32.GetWindowThreadProcessId(hwnd, ctypes.byref(pid))
    return int(pid.value)


def _normalize_macro_name(raw_name: str) -> str:
    name = raw_name.strip().strip("'").strip('"')
    if "!" in name:
        name = name.split("!")[-1]
    return name.strip()


def detect_macro_candidates(workbook_path: Path) -> List[str]:
    candidates: List[str] = []

    with zipfile.ZipFile(workbook_path) as archive:
        drawing_names = [
            name for name in archive.namelist() if name.startswith("xl/drawings/") and name.endswith(".xml")
        ]

        for drawing_name in drawing_names:
            root = ET.fromstring(archive.read(drawing_name))
            for element in root.iter():
                macro = element.attrib.get("macro")
                if not macro:
                    continue
                normalized = _normalize_macro_name(macro)
                if normalized and normalized not in candidates:
                    candidates.append(normalized)

    for macro in FALLBACK_MACROS:
        if macro not in candidates:
            candidates.append(macro)

    return candidates


def update_materials_workbook(workbook_path: Path, output_path: Path, materials: List[Dict[str, object]]) -> None:
    import openpyxl

    workbook = openpyxl.load_workbook(workbook_path, keep_vba=True)
    worksheet = workbook.active

    worksheet.cell(row=1, column=2, value=str(output_path))

    for row in range(4, 201):
        for col in range(1, 9):
            worksheet.cell(row=row, column=col, value=None)

    for idx, material in enumerate(materials, start=4):
        worksheet.cell(row=idx, column=1, value=material.get("name"))
        worksheet.cell(row=idx, column=2, value=material.get("type", "Isotropic"))

        material_type = str(material.get("type", "Isotropic"))
        if material_type == "Orthotropic":
            worksheet.cell(row=idx, column=3, value=material.get("kx"))
            worksheet.cell(row=idx, column=4, value=material.get("ky"))
            worksheet.cell(row=idx, column=5, value=material.get("kz"))
        elif material_type == "Temperature Dependant":
            worksheet.cell(row=idx, column=6, value=material.get("conductivity"))
            worksheet.cell(row=idx, column=7, value=material.get("coefficient"))
            worksheet.cell(row=idx, column=8, value=material.get("ref_temp"))
        else:
            worksheet.cell(row=idx, column=3, value=material.get("kx"))

    workbook.save(workbook_path)
    workbook.close()


def _terminate_pid(pid: Optional[int]) -> None:
    if not pid:
        return
    subprocess.run(
        ["taskkill", "/PID", str(pid), "/T", "/F"],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        check=False,
    )


def _run_macro_helper(workbook_path: Path, macro_name: str, status_file: Path) -> int:
    import win32com.client

    excel = None
    workbook = None
    excel_pid: Optional[int] = None

    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            excel.AutomationSecurity = 1
        except Exception:
            pass

        try:
            excel_pid = _get_pid_from_hwnd(int(excel.Hwnd))
        except Exception:
            excel_pid = None

        _write_status(
            status_file,
            {
                "stage": "excel_started",
                "excel_pid": excel_pid,
                "macro_name": macro_name,
                "workbook_path": str(workbook_path),
            },
        )

        workbook = excel.Workbooks.Open(str(workbook_path))
        _write_status(
            status_file,
            {
                "stage": "workbook_opened",
                "excel_pid": excel_pid,
                "macro_name": macro_name,
                "workbook_path": str(workbook_path),
            },
        )

        excel.Run(f"'{workbook_path.name}'!{macro_name}")
        _write_status(
            status_file,
            {
                "stage": "macro_returned",
                "excel_pid": excel_pid,
                "macro_name": macro_name,
                "workbook_path": str(workbook_path),
            },
        )

        try:
            workbook.Close(False)
        except Exception:
            pass

        try:
            excel.Quit()
        except Exception:
            pass

        _write_status(
            status_file,
            {
                "stage": "finished",
                "excel_pid": excel_pid,
                "macro_name": macro_name,
                "workbook_path": str(workbook_path),
            },
        )
        return 0
    except Exception as exc:
        _write_status(
            status_file,
            {
                "stage": "error",
                "excel_pid": excel_pid,
                "macro_name": macro_name,
                "workbook_path": str(workbook_path),
                "error": str(exc),
            },
        )
        return 1


class ExcelFloXMLGenerator:
    def __init__(self, template_type: str = "materials"):
        if template_type not in TEMPLATE_FILES:
            raise ValueError(f"Unknown template: {template_type}")

        self.template_type = template_type
        self.template_path = TEMPLATES_DIR / TEMPLATE_FILES[template_type]
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template not found: {self.template_path}")

        OUTPUT_DIR.mkdir(exist_ok=True)

    def prepare_workbook(
        self,
        output_floxml: Path,
        materials: Optional[List[Dict[str, object]]] = None,
        workbook_path: Optional[Path] = None,
    ) -> Path:
        if workbook_path is None:
            workbook_path = OUTPUT_DIR / f"{self.template_type}_auto.xlsm"
            shutil.copy2(self.template_path, workbook_path)

        if self.template_type == "materials":
            if materials is None:
                raise ValueError("materials template requires material input data")
            update_materials_workbook(workbook_path, output_floxml, materials)

        return workbook_path

    def run_macro(
        self,
        workbook_path: Path,
        macro_name: str,
        output_floxml: Path,
        timeout_seconds: int = 60,
    ) -> Dict[str, object]:
        status_dir = Path(tempfile.mkdtemp(prefix="floxml_excel_", dir=str(OUTPUT_DIR)))
        status_file = status_dir / "status.json"

        creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
        process = subprocess.Popen(
            [
                sys.executable,
                str(Path(__file__).resolve()),
                "--helper-run-macro",
                str(workbook_path),
                macro_name,
                str(status_file),
            ],
            creationflags=creationflags,
        )

        excel_pid: Optional[int] = None
        started = time.time()
        success = False
        failure_reason: Optional[str] = None

        try:
            while time.time() - started <= timeout_seconds:
                if output_floxml.exists() and output_floxml.stat().st_size > 0:
                    success = True
                    break

                status = _read_status(status_file)
                excel_pid = int(status["excel_pid"]) if status.get("excel_pid") else excel_pid
                if status.get("stage") == "error":
                    failure_reason = str(status.get("error", "Excel macro execution failed"))
                    break

                exit_code = process.poll()
                if exit_code is not None and exit_code != 0:
                    failure_reason = f"helper process exited with code {exit_code}"
                    break

                time.sleep(0.5)

            if not success and failure_reason is None:
                failure_reason = f"timed out after {timeout_seconds}s waiting for FloXML output"
        finally:
            if process.poll() is None:
                process.terminate()
                try:
                    process.wait(timeout=5)
                except subprocess.TimeoutExpired:
                    process.kill()
                    process.wait(timeout=5)

            _terminate_pid(excel_pid)

            try:
                if status_file.exists():
                    status_file.unlink()
            except OSError:
                pass

            try:
                status_dir.rmdir()
            except OSError:
                pass

        return {
            "success": success,
            "macro_name": macro_name,
            "output_floxml": str(output_floxml),
            "workbook_path": str(workbook_path),
            "failure_reason": failure_reason,
        }

    def generate(
        self,
        output_floxml: Path,
        materials: Optional[List[Dict[str, object]]] = None,
        workbook_path: Optional[Path] = None,
        macro_name: Optional[str] = None,
        timeout_seconds: int = 60,
    ) -> Dict[str, object]:
        prepared_workbook = self.prepare_workbook(
            output_floxml=output_floxml,
            materials=materials,
            workbook_path=workbook_path,
        )

        candidates = [macro_name] if macro_name else detect_macro_candidates(prepared_workbook)
        last_result: Optional[Dict[str, object]] = None

        for candidate in candidates:
            if not candidate:
                continue
            result = self.run_macro(
                workbook_path=prepared_workbook,
                macro_name=candidate,
                output_floxml=output_floxml,
                timeout_seconds=timeout_seconds,
            )
            last_result = result
            if result["success"]:
                result["macro_candidates"] = candidates
                return result

        if last_result is None:
            raise RuntimeError("no macro candidates were available")

        last_result["macro_candidates"] = candidates
        return last_result


def load_materials(data_path: Optional[Path]) -> List[Dict[str, object]]:
    if data_path is None:
        return [
            {"name": "Copper", "type": "Isotropic", "kx": 385},
            {"name": "FR4", "type": "Isotropic", "kx": 0.3},
            {"name": "Aluminum", "type": "Isotropic", "kx": 180},
        ]

    payload = json.loads(data_path.read_text(encoding="utf-8"))
    materials = payload.get("materials")
    if not isinstance(materials, list):
        raise ValueError("JSON must contain a top-level 'materials' array")
    return materials


def list_templates() -> None:
    print("\nAvailable FloXML templates:")
    print("-" * 40)
    for key, filename in TEMPLATE_FILES.items():
        template_path = TEMPLATES_DIR / filename
        status = "OK" if template_path.exists() else "MISSING"
        print(f"  {status:8} {key:20} {filename}")
    print()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Generate FloXML via the official Excel templates",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python excel_floxml_generator.py --list
  python excel_floxml_generator.py materials -o c:\\temp\\materials.xml
  python excel_floxml_generator.py materials --data test_materials.json -o c:\\temp\\materials.xml
  python excel_floxml_generator.py materials --workbook floxml_output\\materials_modified.xlsm -o c:\\temp\\materials.xml
        """.strip(),
    )

    parser.add_argument("template", nargs="?", help="template type")
    parser.add_argument("-o", "--output", help="output FloXML path")
    parser.add_argument("--data", type=Path, help="JSON file with template input data")
    parser.add_argument("--workbook", type=Path, help="use an existing workbook instead of copying the template")
    parser.add_argument("--macro", help="explicit macro name to run")
    parser.add_argument("--timeout", type=int, default=60, help="macro wait timeout in seconds")
    parser.add_argument("--list", action="store_true", help="list available templates")
    parser.add_argument(
        "--helper-run-macro",
        nargs=3,
        metavar=("WORKBOOK", "MACRO", "STATUS"),
        help=argparse.SUPPRESS,
    )
    return parser


def main(argv: Optional[Iterable[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(list(argv) if argv is not None else None)

    if args.helper_run_macro:
        workbook, macro, status = args.helper_run_macro
        return _run_macro_helper(Path(workbook), macro, Path(status))

    if args.list:
        list_templates()
        return 0

    if not args.template:
        parser.print_help()
        return 1

    if not args.output:
        parser.error("--output is required")

    output_path = Path(args.output).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists():
        output_path.unlink()

    generator = ExcelFloXMLGenerator(args.template)
    materials = load_materials(args.data) if args.template == "materials" else None

    result = generator.generate(
        output_floxml=output_path,
        materials=materials,
        workbook_path=args.workbook,
        macro_name=args.macro,
        timeout_seconds=args.timeout,
    )

    print("=" * 60)
    print("Excel FloXML Generator")
    print("=" * 60)
    print(f"Workbook : {result['workbook_path']}")
    print(f"Output   : {result['output_floxml']}")
    print(f"Macro    : {result['macro_name']}")
    print(f"Success  : {'yes' if result['success'] else 'no'}")

    if result["success"]:
        print("[OK] FloXML generated successfully.")
        return 0

    print("[ERROR] FloXML generation failed.")
    print(f"Reason   : {result['failure_reason']}")
    print(f"Tried    : {', '.join(result.get('macro_candidates', []))}")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
