#!/usr/bin/env python3
"""
FloXML 完整自动化解决方案

步骤:
1. 修改 Excel 数据
2. 添加 Auto_Open 宏（首次需要手动添加）
3. 运行脚本自动生成 FloXML
"""

import os
import sys
import json
import shutil
from pathlib import Path

# 尝试导入 xlwings
try:
    import xlwings as xw
    HAS_XLWINGS = True
except ImportError:
    HAS_XLWINGS = False

# 路径配置
TEMPLATES_DIR = Path(r"D:\Program Files\Siemens\SimcenterFlotherm\2504\examples\FloXML\Spreadsheets")
OUTPUT_DIR = Path(r"D:\Program Files\Siemens\SimcenterFlotherm\2504\flotherm-automation\floxml_output")

# Auto_Open VBA 代码模板
AUTO_OPEN_VBA = '''
Sub Auto_Open()
    ' 自动运行的宏
    ' 打开 Excel 时会自动执行

    On Error Resume Next

    ' 尝试运行 CREATEMATERIALS 宏
    Application.Run "CREATEMATERIALS"

    If Err.Number <> 0 Then
        ' 如果失败，尝试其他宏
        Err.Clear
        Application.Run "CREATEMODEL"
    End If

    ' 保存并关闭
    ThisWorkbook.Save

    ' 可选：自动关闭 Excel
    ' Application.Quit
End Sub
'''


def setup_excel_with_auto_open(excel_path: str) -> bool:
    """
    在 Excel 中添加 Auto_Open 宏

    注意：这需要 xlwings 和 GUI 环境
    """
    if not HAS_XLWINGS:
        print("[ERROR] 需要安装 xlwings: pip install xlwings")
        return False

    print(f"[INFO] 打开 Excel: {excel_path}")

    try:
        # 使用 xlwings 打开 Excel
        wb = xw.Book(excel_path)

        print("[INFO] Excel 已打开")
        print("[INFO] 请手动添加 Auto_Open 宏:")
        print("  1. 按 Alt+F11 打开 VBA 编辑器")
        print("  2. 在左侧选择你的工作簿")
        print("  3. 右键 → Insert → Module")
        print("  4. 粘贴以下代码:")
        print()
        print(AUTO_OPEN_VBA)
        print()
        print("  5. 保存并关闭 VBA 编辑器")
        print("  6. 保存 Excel 文件")

        input("按 Enter 键继续...")

        wb.save()
        wb.close()

        return True

    except Exception as e:
        print(f"[ERROR] {e}")
        return False


def generate_floxml_via_excel(excel_path: str, materials: list, output_path: str) -> bool:
    """
    通过 Excel 生成 FloXML
    """
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] 需要安装 openpyxl: pip install openpyxl")
        return False

    # 创建输出目录
    OUTPUT_DIR.mkdir(exist_ok=True)

    # 复制模板
    output_excel = OUTPUT_DIR / "floxml_generator.xlsm"
    shutil.copy(excel_path, output_excel)

    # 修改 Excel 数据
    wb = openpyxl.load_workbook(output_excel, keep_vba=True)
    ws = wb.active

    # 设置输出路径
    ws.cell(row=1, column=2, value=output_path)

    # 清除旧数据
    for row in range(4, 101):
        for col in range(1, 9):
            ws.cell(row=row, column=col, value=None)

    # 写入新材料数据
    for idx, mat in enumerate(materials):
        row = idx + 4
        ws.cell(row=row, column=1, value=mat.get("name"))
        ws.cell(row=row, column=2, value=mat.get("type", "Isotropic"))

        if mat.get("type") == "Orthotropic":
            ws.cell(row=row, column=3, value=mat.get("kx"))
            ws.cell(row=row, column=4, value=mat.get("ky"))
            ws.cell(row=row, column=5, value=mat.get("kz"))
        else:
            ws.cell(row=row, column=3, value=mat.get("kx"))

    wb.save(output_excel)
    wb.close()

    print(f"[OK] Excel 已更新: {output_excel}")
    print(f"[OK] 输出路径设置为: {output_path}")

    return True


def main():
    print("=" * 50)
    print("FloXML 自动化生成器")
    print("=" * 50)
    print()

    # 示例材料数据
    materials = [
        {"name": "Copper", "type": "Isotropic", "kx": 385},
        {"name": "FR4", "type": "Isotropic", "kx": 0.3},
        {"name": "Aluminum", "type": "Isotropic", "kx": 180},
    ]

    output_path = r"c:\temp\floxml_output.xml"
    template_path = TEMPLATES_DIR / "Materials.xlsm"

    # Step 1: 更新 Excel 数据
    print("[Step 1] 更新 Excel 数据...")
    success = generate_floxml_via_excel(str(template_path), materials, output_path)

    if not success:
        print("[ERROR] 更新 Excel 失败")
        return

    print()
    print("=" * 50)
    print("下一步操作:")
    print("=" * 50)
    print()
    print("方式 A - 手动运行:")
    print(f"  1. 打开: {OUTPUT_DIR / 'floxml_generator.xlsm'}")
    print("  2. 按 Alt+F8，运行宏")
    print()
    print("方式 B - 自动运行 (需要先添加 Auto_Open):")
    print(f"  运行: powershell -File auto_generate_floxml.ps1")
    print()
    print(f"输出文件: {output_path}")


if __name__ == '__main__':
    main()
