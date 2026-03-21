#!/usr/bin/env python3
"""
Grid Configuration Module

从 Excel 文件读取网格配置，用于 FloTHERM FloXML 项目。

Excel 格式:
  Sheet 1 "system_grid": 全局网格配置
  Sheet 2 "grid_patches": 局部加密补丁 (可选)
  Sheet 3 "grid_constraints": 网格约束 (可选)
  Sheet 4 "assembly_constraints": 装配件网格约束映射 (可选)
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Union
import xml.etree.ElementTree as ET


# ============================================================================
# 数据类定义
# ============================================================================

@dataclass
class SystemGridDirection:
    """单方向的系统网格配置"""
    min_size: float = 0.0001
    grid_type: str = "max_size"  # "max_size" 或 "min_number"
    max_size: float = 0.001
    min_number: int = 10
    smoothing_value: int = 12


@dataclass
class SystemGrid:
    """系统网格配置"""
    smoothing: bool = True
    smoothing_type: str = "v3"  # "v2" 或 "v3"
    dynamic_update: bool = True
    x_grid: SystemGridDirection = field(default_factory=SystemGridDirection)
    y_grid: SystemGridDirection = field(default_factory=SystemGridDirection)
    z_grid: SystemGridDirection = field(default_factory=SystemGridDirection)


@dataclass
class GridPatch:
    """网格补丁配置"""
    name: str = "Patch"
    applies_to: str = "x_direction"  # "x_direction", "y_direction", "z_direction"
    start_location: float = 0.0
    end_location: float = 0.1
    number_of_cells_control: str = "min_number"  # "additional_cells", "min_number", "max_size"
    additional_cells: int = 0
    min_number: int = 10
    max_size: float = 0.001
    cell_distribution: str = "uniform"  # "uniform", "increasing", "decreasing", "symmetrical"
    cell_distribution_index: float = 1.0


@dataclass
class GridInflation:
    """网格膨胀配置"""
    inflation_type: str = "size"  # "size" 或 "percent"
    inflation_size: float = 0.005
    inflation_percent: float = 10.0
    number_cells_control: str = "min_number"  # "max_size" 或 "min_number"
    max_size: float = 0.001
    min_number: int = 10


@dataclass
class GridConstraint:
    """网格约束配置"""
    name: str = "Grid Constraint"
    enable_min_cell_size: bool = True
    min_cell_size: float = 0.001
    number_cells_control: str = "min_number"  # "max_size" 或 "min_number"
    max_size: float = 0.001
    min_number: int = 10
    high_inflation: Optional[GridInflation] = None
    low_inflation: Optional[GridInflation] = None


@dataclass
class AssemblyGridConstraint:
    """装配件网格约束映射"""
    assembly_name: str = ""  # 装配件名称 (支持通配符 *)
    x_constraint: str = ""  # X方向约束名称
    y_constraint: str = ""  # Y方向约束名称
    z_constraint: str = ""  # Z方向约束名称
    all_constraint: str = ""  # 所有方向约束名称


@dataclass
class GridConfig:
    """完整的网格配置"""
    system_grid: SystemGrid = field(default_factory=SystemGrid)
    patches: List[GridPatch] = field(default_factory=list)
    constraints: List[GridConstraint] = field(default_factory=list)
    assembly_constraints: List[AssemblyGridConstraint] = field(default_factory=list)


# ============================================================================
# Excel 读取器
# ============================================================================

class GridExcelReader:
    """从 Excel 读取网格配置"""

    def __init__(self, filepath: str):
        self.filepath = filepath
        self._workbook = None

    def _load_workbook(self):
        """加载 Excel 文件"""
        if self._workbook is not None:
            return

        try:
            # 尝试使用 openpyxl
            from openpyxl import load_workbook
            self._workbook = load_workbook(self.filepath, read_only=True, data_only=True)
            self._use_openpyxl = True
        except ImportError:
            try:
                # 备选：使用 pandas
                import pandas as pd
                self._workbook = pd.ExcelFile(self.filepath)
                self._use_openpyxl = False
            except ImportError:
                raise ImportError("需要安装 openpyxl 或 pandas 来读取 Excel 文件")

    def read_config(self) -> GridConfig:
        """读取完整的网格配置"""
        self._load_workbook()

        config = GridConfig()
        config.system_grid = self._read_system_grid()
        config.patches = self._read_patches()
        config.constraints = self._read_constraints()
        config.assembly_constraints = self._read_assembly_constraints()

        return config

    def _read_system_grid(self) -> SystemGrid:
        """读取 system_grid 配置"""
        grid = SystemGrid()

        if self._use_openpyxl:
            # 查找 system_grid sheet
            sheet = None
            for name in self._workbook.sheetnames:
                if 'system_grid' in name.lower() or 'grid' in name.lower():
                    sheet = self._workbook[name]
                    break

            if sheet is None:
                # 使用第一个 sheet
                sheet = self._workbook.active

            # 读取数据
            data = {}
            for row in sheet.iter_rows(values_only=True):
                if row[0] is None:
                    continue
                key = str(row[0]).strip().lower()
                data[key] = row

            # 解析全局设置
            if 'smoothing' in data:
                grid.smoothing = self._parse_bool(data['smoothing'][1])
            if 'smoothing_type' in data:
                grid.smoothing_type = str(data['smoothing_type'][1] or 'v3').lower()
            if 'dynamic_update' in data:
                grid.dynamic_update = self._parse_bool(data['dynamic_update'][1])

            # 解析各方向网格
            grid.x_grid = self._parse_direction(data, 'x')
            grid.y_grid = self._parse_direction(data, 'y')
            grid.z_grid = self._parse_direction(data, 'z')

        else:
            # 使用 pandas
            import pandas as pd
            df = pd.read_excel(self._workbook, sheet_name=0)

            # 转换为字典
            data = {}
            for _, row in df.iterrows():
                key = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
                if key:
                    data[key] = row.tolist()

            # 解析
            if 'smoothing' in data:
                grid.smoothing = self._parse_bool(data['smoothing'][1])
            if 'smoothing_type' in data:
                grid.smoothing_type = str(data['smoothing_type'][1] or 'v3').lower()
            if 'dynamic_update' in data:
                grid.dynamic_update = self._parse_bool(data['dynamic_update'][1])

            grid.x_grid = self._parse_direction(data, 'x')
            grid.y_grid = self._parse_direction(data, 'y')
            grid.z_grid = self._parse_direction(data, 'z')

        return grid

    def _parse_direction(self, data: Dict, axis: str) -> SystemGridDirection:
        """解析单方向网格配置"""
        direction = SystemGridDirection()

        # 尝试多种键名格式
        min_size_key = f'{axis}_min_size'
        grid_type_key = f'{axis}_grid_type'
        max_size_key = f'{axis}_max_size'
        min_number_key = f'{axis}_min_number'
        smoothing_key = f'{axis}_smoothing_value'

        # 也尝试大写
        axis_upper = axis.upper()

        for key, row in data.items():
            val_x = row[1] if len(row) > 1 else None
            val_y = row[2] if len(row) > 2 else None
            val_z = row[3] if len(row) > 3 else None

            if axis == 'x':
                val = val_x
            elif axis == 'y':
                val = val_y
            else:
                val = val_z

            if val is None:
                continue

            key_lower = key.lower()

            if 'min_size' in key_lower:
                direction.min_size = float(val)
            elif 'grid_type' in key_lower:
                direction.grid_type = str(val).lower()
            elif 'max_size' in key_lower:
                direction.max_size = float(val)
            elif 'min_number' in key_lower:
                direction.min_number = int(val)
            elif 'smoothing_value' in key_lower:
                direction.smoothing_value = int(val)

        return direction

    def _parse_bool(self, value) -> bool:
        """解析布尔值"""
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ('true', '1', 'yes', 'on')
        return bool(value)

    def _read_patches(self) -> List[GridPatch]:
        """读取 grid_patches 配置"""
        patches = []

        # 查找 patches sheet
        sheet_name = None
        if self._use_openpyxl:
            for name in self._workbook.sheetnames:
                if 'patch' in name.lower():
                    sheet_name = name
                    break

            if sheet_name is None:
                return patches

            sheet = self._workbook[sheet_name]

            # 读取表头
            headers = []
            first_row = True
            for row in sheet.iter_rows(values_only=True):
                if first_row:
                    headers = [str(h).lower().strip() if h else '' for h in row]
                    first_row = False
                    continue

                if row[0] is None:
                    continue

                patch = GridPatch()
                for i, header in enumerate(headers):
                    if i >= len(row):
                        break
                    val = row[i]
                    if val is None:
                        continue

                    if 'name' in header:
                        patch.name = str(val)
                    elif 'applies_to' in header:
                        patch.applies_to = str(val).lower()
                    elif 'start' in header:
                        patch.start_location = float(val)
                    elif 'end' in header:
                        patch.end_location = float(val)
                    elif 'control' in header:
                        patch.number_of_cells_control = str(val).lower()
                    elif 'additional' in header:
                        patch.additional_cells = int(val)
                    elif 'min_number' in header or 'min_num' in header:
                        patch.min_number = int(val)
                    elif 'max_size' in header:
                        patch.max_size = float(val)
                    elif 'distribution' in header and 'index' not in header:
                        patch.cell_distribution = str(val).lower()
                    elif 'index' in header:
                        patch.cell_distribution_index = float(val)

                patches.append(patch)

        else:
            # pandas 方式
            import pandas as pd
            try:
                df = pd.read_excel(self._workbook, sheet_name='grid_patches')
            except:
                try:
                    df = pd.read_excel(self._workbook, sheet_name=1)
                except:
                    return patches

            for _, row in df.iterrows():
                patch = GridPatch()

                for col in df.columns:
                    val = row[col]
                    if pd.isna(val):
                        continue

                    col_lower = str(col).lower()

                    if 'name' in col_lower:
                        patch.name = str(val)
                    elif 'applies' in col_lower:
                        patch.applies_to = str(val).lower()
                    elif 'start' in col_lower:
                        patch.start_location = float(val)
                    elif 'end' in col_lower:
                        patch.end_location = float(val)
                    elif 'control' in col_lower:
                        patch.number_of_cells_control = str(val).lower()
                    elif 'additional' in col_lower:
                        patch.additional_cells = int(val)
                    elif 'min_number' in col_lower:
                        patch.min_number = int(val)
                    elif 'max_size' in col_lower:
                        patch.max_size = float(val)
                    elif 'distribution' in col_lower and 'index' not in col_lower:
                        patch.cell_distribution = str(val).lower()
                    elif 'index' in col_lower:
                        patch.cell_distribution_index = float(val)

                patches.append(patch)

        return patches

    def _read_constraints(self) -> List[GridConstraint]:
        """读取 grid_constraints 配置"""
        constraints = []

        # 查找 constraints sheet
        if self._use_openpyxl:
            sheet_name = None
            for name in self._workbook.sheetnames:
                if 'constraint' in name.lower():
                    sheet_name = name
                    break

            if sheet_name is None:
                return constraints

            sheet = self._workbook[sheet_name]

            # 读取表头
            headers = []
            first_row = True
            for row in sheet.iter_rows(values_only=True):
                if first_row:
                    headers = [str(h).lower().strip() if h else '' for h in row]
                    first_row = False
                    continue

                if row[0] is None:
                    continue

                constraint = GridConstraint()
                for i, header in enumerate(headers):
                    if i >= len(row):
                        break
                    val = row[i]
                    if val is None:
                        continue

                    if header == 'name':
                        constraint.name = str(val)
                    elif 'enable_min_cell' in header:
                        constraint.enable_min_cell_size = self._parse_bool(val)
                    elif header == 'min_cell_size':
                        constraint.min_cell_size = float(val)
                    elif header == 'number_cells_control':
                        constraint.number_cells_control = str(val).lower()
                    elif header == 'max_size':
                        constraint.max_size = float(val)
                    elif header == 'min_number':
                        constraint.min_number = int(val)
                    elif 'high_inflation' in header:
                        if constraint.high_inflation is None:
                            constraint.high_inflation = GridInflation()
                        if 'type' in header:
                            constraint.high_inflation.inflation_type = str(val).lower()
                        elif 'size' in header:
                            constraint.high_inflation.inflation_size = float(val)
                        elif 'percent' in header:
                            constraint.high_inflation.inflation_percent = float(val)
                    elif 'low_inflation' in header:
                        if constraint.low_inflation is None:
                            constraint.low_inflation = GridInflation()
                        if 'type' in header:
                            constraint.low_inflation.inflation_type = str(val).lower()
                        elif 'size' in header:
                            constraint.low_inflation.inflation_size = float(val)
                        elif 'percent' in header:
                            constraint.low_inflation.inflation_percent = float(val)

                constraints.append(constraint)

        return constraints

    def _read_assembly_constraints(self) -> List[AssemblyGridConstraint]:
        """读取 assembly_constraints 配置"""
        mappings = []

        # 查找 assembly_constraints sheet
        if self._use_openpyxl:
            sheet_name = None
            for name in self._workbook.sheetnames:
                if 'assembly' in name.lower():
                    sheet_name = name
                    break

            if sheet_name is None:
                return mappings

            sheet = self._workbook[sheet_name]

            # 读取表头
            headers = []
            first_row = True
            for row in sheet.iter_rows(values_only=True):
                if first_row:
                    headers = [str(h).lower().strip() if h else '' for h in row]
                    first_row = False
                    continue

                if row[0] is None:
                    continue

                mapping = AssemblyGridConstraint()
                for i, header in enumerate(headers):
                    if i >= len(row):
                        break
                    val = row[i]
                    if val is None:
                        continue

                    if 'assembly' in header and 'name' in header:
                        mapping.assembly_name = str(val)
                    elif header == 'x_constraint':
                        mapping.x_constraint = str(val)
                    elif header == 'y_constraint':
                        mapping.y_constraint = str(val)
                    elif header == 'z_constraint':
                        mapping.z_constraint = str(val)
                    elif header == 'all_constraint':
                        mapping.all_constraint = str(val)

                if mapping.assembly_name:
                    mappings.append(mapping)

        return mappings


# ============================================================================
# Grid XML 构建器
# ============================================================================

class GridBuilder:
    """构建 Grid XML 元素"""

    def build_grid(self, config: GridConfig) -> ET.Element:
        """构建完整的 grid 元素"""
        grid = ET.Element("grid")

        # system_grid
        grid.append(self._build_system_grid(config.system_grid))

        # patches (如果有)
        if config.patches:
            grid.append(self._build_patches(config.patches))

        return grid

    def _build_system_grid(self, config: SystemGrid) -> ET.Element:
        """构建 system_grid 元素"""
        system_grid = ET.Element("system_grid")

        # 添加子元素
        self._append_text(system_grid, "smoothing", "true" if config.smoothing else "false")
        self._append_text(system_grid, "smoothing_type", config.smoothing_type)
        self._append_text(system_grid, "dynamic_update", "true" if config.dynamic_update else "false")

        # x_grid, y_grid, z_grid
        system_grid.append(self._build_direction("x_grid", config.x_grid))
        system_grid.append(self._build_direction("y_grid", config.y_grid))
        system_grid.append(self._build_direction("z_grid", config.z_grid))

        return system_grid

    def _build_direction(self, tag: str, config: SystemGridDirection) -> ET.Element:
        """构建方向网格元素"""
        elem = ET.Element(tag)

        self._append_text(elem, "min_size", f"{config.min_size:.6g}")
        self._append_text(elem, "grid_type", config.grid_type)

        if config.grid_type == "max_size":
            self._append_text(elem, "max_size", f"{config.max_size:.6g}")
        else:
            self._append_text(elem, "min_number", str(config.min_number))

        self._append_text(elem, "smoothing_value", str(config.smoothing_value))

        return elem

    def _build_patches(self, patches: List[GridPatch]) -> ET.Element:
        """构建 patches 元素"""
        patches_elem = ET.Element("patches")

        for patch in patches:
            patches_elem.append(self._build_patch(patch))

        return patches_elem

    def _build_patch(self, patch: GridPatch) -> ET.Element:
        """构建单个 patch 元素"""
        elem = ET.Element("grid_patch")

        self._append_text(elem, "name", patch.name)
        self._append_text(elem, "applies_to", patch.applies_to)
        self._append_text(elem, "start_location", f"{patch.start_location:.6g}")
        self._append_text(elem, "end_location", f"{patch.end_location:.6g}")
        self._append_text(elem, "number_of_cells_control", patch.number_of_cells_control)

        if patch.number_of_cells_control == "additional_cells":
            self._append_text(elem, "additional_cells", str(patch.additional_cells))
        elif patch.number_of_cells_control == "min_number":
            self._append_text(elem, "min_number", str(patch.min_number))
        else:  # max_size
            self._append_text(elem, "max_size", f"{patch.max_size:.6g}")

        self._append_text(elem, "cell_distribution", patch.cell_distribution)
        self._append_text(elem, "cell_distribution_index", f"{patch.cell_distribution_index:.6g}")

        return elem

    def build_constraints_attributes(self, constraints: List[GridConstraint]) -> ET.Element:
        """构建 grid_constraints 属性元素 (用于 attributes 节)"""
        constraints_elem = ET.Element("grid_constraints")

        for constraint in constraints:
            constraints_elem.append(self._build_constraint(constraint))

        return constraints_elem

    def _build_constraint(self, constraint: GridConstraint) -> ET.Element:
        """构建单个 grid_constraint_att 元素"""
        elem = ET.Element("grid_constraint_att")

        self._append_text(elem, "name", constraint.name)
        self._append_text(elem, "enable_min_cell_size", "true" if constraint.enable_min_cell_size else "false")
        self._append_text(elem, "min_cell_size", f"{constraint.min_cell_size:.6g}")
        self._append_text(elem, "number_cells_control", constraint.number_cells_control)

        if constraint.number_cells_control == "max_size":
            self._append_text(elem, "max_size", f"{constraint.max_size:.6g}")
        else:
            self._append_text(elem, "min_number", str(constraint.min_number))

        if constraint.high_inflation:
            elem.append(self._build_inflation("high_inflation", constraint.high_inflation))

        if constraint.low_inflation:
            elem.append(self._build_inflation("low_inflation", constraint.low_inflation))

        return elem

    def _build_inflation(self, tag: str, inflation: GridInflation) -> ET.Element:
        """构建 inflation 元素"""
        elem = ET.Element(tag)

        self._append_text(elem, "inflation_type", inflation.inflation_type)

        if inflation.inflation_type == "size":
            self._append_text(elem, "inflation_size", f"{inflation.inflation_size:.6g}")
        else:
            self._append_text(elem, "inflation_percent", f"{inflation.inflation_percent:.6g}")

        self._append_text(elem, "number_cells_control", inflation.number_cells_control)

        if inflation.number_cells_control == "max_size":
            self._append_text(elem, "max_size", f"{inflation.max_size:.6g}")
        else:
            self._append_text(elem, "min_number", str(inflation.min_number))

        return elem

    def _append_text(self, parent: ET.Element, tag: str, text: str) -> ET.Element:
        """添加带文本的子元素"""
        elem = ET.SubElement(parent, tag)
        elem.text = text
        return elem


# ============================================================================
# 示例 Excel 模板生成
# ============================================================================

def create_template_excel(output_path: str) -> None:
    """创建示例 Excel 模板"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        wb = Workbook()

        # Sheet 1: system_grid
        ws1 = wb.active
        ws1.title = "system_grid"

        # 表头样式
        header_font = Font(bold=True)
        center_align = Alignment(horizontal='center')

        # 写入表头
        headers = ["参数", "X", "Y", "Z", "说明"]
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align

        # 写入数据
        data = [
            ["min_size", 0.0001, 0.0001, 0.0005, "最小网格尺寸 (米)"],
            ["grid_type", "max_size", "max_size", "max_size", "max_size 或 min_number"],
            ["max_size", 0.001, 0.001, 0.001, "最大网格尺寸 (米)"],
            ["min_number", 10, 10, 10, "最小网格数 (当 grid_type=min_number)"],
            ["smoothing_value", 12, 12, 12, "平滑值 (0-50)"],
            ["", "", "", "", ""],
            ["smoothing", "true", "", "", "是否启用平滑"],
            ["smoothing_type", "v3", "", "", "平滑类型: v2 或 v3"],
            ["dynamic_update", "true", "", "", "是否动态更新"],
        ]

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws1.cell(row=row_idx, column=col_idx, value=value)

        # 调整列宽
        ws1.column_dimensions['A'].width = 20
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 40

        # Sheet 2: grid_patches (可选)
        ws2 = wb.create_sheet("grid_patches")

        headers2 = ["name", "applies_to", "start_location", "end_location",
                    "number_of_cells_control", "value", "cell_distribution", "说明"]
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align

        data2 = [
            ["Fin_Region_X", "x_direction", 0, 0.04, "min_number", 50, "uniform", "鳍片区域X方向加密"],
            ["HeatSource_Y", "y_direction", -0.005, 0.001, "max_size", 0.0002, "symmetrical", "热源区域Y方向加密"],
        ]

        for row_idx, row_data in enumerate(data2, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws2.cell(row=row_idx, column=col_idx, value=value)

        # 调整列宽
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws2.column_dimensions[col].width = 18

        # Sheet 3: grid_constraints
        ws3 = wb.create_sheet("grid_constraints")

        headers3 = ["name", "enable_min_cell_size", "min_cell_size", "number_cells_control",
                    "max_size", "min_number", "说明"]
        for col, header in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align

        data3 = [
            ["Fin_Constraint", "true", 0.0005, "min_number", 0.002, 20, "鳍片区域网格约束"],
            ["Base_Constraint", "true", 0.001, "max_size", 0.002, 10, "基座区域网格约束"],
        ]

        for row_idx, row_data in enumerate(data3, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws3.cell(row=row_idx, column=col_idx, value=value)

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws3.column_dimensions[col].width = 20

        # Sheet 4: assembly_constraints
        ws4 = wb.create_sheet("assembly_constraints")

        headers4 = ["assembly_name", "x_constraint", "y_constraint", "z_constraint", "all_constraint", "说明"]
        for col, header in enumerate(headers4, 1):
            cell = ws4.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align

        data4 = [
            ["Fin*", "Fin_Constraint", "", "", "", "所有 Fin 开头的装配件应用 X 方向约束"],
            ["Base", "", "", "", "Base_Constraint", "Base 装配件应用全方向约束"],
        ]

        for row_idx, row_data in enumerate(data4, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws4.cell(row=row_idx, column=col_idx, value=value)

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws4.column_dimensions[col].width = 20

        wb.save(output_path)
        print(f"模板已创建: {output_path}")

    except ImportError:
        print("需要安装 openpyxl 来创建 Excel 模板: pip install openpyxl")


# ============================================================================
# CLI 接口
# ============================================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="网格配置工具")
    parser.add_argument("--template", "-t", type=str, help="创建示例 Excel 模板")
    parser.add_argument("--read", "-r", type=str, help="读取 Excel 配置并输出 XML")
    parser.add_argument("--output", "-o", type=str, help="输出文件路径")

    args = parser.parse_args()

    if args.template:
        create_template_excel(args.template)
    elif args.read:
        reader = GridExcelReader(args.read)
        config = reader.read_config()

        builder = GridBuilder()
        grid_xml = builder.build_grid(config)

        # 缩进
        from xml.dom import minidom
        xml_str = ET.tostring(grid_xml, encoding='unicode')
        dom = minidom.parseString(xml_str)
        pretty_xml = dom.toprettyxml(indent="  ")

        if args.output:
            Path(args.output).write_text(pretty_xml, encoding='utf-8')
            print(f"Grid XML 已保存: {args.output}")
        else:
            print(pretty_xml)
    else:
        parser.print_help()
