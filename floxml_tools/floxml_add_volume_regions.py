#!/usr/bin/env python3
"""
Add FloTHERM volume regions (<region>) to an existing FloXML project using JSON config.

The same config can also:
- create/update grid_constraint_att definitions
- assign grid constraints directly onto existing geometry objects

Supported region definition modes:
1. Explicit:
   - position: [x, y, z]
   - size: [sx, sy, sz]

2. Derived from existing geometry bounding boxes:
   - bbox_from:
       include_names: ["PCB", "U1", "U2"]
       include_patterns: ["R22 *", "C*"]
       include_tags: ["cuboid"]
       padding: 0.001
       or padding: [0.001, 0.001, 0.0005]

Each region may be inserted at root geometry or under a named assembly.
"""

from __future__ import annotations

import argparse
import copy
import fnmatch
import json
from dataclasses import dataclass
from pathlib import Path
import sys
from typing import Dict, Iterable, List, Optional, Tuple
import xml.etree.ElementTree as ET


Vector3 = Tuple[float, float, float]


@dataclass
class GeometryItem:
    name: str
    tag: str
    position: Vector3
    size: Optional[Vector3]
    global_position: Vector3
    global_size: Optional[Vector3]
    element: ET.Element
    parent_geometry: ET.Element
    assembly_path: Tuple[str, ...]


def _float_text(parent: ET.Element, tag: str, default: float = 0.0) -> float:
    child = parent.find(tag)
    if child is None or child.text is None:
        return default
    try:
        return float(child.text.strip())
    except ValueError:
        return default


def _parse_position(elem: ET.Element) -> Vector3:
    pos = elem.find("position")
    if pos is None:
        return (0.0, 0.0, 0.0)
    return (
        _float_text(pos, "x"),
        _float_text(pos, "y"),
        _float_text(pos, "z"),
    )


def _parse_size(elem: ET.Element) -> Optional[Vector3]:
    size = elem.find("size")
    if size is None:
        return None
    return (
        _float_text(size, "x"),
        _float_text(size, "y"),
        _float_text(size, "z"),
    )


def _append_text(parent: ET.Element, tag: str, text: str) -> ET.Element:
    elem = ET.SubElement(parent, tag)
    elem.text = text
    return elem


def _set_text(parent: ET.Element, tag: str, text: str) -> ET.Element:
    child = parent.find(tag)
    if child is None:
        child = ET.SubElement(parent, tag)
    child.text = text
    return child


def _vector_add(a: Vector3, b: Vector3) -> Vector3:
    return (a[0] + b[0], a[1] + b[1], a[2] + b[2])


def _vector_sub(a: Vector3, b: Vector3) -> Vector3:
    return (a[0] - b[0], a[1] - b[1], a[2] - b[2])


def _normalize_padding(value) -> Vector3:
    if isinstance(value, (int, float)):
        padding = float(value)
        return (padding, padding, padding)
    if isinstance(value, list) and len(value) == 3:
        return (float(value[0]), float(value[1]), float(value[2]))
    return (0.0, 0.0, 0.0)


def _iter_geometry_items(
    geometry_elem: ET.Element,
    parent_global: Vector3 = (0.0, 0.0, 0.0),
    assembly_path: Tuple[str, ...] = (),
) -> Iterable[GeometryItem]:
    for child in list(geometry_elem):
        tag = child.tag
        position = _parse_position(child)
        size = _parse_size(child)
        global_position = _vector_add(parent_global, position)
        global_size = size
        name = (child.findtext("name") or "").strip()

        yield GeometryItem(
            name=name,
            tag=tag,
            position=position,
            size=size,
            global_position=global_position,
            global_size=global_size,
            element=child,
            parent_geometry=geometry_elem,
            assembly_path=assembly_path,
        )

        child_geometry = child.find("geometry")
        if child_geometry is not None:
            next_path = assembly_path
            if tag == "assembly" and name:
                next_path = assembly_path + (name,)
            yield from _iter_geometry_items(
                child_geometry,
                parent_global=global_position,
                assembly_path=next_path,
            )


def _iter_all_geometry_items(root_geometry: ET.Element) -> Iterable[GeometryItem]:
    for item in _iter_geometry_items(root_geometry):
        yield item
        child_geometry = item.element.find("geometry")
        if child_geometry is not None:
            next_path = item.assembly_path
            if item.tag == "assembly" and item.name:
                next_path = item.assembly_path + (item.name,)
            yield from _iter_geometry_items(
                child_geometry,
                parent_global=item.global_position,
                assembly_path=next_path,
            )


def _find_root_geometry(root: ET.Element) -> ET.Element:
    geometry = root.find("geometry")
    if geometry is None:
        raise ValueError("FloXML missing <geometry> section")
    return geometry


def _find_or_create_attributes(root: ET.Element) -> ET.Element:
    attributes = root.find("attributes")
    if attributes is None:
        insert_at = 0
        for idx, child in enumerate(list(root)):
            if child.tag in {"geometry", "solution_domain"}:
                insert_at = idx
                break
            insert_at = idx + 1
        attributes = ET.Element("attributes")
        root.insert(insert_at, attributes)
    return attributes


def _find_or_create_grid_constraints(attributes: ET.Element) -> ET.Element:
    grid_constraints = attributes.find("grid_constraints")
    if grid_constraints is None:
        grid_constraints = ET.SubElement(attributes, "grid_constraints")
    return grid_constraints


def _find_assembly_element(root_geometry: ET.Element, assembly_name: str) -> Optional[ET.Element]:
    for item in _iter_all_geometry_items(root_geometry):
        if item.tag == "assembly" and item.name == assembly_name:
            return item.element
    return None


def _ensure_geometry_container(parent_elem: ET.Element) -> ET.Element:
    geometry = parent_elem.find("geometry")
    if geometry is None:
        geometry = ET.SubElement(parent_elem, "geometry")
    return geometry


def _compute_bbox(items: List[GeometryItem]) -> Tuple[Vector3, Vector3]:
    if not items:
        raise ValueError("No matching geometry items found for bbox_from")

    mins = [float("inf")] * 3
    maxs = [float("-inf")] * 3

    for item in items:
        if item.global_size is None:
            continue
        for axis in range(3):
            mins[axis] = min(mins[axis], item.global_position[axis])
            maxs[axis] = max(maxs[axis], item.global_position[axis] + item.global_size[axis])

    if mins[0] == float("inf"):
        no_size = [f"{item.tag}/{item.name}" for item in items if item.global_size is None]
        raise ValueError(
            f"Matching items did not have usable size data for bbox_from. "
            f"Items without size: {no_size}"
        )

    return (tuple(mins), tuple(maxs))  # type: ignore[arg-type]


def _match_items(
    root_geometry: ET.Element,
    include_names: List[str],
    include_patterns: List[str],
    include_tags: List[str],
    scope_assembly: Optional[str],
) -> List[GeometryItem]:
    matches: List[GeometryItem] = []
    for item in _iter_all_geometry_items(root_geometry):
        if not item.name:
            continue
        if scope_assembly and scope_assembly not in item.assembly_path and item.name != scope_assembly:
            continue
        if include_tags and item.tag not in include_tags:
            continue
        if item.name in include_names or any(fnmatch.fnmatch(item.name, pattern) for pattern in include_patterns):
            matches.append(item)
    return matches


def _apply_object_constraint(item: GeometryItem, cfg: Dict) -> None:
    for constraint_tag in ("x_grid_constraint", "y_grid_constraint", "z_grid_constraint", "all_grid_constraint"):
        if cfg.get(constraint_tag):
            _set_text(item.element, constraint_tag, str(cfg[constraint_tag]))
    if cfg.get("localized_grid") is not None:
        _set_text(item.element, "localized_grid", "true" if cfg["localized_grid"] else "false")


def _apply_object_constraints(root_geometry: ET.Element, config: Dict) -> int:
    applied = 0
    for obj_cfg in config.get("object_constraints", []):
        include_names = list(obj_cfg.get("target_names", []))
        include_patterns = list(obj_cfg.get("target_patterns", []))
        include_tags = list(obj_cfg.get("target_tags", []))
        scope_assembly = obj_cfg.get("scope_assembly")
        matches = _match_items(root_geometry, include_names, include_patterns, include_tags, scope_assembly)
        if not matches:
            raise ValueError(
                f"object_constraints target did not match any geometry: "
                f"names={include_names}, patterns={include_patterns}, tags={include_tags}, "
                f"scope_assembly={scope_assembly}"
            )
        for item in matches:
            _apply_object_constraint(item, obj_cfg)
            applied += 1
    return applied


def _build_region_element(name: str, position: Vector3, size: Vector3, cfg: Dict) -> ET.Element:
    region = ET.Element("region")
    _append_text(region, "name", name)
    _append_text(region, "active", "true" if cfg.get("active", True) else "false")
    if cfg.get("hidden") is not None:
        _append_text(region, "hidden", "true" if cfg["hidden"] else "false")

    pos = ET.SubElement(region, "position")
    _append_text(pos, "x", f"{position[0]:.6g}")
    _append_text(pos, "y", f"{position[1]:.6g}")
    _append_text(pos, "z", f"{position[2]:.6g}")

    size_elem = ET.SubElement(region, "size")
    _append_text(size_elem, "x", f"{size[0]:.6g}")
    _append_text(size_elem, "y", f"{size[1]:.6g}")
    _append_text(size_elem, "z", f"{size[2]:.6g}")

    orientation = ET.SubElement(region, "orientation")
    for axis_tag, vec in (
        ("local_x", ("1", "0", "0")),
        ("local_y", ("0", "0", "1")),
        ("local_z", ("0", "1", "0")),
    ):
        axis = ET.SubElement(orientation, axis_tag)
        _append_text(axis, "i", vec[0])
        _append_text(axis, "j", vec[1])
        _append_text(axis, "k", vec[2])

    for constraint_tag in ("x_grid_constraint", "y_grid_constraint", "z_grid_constraint"):
        if cfg.get(constraint_tag):
            _append_text(region, constraint_tag, str(cfg[constraint_tag]))

    if cfg.get("all_grid_constraint"):
        _append_text(region, "all_grid_constraint", str(cfg["all_grid_constraint"]))

    _append_text(region, "localized_grid", "true" if cfg.get("localized_grid", True) else "false")
    return region


def _upsert_grid_constraint(grid_constraints_elem: ET.Element, cfg: Dict) -> ET.Element:
    name = cfg.get("name")
    if not name:
        raise ValueError("Each grid constraint requires a name")

    existing = None
    for child in grid_constraints_elem.findall("grid_constraint_att"):
        if (child.findtext("name") or "").strip() == name:
            existing = child
            break

    elem = existing if existing is not None else ET.SubElement(grid_constraints_elem, "grid_constraint_att")
    _set_text(elem, "name", str(name))
    _set_text(elem, "enable_min_cell_size", "true" if cfg.get("enable_min_cell_size", True) else "false")

    if cfg.get("min_cell_size") is not None:
        _set_text(elem, "min_cell_size", f"{float(cfg['min_cell_size']):.6g}")

    _set_text(elem, "number_cells_control", str(cfg.get("number_cells_control", "min_number")))
    if cfg.get("min_number") is not None:
        _set_text(elem, "min_number", str(cfg["min_number"]))

    hi_cfg = cfg.get("high_inflation")
    if hi_cfg:
        hi = elem.find("high_inflation")
        if hi is None:
            hi = ET.SubElement(elem, "high_inflation")
        _set_text(hi, "inflation_type", str(hi_cfg.get("inflation_type", "size")))
        if hi_cfg.get("inflation_size") is not None:
            _set_text(hi, "inflation_size", f"{float(hi_cfg['inflation_size']):.6g}")
        _set_text(hi, "number_cells_control", str(hi_cfg.get("number_cells_control", "min_number")))
        if hi_cfg.get("min_number") is not None:
            _set_text(hi, "min_number", str(hi_cfg["min_number"]))

    return elem


def _resolve_region_geometry(
    root_geometry: ET.Element,
    region_cfg: Dict,
) -> Tuple[Vector3, Vector3]:
    if "position" in region_cfg and "size" in region_cfg:
        position = tuple(float(v) for v in region_cfg["position"])
        size = tuple(float(v) for v in region_cfg["size"])
        return position, size  # type: ignore[return-value]

    bbox_cfg = region_cfg.get("bbox_from")
    if not bbox_cfg:
        raise ValueError(f"Region '{region_cfg.get('name', '<unnamed>')}' missing position/size or bbox_from")

    include_names = list(bbox_cfg.get("include_names", []))
    include_patterns = list(bbox_cfg.get("include_patterns", []))
    include_tags = list(bbox_cfg.get("include_tags", []))
    scope_assembly = bbox_cfg.get("scope_assembly")
    matches = _match_items(root_geometry, include_names, include_patterns, include_tags, scope_assembly)
    lower, upper = _compute_bbox(matches)
    padding = _normalize_padding(bbox_cfg.get("padding", 0.0))
    position = tuple(lower[i] - padding[i] for i in range(3))
    size = tuple((upper[i] - lower[i]) + (2.0 * padding[i]) for i in range(3))
    return position, size  # type: ignore[return-value]


def _bbox_volume(items: List[GeometryItem]) -> float:
    """Compute bbox volume of a group of items."""
    lower, upper = _compute_bbox(items)
    return (upper[0] - lower[0]) * (upper[1] - lower[1]) * (upper[2] - lower[2])


def _decompose_selected_items(
    selected_items: List[GeometryItem],
    root_geometry: ET.Element,
    min_volume_reduction: float = 0.2,
    obstacles: Optional[List[GeometryItem]] = None,
    ignore_obstacles: bool = False,
) -> List[List[GeometryItem]]:
    """
    Line-first directional merge: minimize region count without wrapping obstacles.

    Algorithm overview:
        1. n matched items → n initial regions (each item's bbox + padding)
        2. Phase 1 – overlap merge: items occupying the same 3D space
           (e.g. die + source on die) are merged into one group.
        3. Phase 2 – directional line merge: for each axis (x → y → z),
           find pairs of regions that are collinear (overlap on the other 2 axes)
           and adjacent (small gap on this axis). Merge if no obstacles in the
           combined bbox. Repeat until no more merges are possible.
        4. Result: minimum number of rectangular regions that cover all selected
           items without including any non-selected geometry (obstacles).

    Obstacle check uses XY projection (2D) because sources are very thin in Z
    and a 3D overlap check would miss XY-plane conflicts.
    """
    if len(selected_items) <= 1:
        return [selected_items] if selected_items else []

    # ------------------------------------------------------------------
    # Build obstacle context
    # ------------------------------------------------------------------
    # usable_set: ALL geometry in the same assemblies as selected items is
    # excluded from obstacles. When user selects a source from a chip, the
    # entire chip (substrate, die, mold, etc.) is "selected territory".
    all_geometry_items = list(_iter_geometry_items(root_geometry))
    selected_paths = {item.assembly_path for item in selected_items}
    usable_set: Set[int] = set()
    for gitem in all_geometry_items:
        if gitem.assembly_path in selected_paths:
            usable_set.add(id(gitem.element))
            continue
        # Hierarchy: ancestor containers (path is prefix of a selected path)
        # are not obstacles — they enclose the selected items.
        for sp in selected_paths:
            if sp[:len(gitem.assembly_path)] == gitem.assembly_path:
                usable_set.add(id(gitem.element))
                break

    # ------------------------------------------------------------------
    # Phase 1: merge items that physically overlap in 3D space
    # ------------------------------------------------------------------
    # Items in the same component (e.g. die + source) occupy the same space
    # and must form one group. Spatially separated items stay separate.
    tol = 1e-9
    groups: List[List[GeometryItem]] = [[item] for item in selected_items]

    def _items_overlap_3d(a: List[GeometryItem], b: List[GeometryItem]) -> bool:
        """Check if two groups overlap on ALL 3 axes (AABB intersection)."""
        lo_a, hi_a = _compute_bbox(a)
        lo_b, hi_b = _compute_bbox(b)
        for axis in range(3):
            overlap = min(hi_a[axis], hi_b[axis]) - max(lo_a[axis], lo_b[axis])
            if overlap <= tol:
                return False
        return True

    # Keep merging overlapping pairs until stable
    changed = True
    while changed:
        changed = False
        for i in range(len(groups)):
            for j in range(i + 1, len(groups)):
                if _items_overlap_3d(groups[i], groups[j]):
                    groups[i] = groups[i] + groups[j]
                    groups.pop(j)
                    changed = True
                    break
            if changed:
                break

    if len(groups) <= 1:
        return groups

    # ------------------------------------------------------------------
    # Obstacle helper (XY projection)
    # ------------------------------------------------------------------
    def _has_obstacles(items: List[GeometryItem], return_blocker: bool = False):
        """Check if the bbox of items contains any non-selected geometry.

        Uses full 3D AABB overlap test (all 3 axes). Items at different Z
        levels (e.g. a PCB below thin sources) won't block merges, while
        items at the same level (e.g. other sources) will.
        """
        lower, upper = _compute_bbox(items)
        for gitem in all_geometry_items:
            if gitem.global_size is None:
                continue
            if id(gitem.element) in usable_set:
                continue
            # Spatial: if this item fully contains the merged bbox, it's a
            # container — not an obstacle.  The volume region sits inside it.
            item_lo = gitem.global_position
            item_hi = _vector_add(item_lo, gitem.global_size)
            contains = all(
                item_lo[axis] - tol <= lower[axis] and upper[axis] <= item_hi[axis] + tol
                for axis in range(3)
            )
            if contains:
                continue
            # 3D overlap check (all axes)
            # Special case: if both the region bbox and the obstacle have
            # zero extent on an axis and are at the same position, treat
            # as overlapping (handles thin/zero-thickness sources).
            blocked = True
            for axis in range(3):
                item_lo = gitem.global_position[axis]
                item_hi = item_lo + gitem.global_size[axis]
                overlap = min(item_hi, upper[axis]) - max(item_lo, lower[axis])
                if overlap > tol:
                    continue
                # Zero-extent special case
                extent_item = item_hi - item_lo
                extent_bbox = upper[axis] - lower[axis]
                if extent_item <= tol and extent_bbox <= tol and abs(item_lo - lower[axis]) <= tol:
                    continue
                blocked = False
                break
            if blocked:
                if return_blocker:
                    return gitem.name or gitem.tag
                return True
        if return_blocker:
            return None
        return False

    # ------------------------------------------------------------------
    # Phase 2: directional line merge along each axis
    # ------------------------------------------------------------------
    # For each axis, repeatedly find pairs of groups that are:
    #   - Collinear: overlap on the other 2 axes (perpendicular to merge axis)
    #   - Adjacent: gap on the merge axis is within tolerance
    #   - Obstacle-free: merged bbox doesn't contain non-selected geometry
    # Merge the pair that produces the smallest combined bbox volume,
    # then repeat until no more merges are possible for that axis.

    def _axes_overlap(a: List[GeometryItem], b: List[GeometryItem],
                      axes: Tuple[int, ...]) -> bool:
        """Check if two groups overlap on the specified axes.

        Special case: if both groups have zero or near-zero extent on an
        axis (e.g. thin sources with z size = 0) and are at the same
        position on that axis, they are considered overlapping. This
        prevents zero-thickness items from failing collinearity checks.
        """
        lo_a, hi_a = _compute_bbox(a)
        lo_b, hi_b = _compute_bbox(b)
        for axis in axes:
            overlap = min(hi_a[axis], hi_b[axis]) - max(lo_a[axis], lo_b[axis])
            if overlap > tol:
                continue
            # Zero-extent special case: both groups are degenerate on this
            # axis and at the same position → treat as overlapping
            extent_a = hi_a[axis] - lo_a[axis]
            extent_b = hi_b[axis] - lo_b[axis]
            if extent_a <= tol and extent_b <= tol and abs(lo_a[axis] - lo_b[axis]) <= tol:
                continue
            return False
        return True

    def _axes_range_match(a: List[GeometryItem], b: List[GeometryItem],
                          axes: Tuple[int, ...]) -> bool:
        """Check if two groups have matching range on the given axes.

        For merging along axis A, perpendicular axes must have aligned
        ranges so the merged bbox doesn't cover empty space.  E.g. a
        2-wide horizontal strip should not merge with a 1-wide item
        below it, but two 2-wide strips can merge.
        """
        lo_a, hi_a = _compute_bbox(a)
        lo_b, hi_b = _compute_bbox(b)
        for axis in axes:
            extent_a = hi_a[axis] - lo_a[axis]
            extent_b = hi_b[axis] - lo_b[axis]
            # Both zero-extent: positions must match
            if extent_a <= tol and extent_b <= tol:
                if abs(lo_a[axis] - lo_b[axis]) > tol:
                    return False
                continue
            # At least one has real extent: ranges must match
            if abs(lo_a[axis] - lo_b[axis]) > tol or abs(hi_a[axis] - hi_b[axis]) > tol:
                return False
        return True

    def _axis_gap(a: List[GeometryItem], b: List[GeometryItem],
                  axis: int) -> float:
        """Compute the gap between two groups along a given axis.

        Positive = separated, zero = touching, negative = overlapping.
        """
        lo_a, hi_a = _compute_bbox(a)
        lo_b, hi_b = _compute_bbox(b)
        return max(lo_a[axis], lo_b[axis]) - min(hi_a[axis], hi_b[axis])

    # Process each axis: x(0), y(1), z(2)
    # No gap threshold — rely purely on collinearity + obstacle check.
    # If two regions are collinear and the merged bbox has no obstacles,
    # they can merge regardless of the gap between them. This handles
    # the common case of small sources spread apart on a PCB where the
    # gap far exceeds the source size.
    for merge_axis in range(3):
        # Perpendicular axes: the 2 axes that must overlap for collinearity
        perp_axes = tuple(a for a in range(3) if a != merge_axis)

        # Repeatedly find and apply the best valid merge for this axis
        changed = True
        while changed:
            changed = False
            best = None  # (merged_volume, i, j)

            for i in range(len(groups)):
                for j in range(i + 1, len(groups)):
                    # Must be collinear: overlap on both perpendicular axes
                    if not _axes_overlap(groups[i], groups[j], perp_axes):
                        continue

                    # Must have matching range on perpendicular axes to
                    # avoid creating regions that cover empty space.
                    if not _axes_range_match(groups[i], groups[j], perp_axes):
                        continue

                    # Volume efficiency: merged bbox shouldn't be mostly empty.
                    # Per-axis ratio: (extent_a + extent_b) / extent_merged,
                    # skip axes where both are zero-thickness. Product gives
                    # overall fill ratio.
                    lo_a, hi_a = _compute_bbox(groups[i])
                    lo_b, hi_b = _compute_bbox(groups[j])
                    efficiency = 1.0
                    for ax in range(3):
                        ext_a = hi_a[ax] - lo_a[ax]
                        ext_b = hi_b[ax] - lo_b[ax]
                        ext_m = max(hi_a[ax], hi_b[ax]) - min(lo_a[ax], lo_b[ax])
                        if ext_m <= tol:
                            continue  # zero-extent axis, irrelevant
                        efficiency *= min((ext_a + ext_b) / ext_m, 1.0)
                    if efficiency < 0.2:
                        continue

                    # Must be obstacle-free (obstacle check prevents
                    # merging regions that would wrap non-selected geometry)
                    merged = groups[i] + groups[j]
                    if not ignore_obstacles:
                        blocked_by = _has_obstacles(merged, return_blocker=True)
                        if blocked_by:
                            n_i = groups[i][0].name
                            n_j = groups[j][0].name
                            print(
                                f"[D] {'xyz'[merge_axis]}: {n_i}+{n_j} <- {blocked_by}",
                                file=sys.stderr,
                            )
                            continue

                    # Pick the merge that produces the smallest volume
                    vol = _bbox_volume(merged)
                    if best is None or vol < best[0]:
                        best = (vol, i, j)

            if best is not None:
                _, i, j = best
                groups[i] = groups[i] + groups[j]
                groups.pop(j)
                changed = True

    print(f"[D] {len(groups)} regions", file=sys.stderr)
    return groups


def _target_geometry(root_geometry: ET.Element, region_cfg: Dict) -> Tuple[ET.Element, Vector3]:
    parent_assembly = region_cfg.get("parent_assembly")
    if not parent_assembly:
        return root_geometry, (0.0, 0.0, 0.0)

    assembly_elem = _find_assembly_element(root_geometry, parent_assembly)
    if assembly_elem is None:
        raise ValueError(f"parent_assembly not found: {parent_assembly}")

    geometry_elem = _ensure_geometry_container(assembly_elem)

    # Assembly positions are local; convert global region position to parent-local.
    global_offset = (0.0, 0.0, 0.0)
    for item in _iter_all_geometry_items(root_geometry):
        if item.element is assembly_elem:
            global_offset = item.global_position
            break

    return geometry_elem, global_offset


def add_regions(root: ET.Element, config: Dict) -> ET.Element:
    geometry = _find_root_geometry(root)
    regions = config.get("regions", [])
    grid_constraints = config.get("grid_constraints", [])
    object_constraints = config.get("object_constraints", [])
    if not regions and not grid_constraints and not object_constraints:
        raise ValueError("JSON config must contain 'regions', 'grid_constraints', and/or 'object_constraints'")

    if grid_constraints:
        attributes = _find_or_create_attributes(root)
        grid_constraints_elem = _find_or_create_grid_constraints(attributes)
        for constraint_cfg in grid_constraints:
            _upsert_grid_constraint(grid_constraints_elem, constraint_cfg)

    _apply_object_constraints(geometry, config)

    for region_cfg in regions:
        name = region_cfg.get("name")
        if not name:
            raise ValueError("Each region requires a name")

        bbox_cfg = region_cfg.get("bbox_from")
        split = bbox_cfg.get("split_regions", False) if bbox_cfg else False

        if split:
            # Decompose into multiple rectangular regions avoiding obstacles
            include_names = list(bbox_cfg.get("include_names", []))
            include_patterns = list(bbox_cfg.get("include_patterns", []))
            include_tags = list(bbox_cfg.get("include_tags", []))
            scope_assembly = bbox_cfg.get("scope_assembly")
            matches = _match_items(geometry, include_names, include_patterns, include_tags, scope_assembly)
            if not matches:
                raise ValueError(
                    f"bbox_from target did not match any geometry for region '{name}': "
                    f"names={include_names}, patterns={include_patterns}, tags={include_tags}"
                )

            # Filter out items without usable size data
            usable = [m for m in matches if m.global_size is not None]
            skipped = [m for m in matches if m.global_size is None]
            if skipped:
                print(
                    f"Warning: region '{name}' skipped {len(skipped)} item(s) without size: "
                    f"{[f'{m.tag}/{m.name}' for m in skipped]}",
                    file=sys.stderr,
                )
            if not usable:
                raise ValueError(
                    f"bbox_from for region '{name}' matched {len(matches)} item(s) but none have size data. "
                    f"Matched: {[f'{m.tag}/{m.name}' for m in matches]}"
                )

            # Compute obstacles: all geometry items not in the selected set
            usable_set = {id(m.element) for m in usable}
            obstacles = [
                item for item in _iter_all_geometry_items(geometry)
                if item.global_size is not None
                and id(item.element) not in usable_set
            ]

            groups = _decompose_selected_items(
                usable, geometry,
                min_volume_reduction=bbox_cfg.get("min_volume_reduction", 0.2),
                obstacles=obstacles,
                ignore_obstacles=bbox_cfg.get("ignore_obstacles", False),
            )
            padding = _normalize_padding(bbox_cfg.get("padding", 0.0))

            for i, group in enumerate(groups):
                if not group:
                    continue
                sub_name = f"{name}_{i+1}" if len(groups) > 1 else name
                lower, upper = _compute_bbox(group)
                global_position = tuple(lower[j] - padding[j] for j in range(3))
                size = tuple((upper[j] - lower[j]) + (2.0 * padding[j]) for j in range(3))
                target_geometry_elem, parent_offset = _target_geometry(geometry, region_cfg)
                local_position = _vector_sub(global_position, parent_offset)
                region_elem = _build_region_element(sub_name, local_position, size, region_cfg)
                target_geometry_elem.append(region_elem)
        else:
            global_position, size = _resolve_region_geometry(geometry, region_cfg)
            target_geometry_elem, parent_offset = _target_geometry(geometry, region_cfg)
            local_position = _vector_sub(global_position, parent_offset)
            region_elem = _build_region_element(name, local_position, size, region_cfg)
            target_geometry_elem.append(region_elem)

    return root


def indent_xml(elem: ET.Element, level: int = 0) -> None:
    indent = "\n" + ("    " * level)
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = indent + "    "
        for child in elem:
            indent_xml(child, level + 1)
        if not elem[-1].tail or not elem[-1].tail.strip():
            elem[-1].tail = indent
    if level and (not elem.tail or not elem.tail.strip()):
        elem.tail = indent


def load_json(path: Path) -> Dict:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


# ============================================================================
# Excel 配置读取
# ============================================================================

def _split_list(value) -> List[str]:
    """将逗号分隔的字符串拆分为列表，非字符串原样返回"""
    if value is None:
        return []
    if isinstance(value, str):
        return [s.strip() for s in value.split(",") if s.strip()]
    if isinstance(value, (list, tuple)):
        return [str(v).strip() for v in value if str(v).strip()]
    return [str(value)]


def _parse_bool(value) -> Optional[bool]:
    """解析布尔值"""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in ("true", "1", "yes"):
        return True
    if s in ("false", "0", "no", ""):
        return False
    return None


def _parse_number(value):
    """解析数值，失败则返回 None"""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _row_dict(ws, row: int, headers: List[str]) -> Dict:
    """将一行数据转为 {header: value} dict，跳过全空行"""
    values = []
    for col_idx in range(1, len(headers) + 1):
        values.append(ws.cell(row=row, column=col_idx).value)
    if all(v is None for v in values):
        return {}
    return dict(zip(headers, values))


def _read_grid_constraints_sheet(ws) -> List[Dict]:
    """读取 grid_constraints sheet"""
    headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    if not headers or not headers[0]:
        return []

    results = []
    for row in range(2, ws.max_row + 1):
        rd = _row_dict(ws, row, headers)
        if not rd:
            continue
        name = str(rd.get("name", "")).strip()
        if not name:
            continue

        entry: Dict = {"name": name}

        v = _parse_bool(rd.get("enable_min_cell_size"))
        if v is not None:
            entry["enable_min_cell_size"] = v

        v = _parse_number(rd.get("min_cell_size"))
        if v is not None:
            entry["min_cell_size"] = v

        v = rd.get("number_cells_control")
        if v and str(v).strip():
            entry["number_cells_control"] = str(v).strip()

        v = _parse_number(rd.get("min_number"))
        if v is not None:
            entry["min_number"] = int(v)

        # high_inflation 子对象
        hi = {}
        for key, json_key in [
            ("high_inflation_type", "inflation_type"),
            ("high_inflation_size", "inflation_size"),
            ("high_inflation_number_cells_control", "number_cells_control"),
            ("high_inflation_min_number", "min_number"),
        ]:
            raw = rd.get(key)
            if raw is None or str(raw).strip() == "":
                continue
            if json_key == "min_number":
                hi[json_key] = int(float(raw))
            elif json_key == "inflation_size":
                hi[json_key] = float(raw)
            else:
                hi[json_key] = str(raw).strip()
        if hi:
            entry["high_inflation"] = hi

        results.append(entry)
    return results


def _read_object_constraints_sheet(ws) -> List[Dict]:
    """读取 object_constraints sheet"""
    headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    if not headers or not headers[0]:
        return []

    results = []
    for row in range(2, ws.max_row + 1):
        rd = _row_dict(ws, row, headers)
        if not rd:
            continue

        entry: Dict = {}

        names = _split_list(rd.get("target_names"))
        if names:
            entry["target_names"] = names

        patterns = _split_list(rd.get("target_patterns"))
        if patterns:
            entry["target_patterns"] = patterns

        tags = _split_list(rd.get("target_tags"))
        if tags:
            entry["target_tags"] = tags

        v = rd.get("scope_assembly")
        if v and str(v).strip():
            entry["scope_assembly"] = str(v).strip()

        for col_tag in ("x_grid_constraint", "y_grid_constraint", "z_grid_constraint", "all_grid_constraint"):
            v = rd.get(col_tag)
            if v and str(v).strip():
                entry[col_tag] = str(v).strip()

        v = _parse_bool(rd.get("localized_grid"))
        if v is not None:
            entry["localized_grid"] = v

        if entry:
            results.append(entry)
    return results


def _read_regions_sheet(ws) -> List[Dict]:
    """读取 regions sheet（支持 explicit 和 bbox 两种模式）"""
    headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    if not headers or not headers[0]:
        return []

    results = []
    for row in range(2, ws.max_row + 1):
        rd = _row_dict(ws, row, headers)
        if not rd:
            continue

        name = str(rd.get("name", "")).strip()
        if not name:
            continue

        entry: Dict = {"name": name}

        # parent_assembly
        v = rd.get("parent_assembly")
        if v and str(v).strip():
            entry["parent_assembly"] = str(v).strip()

        # 判断 bbox 模式还是 explicit 模式
        bbox_names = _split_list(rd.get("bbox_include_names"))
        bbox_patterns = _split_list(rd.get("bbox_include_patterns"))

        if bbox_names or bbox_patterns:
            # bbox 模式
            bbox: Dict = {}
            if bbox_names:
                bbox["include_names"] = bbox_names
            if bbox_patterns:
                bbox["include_patterns"] = bbox_patterns
            tags = _split_list(rd.get("bbox_include_tags"))
            if tags:
                bbox["include_tags"] = tags
            scope = rd.get("bbox_scope_assembly")
            if scope and str(scope).strip():
                bbox["scope_assembly"] = str(scope).strip()

            # padding: 单值或逗号分隔 3 值
            raw_padding = rd.get("bbox_padding")
            if raw_padding is not None and str(raw_padding).strip():
                parts = [s.strip() for s in str(raw_padding).split(",") if s.strip()]
                if len(parts) == 1:
                    bbox["padding"] = float(parts[0])
                elif len(parts) == 3:
                    bbox["padding"] = [float(p) for p in parts]

            # split_regions: auto-split to minimize region volume
            v = _parse_bool(rd.get("split_regions"))
            if v:
                bbox["split_regions"] = True
                mvr = _parse_number(rd.get("split_min_reduction"))
                if mvr is not None:
                    bbox["min_volume_reduction"] = float(mvr)

            entry["bbox_from"] = bbox
        else:
            # explicit 模式
            pos_x = _parse_number(rd.get("position_x"))
            pos_y = _parse_number(rd.get("position_y"))
            pos_z = _parse_number(rd.get("position_z"))
            if pos_x is not None or pos_y is not None or pos_z is not None:
                entry["position"] = [pos_x or 0.0, pos_y or 0.0, pos_z or 0.0]

            size_x = _parse_number(rd.get("size_x"))
            size_y = _parse_number(rd.get("size_y"))
            size_z = _parse_number(rd.get("size_z"))
            if size_x is not None or size_y is not None or size_z is not None:
                entry["size"] = [size_x or 0.0, size_y or 0.0, size_z or 0.0]

        # 公共字段
        for col_tag in ("x_grid_constraint", "y_grid_constraint", "z_grid_constraint", "all_grid_constraint"):
            v = rd.get(col_tag)
            if v and str(v).strip():
                entry[col_tag] = str(v).strip()

        v = _parse_bool(rd.get("active"))
        if v is not None:
            entry["active"] = v

        v = _parse_bool(rd.get("hidden"))
        if v is not None:
            entry["hidden"] = v

        v = _parse_bool(rd.get("localized_grid"))
        if v is not None:
            entry["localized_grid"] = v

        results.append(entry)
    return results


def load_excel(path: Path) -> Dict:
    """从 Excel 文件读取配置，返回与 JSON 格式相同的 dict"""
    try:
        from openpyxl import load_workbook as _load_wb
    except ImportError:
        raise ImportError("需要安装 openpyxl 来读取 Excel 配置: pip install openpyxl")

    wb = _load_wb(str(path), read_only=True, data_only=True)
    config: Dict = {}

    for sheet_name in ("grid_constraints", "object_constraints", "regions"):
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
        if ws is None:
            continue
        if sheet_name == "grid_constraints":
            data = _read_grid_constraints_sheet(ws)
            if data:
                config["grid_constraints"] = data
        elif sheet_name == "object_constraints":
            data = _read_object_constraints_sheet(ws)
            if data:
                config["object_constraints"] = data
        elif sheet_name == "regions":
            data = _read_regions_sheet(ws)
            if data:
                config["regions"] = data

    wb.close()
    return config


def load_config(path: Path) -> Dict:
    """自动识别 JSON / Excel 配置文件"""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls", ".xlsm"):
        return load_excel(path)
    return load_json(path)


def create_template_excel(output_path: str) -> None:
    """创建 volume regions 配置的 Excel 模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center')

    def _write_headers(ws, headers):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align

    def _set_col_widths(ws, width=18):
        for idx in range(1, ws.max_column + 1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    # ── Sheet 1: grid_constraints ──
    ws1 = wb.active
    ws1.title = "grid_constraints"
    _write_headers(ws1, [
        "name", "enable_min_cell_size", "min_cell_size", "number_cells_control", "min_number",
        "high_inflation_type", "high_inflation_size",
        "high_inflation_number_cells_control", "high_inflation_min_number",
    ])
    example1 = [
        ["Grid Constraint 1", "true", 0.001, "min_number", 43, "size", 0.005, "min_number", 23],
    ]
    for r, row_data in enumerate(example1, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)
    _set_col_widths(ws1)

    # ── Sheet 2: object_constraints ──
    ws2 = wb.create_sheet("object_constraints")
    _write_headers(ws2, [
        "target_names", "target_patterns", "target_tags", "scope_assembly",
        "x_grid_constraint", "y_grid_constraint", "z_grid_constraint", "all_grid_constraint",
        "localized_grid",
    ])
    example2 = [
        ["PCB", "", "cuboid", "", "", "", "", "Grid Constraint 1", "false"],
    ]
    for r, row_data in enumerate(example2, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)
    _set_col_widths(ws2)

    # ── Sheet 3: regions ──
    ws3 = wb.create_sheet("regions")
    _write_headers(ws3, [
        "name", "parent_assembly",
        "position_x", "position_y", "position_z", "size_x", "size_y", "size_z",
        "bbox_include_names", "bbox_include_patterns", "bbox_include_tags",
        "bbox_scope_assembly", "bbox_padding", "split_regions", "split_min_reduction",
        "active", "hidden", "localized_grid",
        "x_grid_constraint", "y_grid_constraint", "z_grid_constraint", "all_grid_constraint",
    ])
    example3 = [
        # explicit 模式示例
        ["Explicit Volume Region", "", -0.01, -0.01, -0.002, 0.12, 0.08, 0.01,
         "", "", "", "", "",
         "", "", "true", "Grid Constraint 1", "", "", ""],
        # bbox 模式示例
        ["BBox Region Around PCB", "DemoBoard_Assembly",
         "", "", "", "", "", "",
         "PCB", "U*", "cuboid", "", "0.001,0.001,0.0005", "", "",
         "", "", "true", "", "", "", "Grid Constraint 1"],
        # bbox + split_regions 模式示例
        ["Split Region L-Shape", "",
         "", "", "", "", "", "",
         "C1,C2,C3,C4,C7", "", "", "", "0.05", "true", "",
         "", "true", "", "", "", "Grid Constraint 1"],
        # bbox + split_regions + 自定义阈值 (10%)
        ["Split Region Custom", "",
         "", "", "", "", "", "",
         "S1,S2,S3", "", "", "", "0.05", "true", "0.1",
         "", "true", "", "", "", "Grid Constraint 1"],
    ]
    for r, row_data in enumerate(example3, 2):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)
    _set_col_widths(ws3)

    wb.save(output_path)
    print(f"[OK] 模板已创建: {output_path}")


# ============================================================================
# CLI 接口
# ============================================================================

def main() -> int:
    parser = argparse.ArgumentParser(description="Add volume regions to FloXML using JSON or Excel config")
    parser.add_argument("input", nargs="?", help="Input FloXML file")
    parser.add_argument("--config", help="Region config file (.json or .xlsx)")
    parser.add_argument("-o", "--output", help="Output FloXML file")
    parser.add_argument("--create-template", metavar="PATH",
                        help="Create an example Excel template at PATH")
    args = parser.parse_args()

    if args.create_template:
        create_template_excel(args.create_template)
        return 0

    if not args.input or not args.config:
        parser.error("input and --config are required (unless using --create-template)")

    input_path = Path(args.input)
    config_path = Path(args.config)
    output_path = Path(args.output) if args.output else input_path.with_name(f"{input_path.stem}_with_regions{input_path.suffix}")

    tree = ET.parse(input_path)
    root = tree.getroot()
    config = load_config(config_path)
    root = add_regions(root, config)
    indent_xml(root)
    tree.write(output_path, encoding="utf-8", xml_declaration=True)

    print(
        f"[OK] Added {len(config.get('regions', []))} region(s), "
        f"upserted {len(config.get('grid_constraints', []))} grid constraint(s), "
        f"applied {len(config.get('object_constraints', []))} object constraint rule(s): {output_path}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
