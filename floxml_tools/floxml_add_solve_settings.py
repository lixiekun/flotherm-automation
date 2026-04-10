#!/usr/bin/env python3
"""
往 FloXML 里塞求解设置和瞬态设置，支持 JSON 和 Excel 两种配置方式。

能配的东西：
  <solve> 下面：
    - overall_control   迭代次数、求解器类型、收敛判据那些
    - variable_controls 每个变量（速度、温度）的求解调参
    - solver_controls   压力、温度的线性松弛之类的

  <model> 下面：
    - modeling          求解类型（流动+传热/纯导热等）、辐射开关
    - turbulence        层流/湍流
    - gravity           重力
    - global            基准压力、环境温度
    - initial_variables 初始值
    - transient         瞬态时间设置（时间段、保存时刻、时间片）

用法：
  python -m floxml_tools.floxml_add_solve_settings model.xml --config solve.json -o out.xml
  python -m floxml_tools.floxml_add_solve_settings model.xml --config solve.xlsx -o out.xml
  python -m floxml_tools.floxml_add_solve_settings --create-template template.xlsx
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Dict, List, Optional
import xml.etree.ElementTree as ET


# ============================================================================
# XML 操作的小工具
# ============================================================================

def _set_text(parent: ET.Element, tag: str, text: str) -> ET.Element:
    child = parent.find(tag)
    if child is None:
        child = ET.SubElement(parent, tag)
    child.text = text
    return child


def _find_or_create(parent: ET.Element, tag: str) -> ET.Element:
    child = parent.find(tag)
    if child is None:
        child = ET.SubElement(parent, tag)
    return child


def _find_or_create_solve(root: ET.Element) -> ET.Element:
    """找 <solve>，没有就建一个，尽量插到 <grid> 前面"""
    solve = root.find("solve")
    if solve is not None:
        return solve

    solve = ET.Element("solve")
    insert_before = None
    for idx, child in enumerate(list(root)):
        if child.tag in ("grid", "solution_domain"):
            insert_before = idx
            break
    if insert_before is not None:
        root.insert(insert_before, solve)
    else:
        root.append(solve)
    return solve


# ============================================================================
# <solve> 下面几个子节点
# ============================================================================

BOOL_FIELDS = (
    "monitor_convergence",
    "active_plate_conduction",
    "use_double_precision",
    "multi_grid_damping",
    "network_assembly_block_correction",
    "freeze_flow",
    "store_error_field",
)

INT_FIELDS = (
    "outer_iterations",
)

FLOAT_FIELDS = (
    "fan_relaxation",
    "estimated_free_convection_velocity",
)

STRING_FIELDS = (
    "solver_option",
    "error_field_variable",
)

CONVERGENCE_FIELDS = {
    "required_accuracy": float,
    "num_iterations": int,
    "residual_threshold": int,
}


def _format_value(value, python_type):
    """把配置值转成 XML 里要写的字符串，顺便处理格式"""
    if python_type is bool:
        if isinstance(value, bool):
            return "true" if value else "false"
        s = str(value).strip().lower()
        return "true" if s in ("true", "1", "yes") else "false"
    if python_type is int:
        return str(int(float(value)))
    if python_type is float:
        fval = float(value)
        if fval == int(fval):
            return str(int(fval))
        # 用定点格式避免科学计数法，去掉尾巴上的0
        return f"{fval:.15f}".rstrip('0').rstrip('.')
    return str(value)


def _apply_overall_control(solve: ET.Element, cfg: Dict) -> None:
    """写 <overall_control>"""
    oc = _find_or_create(solve, "overall_control")

    for field in BOOL_FIELDS:
        if field in cfg:
            _set_text(oc, field, _format_value(cfg[field], bool))

    for field in INT_FIELDS:
        if field in cfg:
            _set_text(oc, field, _format_value(cfg[field], int))

    for field in FLOAT_FIELDS:
        if field in cfg:
            _set_text(oc, field, _format_value(cfg[field], float))

    for field in STRING_FIELDS:
        if field in cfg:
            _set_text(oc, field, _format_value(cfg[field], str))

    # convergence_values 子节点
    cv_cfg = cfg.get("convergence_values")
    if cv_cfg and isinstance(cv_cfg, dict):
        cv = _find_or_create(oc, "convergence_values")
        for fname, ftype in CONVERGENCE_FIELDS.items():
            if fname in cv_cfg:
                _set_text(cv, fname, _format_value(cv_cfg[fname], ftype))

    # 监控点瞬态终止条件
    mpttc_cfg = cfg.get("monitor_point_transient_termination_criteria")
    if mpttc_cfg and isinstance(mpttc_cfg, dict):
        mpttc = _find_or_create(oc, "monitor_point_transient_termination_criteria")
        if "use_monitor_point_transient_termination_criteria" in mpttc_cfg:
            _set_text(mpttc, "use_monitor_point_transient_termination_criteria",
                      _format_value(mpttc_cfg["use_monitor_point_transient_termination_criteria"], bool))
        tmp_cfg = mpttc_cfg.get("terminating_monitor_points")
        if tmp_cfg and isinstance(tmp_cfg, list):
            tmp_parent = _find_or_create(mpttc, "terminating_monitor_points")
            for old in tmp_parent.findall("terminating_monitor_point"):
                tmp_parent.remove(old)
            for mp_cfg in tmp_cfg:
                mp = ET.SubElement(tmp_parent, "terminating_monitor_point")
                if "monitor_point" in mp_cfg:
                    _set_text(mp, "monitor_point", str(mp_cfg["monitor_point"]))
                if "temperature" in mp_cfg:
                    _set_text(mp, "temperature", _format_value(mp_cfg["temperature"], float))


def _apply_variable_controls(solve: ET.Element, controls: List[Dict]) -> None:
    """写 <variable_controls>，每个变量一条"""
    vc_parent = _find_or_create(solve, "variable_controls")

    for ctrl_cfg in controls:
        variable = ctrl_cfg.get("variable")
        if not variable:
            continue

        # 找已有的同名变量，没有就新建
        entry = None
        for child in vc_parent.findall("variable_control"):
            v_elem = child.find("variable")
            if v_elem is not None and (v_elem.text or "").strip() == variable:
                entry = child
                break
        if entry is None:
            entry = ET.SubElement(vc_parent, "variable_control")

        _set_text(entry, "variable", variable)

        for field, ftype in (
            ("false_time_step", str),
            ("false_time_step_user_value", float),
            ("false_time_step_damping_auto_multiplier", float),
            ("terminal_residual", str),
            ("terminal_residual_auto_multiplier", float),
            ("terminal_residual_user_value", float),
            ("inner_iterations", int),
        ):
            if field in ctrl_cfg and ctrl_cfg[field] is not None:
                _set_text(entry, field, _format_value(ctrl_cfg[field], ftype))


def _apply_solver_controls(solve: ET.Element, controls: List[Dict]) -> None:
    """写 <solver_controls>，线性松弛那些"""
    sc_parent = _find_or_create(solve, "solver_controls")

    for ctrl_cfg in controls:
        variable = ctrl_cfg.get("variable")
        if not variable:
            continue

        entry = None
        for child in sc_parent.findall("solver_control"):
            v_elem = child.find("variable")
            if v_elem is not None and (v_elem.text or "").strip() == variable:
                entry = child
                break
        if entry is None:
            entry = ET.SubElement(sc_parent, "solver_control")

        _set_text(entry, "variable", variable)

        for field, ftype in (
            ("linear_relaxation", float),
            ("error_compute_frequency", int),
        ):
            if field in ctrl_cfg:
                _set_text(entry, field, _format_value(ctrl_cfg[field], ftype))


# ============================================================================
# 瞬态设置 <model>/<transient>
# ============================================================================

OVERALL_TRANSIENT_FIELDS = {
    "start_time": float,
    "end_time": float,
    "duration": float,
    "keypoint_tolerance": float,
}

TIME_PATCH_FIELDS = {
    "start_time": float,
    "end_time": float,
    "step_control": str,
    "additional_steps": int,
    "minimum_number": int,
    "maximum_size": float,
    "step_distribution": str,
    "distribution_index": float,
}


def _find_or_create_model(root: ET.Element) -> ET.Element:
    model = root.find("model")
    if model is None:
        model = ET.SubElement(root, "model")
    return model


def _find_or_create_modeling(model: ET.Element) -> ET.Element:
    modeling = model.find("modeling")
    if modeling is None:
        modeling = ET.SubElement(model, "modeling")
    return modeling


def _apply_transient_toggle(model: ET.Element, enabled: bool) -> None:
    """把 <modeling>/<transient> 设成 true/false"""
    modeling = _find_or_create_modeling(model)
    _set_text(modeling, "transient", "true" if enabled else "false")


def _apply_overall_transient(model: ET.Element, cfg: Dict) -> None:
    """写整体瞬态参数（起始时间、终止时间、duration、keypoint容差）"""
    transient = _find_or_create(model, "transient")
    ot = _find_or_create(transient, "overall_transient")

    for fname, ftype in OVERALL_TRANSIENT_FIELDS.items():
        if fname in cfg:
            _set_text(ot, fname, _format_value(cfg[fname], ftype))


def _apply_save_times(model: ET.Element, times: List[float]) -> None:
    """写保存时刻，先清掉旧的再写新的"""
    transient = _find_or_create(model, "transient")
    st_parent = _find_or_create(transient, "transient_save_times")

    for old in st_parent.findall("save_time"):
        st_parent.remove(old)

    for t in times:
        st = ET.SubElement(st_parent, "save_time")
        st.text = _format_value(t, float)


def _apply_time_patches(model: ET.Element, patches: List[Dict]) -> None:
    """写时间片，按 name 匹配已有的"""
    transient = _find_or_create(model, "transient")
    tp_parent = _find_or_create(transient, "time_patches")

    for patch_cfg in patches:
        name = patch_cfg.get("name")
        if not name:
            continue

        entry = None
        for child in tp_parent.findall("time_patch"):
            n = child.find("name")
            if n is not None and (n.text or "").strip() == name:
                entry = child
                break
        if entry is None:
            entry = ET.SubElement(tp_parent, "time_patch")

        _set_text(entry, "name", name)

        for fname, ftype in TIME_PATCH_FIELDS.items():
            if fname in patch_cfg:
                _set_text(entry, fname, _format_value(patch_cfg[fname], ftype))


def _apply_transient_settings(root: ET.Element, config: Dict) -> None:
    """把 transient 那一坨都写上"""
    transient_cfg = config.get("transient")
    if not transient_cfg:
        return

    model = _find_or_create_model(root)

    # 先把 modeling 里的 transient 开关打开
    _apply_transient_toggle(model, True)

    if "overall_transient" in transient_cfg:
        _apply_overall_transient(model, transient_cfg["overall_transient"])

    if "save_times" in transient_cfg:
        _apply_save_times(model, transient_cfg["save_times"])

    if "time_patches" in transient_cfg:
        _apply_time_patches(model, transient_cfg["time_patches"])


# ============================================================================
# <model> 下面其他几个子节点
# ============================================================================

MODELING_STRING_FIELDS = (
    "solution",       # flow_heat | flow_only | conduction_only
    "dimensionality", # 2d | 3d
    "radiation",      # off | on | high_accuracy
)

MODELING_BOOL_FIELDS = (
    "joule_heating",
    "store_mass_flux",
    "store_heat_flux",
    "store_surface_temp",
    "store_grad_t",
    "store_bn_sc",
    "store_power_density",
    "store_mean_radiant_temperature",
    "compute_capture_index",
    "user_defined_subgroups",
    "store_lma",
    "store_visibility",
)

MODELING_FLOAT_FIELDS = (
    "recirculation_ratio",
)


def _apply_modeling(model: ET.Element, cfg: Dict) -> None:
    """写 <modeling>，求解类型、辐射、各种存储开关"""
    modeling = _find_or_create_modeling(model)

    for field in MODELING_STRING_FIELDS:
        if field in cfg:
            _set_text(modeling, field, str(cfg[field]))

    for field in MODELING_BOOL_FIELDS:
        if field in cfg:
            _set_text(modeling, field, _format_value(cfg[field], bool))

    for field in MODELING_FLOAT_FIELDS:
        if field in cfg:
            _set_text(modeling, field, _format_value(cfg[field], float))


def _apply_turbulence(model: ET.Element, cfg: Dict) -> None:
    """写 <turbulence>，层流/湍流模型"""
    turb = _find_or_create(model, "turbulence")

    for field in ("type", "turbulence_type"):
        if field in cfg:
            _set_text(turb, field, str(cfg[field]))

    if "lvel_k_epsilon_stratification" in cfg:
        _set_text(turb, "lvel_k_epsilon_stratification",
                  _format_value(cfg["lvel_k_epsilon_stratification"], bool))
    if "capped_lvel_multiplier" in cfg:
        _set_text(turb, "capped_lvel_multiplier",
                  _format_value(cfg["capped_lvel_multiplier"], float))

    # revised_algebraic 里有个 velocity 和 length
    ra_cfg = cfg.get("revised_algebraic")
    if ra_cfg and isinstance(ra_cfg, dict):
        ra = _find_or_create(turb, "revised_algebraic")
        if "velocity" in ra_cfg:
            _set_text(ra, "velocity", _format_value(ra_cfg["velocity"], float))
        if "length" in ra_cfg:
            _set_text(ra, "length", _format_value(ra_cfg["length"], float))


def _apply_gravity(model: ET.Element, cfg: Dict) -> None:
    """写 <gravity>"""
    grav = _find_or_create(model, "gravity")

    for field in ("type", "normal_direction", "value_type"):
        if field in cfg:
            _set_text(grav, field, str(cfg[field]))

    if "gravity_value" in cfg:
        _set_text(grav, "gravity_value", _format_value(cfg["gravity_value"], float))

    # 角度方向是个三元组
    ad_cfg = cfg.get("angled_direction")
    if ad_cfg and isinstance(ad_cfg, dict):
        ad = _find_or_create(grav, "angled_direction")
        for axis in ("x", "y", "z"):
            if axis in ad_cfg:
                _set_text(ad, axis, _format_value(ad_cfg[axis], float))


GLOBAL_FIELDS = {
    "datum_pressure": float,
    "radiant_temperature": float,
    "ambient_temperature": float,
    "ambient_transient": str,
    "radiant_transient": str,
}


def _apply_global(model: ET.Element, cfg: Dict) -> None:
    """写 <global>，基准压力、环境温度那些"""
    g = _find_or_create(model, "global")

    for fname, ftype in GLOBAL_FIELDS.items():
        if fname in cfg:
            _set_text(g, fname, _format_value(cfg[fname], ftype))

    # 浓度字段 concentration_1 ~ concentration_15，一般用不到
    for i in range(1, 16):
        key = f"concentration_{i}"
        if key in cfg:
            _set_text(g, key, _format_value(cfg[key], float))


_INITIAL_VAR_NAMES = [
    "pressure", "x_velocity", "y_velocity", "z_velocity", "temperature",
    "ke_turb", "diss_turb", "turb_vis", "density", "potential",
] + [f"concentration_{i}" for i in range(1, 16)]


def _apply_initial_variables(model: ET.Element, cfg: Dict) -> None:
    """写 <initial_variables>，每个变量可以设 type 和 value"""
    iv = _find_or_create(model, "initial_variables")

    for field in ("initial_solution_name", "initial_solution_id"):
        if field in cfg:
            _set_text(iv, field, str(cfg[field]))

    if "use_initial_for_all" in cfg:
        _set_text(iv, "use_initial_for_all", _format_value(cfg["use_initial_for_all"], bool))

    for var_name in _INITIAL_VAR_NAMES:
        var_cfg = cfg.get(var_name)
        if var_cfg and isinstance(var_cfg, dict):
            var_elem = _find_or_create(iv, var_name)
            if "type" in var_cfg:
                _set_text(var_elem, "type", str(var_cfg["type"]))
            if "value" in var_cfg:
                _set_text(var_elem, "value", _format_value(var_cfg["value"], float))


# ============================================================================
# 总入口
# ============================================================================

def apply_solve_settings(root: ET.Element, config: Dict) -> ET.Element:
    """主函数，把 JSON/Excel 里配的东西全部写到 FloXML 上"""
    solve = _find_or_create_solve(root)

    if "overall_control" in config:
        _apply_overall_control(solve, config["overall_control"])

    if "variable_controls" in config:
        _apply_variable_controls(solve, config["variable_controls"])

    if "solver_controls" in config:
        _apply_solver_controls(solve, config["solver_controls"])

    # <model> 下面的设置
    model_keys = ("modeling", "turbulence", "gravity", "global", "initial_variables", "transient")
    if any(k in config for k in model_keys):
        model = _find_or_create_model(root)
        if "modeling" in config:
            _apply_modeling(model, config["modeling"])
        if "turbulence" in config:
            _apply_turbulence(model, config["turbulence"])
        if "gravity" in config:
            _apply_gravity(model, config["gravity"])
        if "global" in config:
            _apply_global(model, config["global"])
        if "initial_variables" in config:
            _apply_initial_variables(model, config["initial_variables"])
        if "transient" in config:
            _apply_transient_settings(root, config)

    return root


# ============================================================================
# 读 JSON 配置
# ============================================================================

def load_json(path: Path) -> Dict:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


# ============================================================================
# 读 Excel 配置
# ============================================================================

def _parse_bool(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in ("true", "1", "yes"):
        return True
    if s in ("false", "0", "no", ""):
        return False
    return None


def _parse_number(value):
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _row_dict(ws, row: int, headers: List[str]) -> Dict:
    """把一行读成 dict"""
    values = [ws.cell(row=row, column=c).value for c in range(1, len(headers) + 1)]
    if all(v is None for v in values):
        return {}
    return dict(zip(headers, values))


def _read_overall_control_sheet(ws) -> Dict:
    """读 overall_control sheet，多行合并成一个 dict"""
    headers = [str(ws.cell(row=1, column=c).value or "").strip()
               for c in range(1, ws.max_column + 1)]
    if not headers or not headers[0]:
        return {}

    result: Dict = {}
    for row in range(2, ws.max_row + 1):
        rd = _row_dict(ws, row, headers)
        if not rd:
            continue

        # 数值和字符串字段
        for col_name, target_name, parser in (
            ("outer_iterations", "outer_iterations", lambda v: int(float(v)) if v is not None else None),
            ("fan_relaxation", "fan_relaxation", _parse_number),
            ("estimated_free_convection_velocity", "estimated_free_convection_velocity", _parse_number),
            ("solver_option", "solver_option", lambda v: str(v).strip() if v and str(v).strip() else None),
            ("error_field_variable", "error_field_variable", lambda v: str(v).strip() if v and str(v).strip() else None),
        ):
            v = rd.get(col_name)
            if v is not None and (not isinstance(v, str) or v.strip()):
                parsed = parser(v)
                if parsed is not None:
                    result[target_name] = parsed

        # 布尔字段
        for col_name in BOOL_FIELDS:
            v = _parse_bool(rd.get(col_name))
            if v is not None:
                result[col_name] = v

        # 收敛判据
        cv: Dict = {}
        for col_name, ftype in CONVERGENCE_FIELDS.items():
            raw = rd.get(col_name)
            if raw is not None and (not isinstance(raw, str) or raw.strip()):
                try:
                    cv[col_name] = ftype(float(raw))
                except (ValueError, TypeError):
                    pass
        if cv:
            result.setdefault("convergence_values", {}).update(cv)

    return {"overall_control": result} if result else {}


def _read_controls_sheet(ws, entry_tag: str) -> List[Dict]:
    """读 variable_controls 或 solver_controls sheet，每行一条"""
    headers = [str(ws.cell(row=1, column=c).value or "").strip()
               for c in range(1, ws.max_column + 1)]
    if not headers or not headers[0]:
        return []

    results: List[Dict] = []
    for row in range(2, ws.max_row + 1):
        rd = _row_dict(ws, row, headers)
        if not rd:
            continue

        variable = rd.get("variable")
        if not variable or not str(variable).strip():
            continue

        entry: Dict = {"variable": str(variable).strip()}

        if entry_tag == "variable_controls":
            for col, key, ftype in (
                ("false_time_step", "false_time_step", str),
                ("false_time_step_user_value", "false_time_step_user_value", float),
                ("false_time_step_damping_auto_multiplier", "false_time_step_damping_auto_multiplier", float),
                ("terminal_residual", "terminal_residual", str),
                ("terminal_residual_auto_multiplier", "terminal_residual_auto_multiplier", float),
                ("terminal_residual_user_value", "terminal_residual_user_value", float),
                ("inner_iterations", "inner_iterations", int),
            ):
                raw = rd.get(col)
                if raw is not None and (not isinstance(raw, str) or raw.strip()):
                    try:
                        entry[key] = ftype(float(raw)) if ftype in (int, float) else ftype(raw)
                    except (ValueError, TypeError):
                        pass

        elif entry_tag == "solver_controls":
            for col, key, ftype in (
                ("linear_relaxation", "linear_relaxation", float),
                ("error_compute_frequency", "error_compute_frequency", int),
            ):
                raw = rd.get(col)
                if raw is not None and (not isinstance(raw, str) or raw.strip()):
                    try:
                        entry[key] = ftype(float(raw))
                    except (ValueError, TypeError):
                        pass

        if len(entry) > 1:
            results.append(entry)
    return results


def _read_transient_overall_sheet(ws) -> Dict:
    """读 transient_overall sheet"""
    headers = [str(ws.cell(row=1, column=c).value or "").strip()
               for c in range(1, ws.max_column + 1)]
    if not headers or not headers[0]:
        return {}

    result: Dict = {}
    for row in range(2, ws.max_row + 1):
        rd = _row_dict(ws, row, headers)
        if not rd:
            continue
        for col_name, ftype in OVERALL_TRANSIENT_FIELDS.items():
            raw = rd.get(col_name)
            if raw is not None and (not isinstance(raw, str) or raw.strip()):
                try:
                    result[col_name] = ftype(float(raw))
                except (ValueError, TypeError):
                    pass
    return result


def _read_transient_save_times_sheet(ws) -> List[float]:
    """读 save_times，一列数值"""
    headers = [str(ws.cell(row=1, column=c).value or "").strip()
               for c in range(1, ws.max_column + 1)]
    if not headers or not headers[0]:
        return []

    times: List[float] = []
    for row in range(2, ws.max_row + 1):
        rd = _row_dict(ws, row, headers)
        if not rd:
            continue
        raw = rd.get("save_time")
        if raw is not None and (not isinstance(raw, str) or raw.strip()):
            try:
                times.append(float(raw))
            except (ValueError, TypeError):
                pass
    return times


def _read_time_patches_sheet(ws) -> List[Dict]:
    """读 time_patches，每行一个时间片"""
    headers = [str(ws.cell(row=1, column=c).value or "").strip()
               for c in range(1, ws.max_column + 1)]
    if not headers or not headers[0]:
        return []

    results: List[Dict] = []
    for row in range(2, ws.max_row + 1):
        rd = _row_dict(ws, row, headers)
        if not rd:
            continue

        name = rd.get("name")
        if not name or not str(name).strip():
            continue

        entry: Dict = {"name": str(name).strip()}
        for col_name, ftype in TIME_PATCH_FIELDS.items():
            raw = rd.get(col_name)
            if raw is not None and (not isinstance(raw, str) or raw.strip()):
                try:
                    entry[col_name] = ftype(float(raw)) if ftype in (int, float) else str(raw).strip()
                except (ValueError, TypeError):
                    pass

        if len(entry) > 1:
            results.append(entry)
    return results


def _read_key_value_sheet(ws, string_fields=(), bool_fields=(), float_fields=()) -> Dict:
    """通用的 key-value sheet 读取，适合 modeling/gravity/global 这种简单结构"""
    headers = [str(ws.cell(row=1, column=c).value or "").strip()
               for c in range(1, ws.max_column + 1)]
    if not headers or not headers[0]:
        return {}

    result: Dict = {}
    for row in range(2, ws.max_row + 1):
        rd = _row_dict(ws, row, headers)
        if not rd:
            continue

        for col_name in string_fields:
            raw = rd.get(col_name)
            if raw is not None and str(raw).strip():
                result[col_name] = str(raw).strip()

        for col_name in bool_fields:
            v = _parse_bool(rd.get(col_name))
            if v is not None:
                result[col_name] = v

        for col_name in float_fields:
            raw = rd.get(col_name)
            if raw is not None and (not isinstance(raw, str) or raw.strip()):
                try:
                    result[col_name] = float(raw)
                except (ValueError, TypeError):
                    pass

    return result


def _read_modeling_sheet(ws) -> Dict:
    data = _read_key_value_sheet(
        ws,
        string_fields=MODELING_STRING_FIELDS,
        bool_fields=MODELING_BOOL_FIELDS,
        float_fields=MODELING_FLOAT_FIELDS,
    )
    return {"modeling": data} if data else {}


def _read_turbulence_sheet(ws) -> Dict:
    data = _read_key_value_sheet(
        ws,
        string_fields=("type", "turbulence_type"),
        bool_fields=("lvel_k_epsilon_stratification",),
        float_fields=("capped_lvel_multiplier",),
    )
    # revised_algebraic 的两个子列
    headers = [str(ws.cell(row=1, column=c).value or "").strip()
               for c in range(1, ws.max_column + 1)]
    for row in range(2, ws.max_row + 1):
        rd = _row_dict(ws, row, headers)
        ra: Dict = {}
        for sub_key, col in (("velocity", "ra_velocity"), ("length", "ra_length")):
            raw = rd.get(col)
            if raw is not None and (not isinstance(raw, str) or raw.strip()):
                try:
                    ra[sub_key] = float(raw)
                except (ValueError, TypeError):
                    pass
        if ra:
            data["revised_algebraic"] = ra
    return {"turbulence": data} if data else {}


def _read_gravity_sheet(ws) -> Dict:
    data = _read_key_value_sheet(
        ws,
        string_fields=("type", "normal_direction", "value_type"),
        float_fields=("gravity_value", "angled_x", "angled_y", "angled_z"),
    )
    # angled_direction 拆出来
    ad: Dict = {}
    for axis, col in (("x", "angled_x"), ("y", "angled_y"), ("z", "angled_z")):
        if col in data:
            ad[axis] = data.pop(col)
    if ad:
        data["angled_direction"] = ad
    return {"gravity": data} if data else {}


def _read_global_sheet(ws) -> Dict:
    float_fields = list(GLOBAL_FIELDS.keys()) + [f"concentration_{i}" for i in range(1, 16)]
    data = _read_key_value_sheet(
        ws,
        string_fields=("ambient_transient", "radiant_transient"),
        float_fields=[f for f in float_fields if GLOBAL_FIELDS.get(f) is float or f.startswith("concentration_")],
    )
    return {"global": data} if data else {}


def _read_initial_variables_sheet(ws) -> Dict:
    """读 initial_variables，列名格式是 pressure_type, pressure_value 这样"""
    headers = [str(ws.cell(row=1, column=c).value or "").strip()
               for c in range(1, ws.max_column + 1)]
    if not headers or not headers[0]:
        return {}

    result: Dict = {}
    for row in range(2, ws.max_row + 1):
        rd = _row_dict(ws, row, headers)
        if not rd:
            continue

        for field in ("initial_solution_name", "initial_solution_id"):
            raw = rd.get(field)
            if raw is not None and str(raw).strip():
                result[field] = str(raw).strip()

        if "use_initial_for_all" in headers:
            v = _parse_bool(rd.get("use_initial_for_all"))
            if v is not None:
                result["use_initial_for_all"] = v

        for var_name in _INITIAL_VAR_NAMES:
            type_col = f"{var_name}_type"
            val_col = f"{var_name}_value"
            if type_col in headers or val_col in headers:
                var_cfg: Dict = {}
                raw_type = rd.get(type_col)
                if raw_type is not None and str(raw_type).strip():
                    var_cfg["type"] = str(raw_type).strip()
                raw_val = rd.get(val_col)
                if raw_val is not None and (not isinstance(raw_val, str) or raw_val.strip()):
                    try:
                        var_cfg["value"] = float(raw_val)
                    except (ValueError, TypeError):
                        pass
                if var_cfg:
                    result[var_name] = var_cfg

    return {"initial_variables": result} if result else {}


def load_excel(path: Path) -> Dict:
    try:
        from openpyxl import load_workbook as _load_wb
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")

    wb = _load_wb(str(path), read_only=True, data_only=True)
    config: Dict = {}

    # 按顺序读各个 sheet，有哪个读哪个
    if "modeling" in wb.sheetnames:
        config.update(_read_modeling_sheet(wb["modeling"]))

    if "turbulence" in wb.sheetnames:
        config.update(_read_turbulence_sheet(wb["turbulence"]))

    if "gravity" in wb.sheetnames:
        config.update(_read_gravity_sheet(wb["gravity"]))

    if "global" in wb.sheetnames:
        config.update(_read_global_sheet(wb["global"]))

    if "initial_variables" in wb.sheetnames:
        config.update(_read_initial_variables_sheet(wb["initial_variables"]))

    if "overall_control" in wb.sheetnames:
        config.update(_read_overall_control_sheet(wb["overall_control"]))

    if "variable_controls" in wb.sheetnames:
        data = _read_controls_sheet(wb["variable_controls"], "variable_controls")
        if data:
            config["variable_controls"] = data

    if "solver_controls" in wb.sheetnames:
        data = _read_controls_sheet(wb["solver_controls"], "solver_controls")
        if data:
            config["solver_controls"] = data

    # 瞬态相关的三个 sheet 合并到 transient 下面
    if "transient_overall" in wb.sheetnames:
        data = _read_transient_overall_sheet(wb["transient_overall"])
        if data:
            config.setdefault("transient", {}).setdefault("overall_transient", {}).update(data)

    if "transient_save_times" in wb.sheetnames:
        data = _read_transient_save_times_sheet(wb["transient_save_times"])
        if data:
            config.setdefault("transient", {})["save_times"] = data

    if "time_patches" in wb.sheetnames:
        data = _read_time_patches_sheet(wb["time_patches"])
        if data:
            config.setdefault("transient", {})["time_patches"] = data

    wb.close()
    return config


def load_config(path: Path) -> Dict:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls", ".xlsm"):
        return load_excel(path)
    return load_json(path)


# ============================================================================
# 生成 Excel 模板
# ============================================================================

def create_template_excel(output_path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    header_font = Font(bold=True)
    center = Alignment(horizontal="center")

    def _headers(ws, names):
        for col, h in enumerate(names, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.alignment = center

    def _widths(ws, w=20):
        for idx in range(1, ws.max_column + 1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w

    # ── Sheet 1: modeling ──
    ws1 = wb.active
    ws1.title = "modeling"
    _headers(ws1, [
        "solution", "dimensionality", "radiation", "joule_heating",
        "store_mass_flux", "store_heat_flux", "store_surface_temp",
        "store_grad_t", "store_bn_sc", "store_power_density",
        "store_mean_radiant_temperature",
    ])
    example1 = ["flow_heat", "3d", "on", "false",
                "false", "false", "false", "false", "false", "false", "false"]
    for c, val in enumerate(example1, 1):
        ws1.cell(row=2, column=c, value=val)
    _widths(ws1)

    # ── Sheet 2: turbulence ──
    ws2 = wb.create_sheet("turbulence")
    _headers(ws2, ["type", "turbulence_type", "lvel_k_epsilon_stratification",
                    "capped_lvel_multiplier", "ra_velocity", "ra_length"])
    example2 = ["laminar", "", "false", "", "", ""]
    for c, val in enumerate(example2, 1):
        if val != "":
            ws2.cell(row=2, column=c, value=val)
    _widths(ws2)

    # ── Sheet 3: gravity ──
    ws3 = wb.create_sheet("gravity")
    _headers(ws3, ["type", "normal_direction", "value_type", "gravity_value",
                    "angled_x", "angled_y", "angled_z"])
    example3 = ["normal", "neg_y", "default", "", "", "", ""]
    for c, val in enumerate(example3, 1):
        if val != "":
            ws3.cell(row=2, column=c, value=val)
    _widths(ws3)

    # ── Sheet 4: global ──
    ws4 = wb.create_sheet("global")
    _headers(ws4, ["datum_pressure", "ambient_temperature", "radiant_temperature",
                    "ambient_transient", "radiant_transient"])
    example4 = [101325.0, 25.0, 25.0, "", ""]
    for c, val in enumerate(example4, 1):
        if val != "":
            ws4.cell(row=2, column=c, value=val)
    _widths(ws4)

    # ── Sheet 5: initial_variables ──
    ws5 = wb.create_sheet("initial_variables")
    iv_headers = ["initial_solution_name", "initial_solution_id", "use_initial_for_all"]
    for var_name in ("pressure", "x_velocity", "y_velocity", "z_velocity", "temperature"):
        iv_headers.extend([f"{var_name}_type", f"{var_name}_value"])
    _headers(ws5, iv_headers)
    example5 = ["", "", "false",
                "automatic", "", "automatic", "", "automatic", "", "automatic", "",
                "user_specified", 25.0]
    for c, val in enumerate(example5, 1):
        if val != "":
            ws5.cell(row=2, column=c, value=val)
    _widths(ws5)

    # ── Sheet 6: overall_control ──
    ws6 = wb.create_sheet("overall_control")
    _headers(ws6, [
        "outer_iterations", "fan_relaxation", "estimated_free_convection_velocity",
        "solver_option", "monitor_convergence",
        "active_plate_conduction", "use_double_precision", "multi_grid_damping",
        "network_assembly_block_correction", "freeze_flow", "store_error_field",
        "error_field_variable",
        "required_accuracy", "num_iterations", "residual_threshold",
    ])
    example6 = [500, 0.9, 0.2, "multi_grid", "true",
                "false", "false", "false", "false", "false", "false",
                "", 0.2, 45, 200]
    for c, val in enumerate(example6, 1):
        ws6.cell(row=2, column=c, value=val)
    _widths(ws6)

    # ── Sheet 7: variable_controls ──
    ws7 = wb.create_sheet("variable_controls")
    _headers(ws7, [
        "variable", "false_time_step", "false_time_step_user_value",
        "false_time_step_damping_auto_multiplier",
        "terminal_residual", "terminal_residual_auto_multiplier",
        "terminal_residual_user_value", "inner_iterations",
    ])
    for r, row_data in enumerate([
        ("x_velocity", "user", 1.5, 0, "automatic", 1, "", 1),
        ("y_velocity", "user", 1.5, 0, "automatic", 1, "", 1),
        ("z_velocity", "user", 1.5, 0, "automatic", 1, "", 1),
        ("temperature", "user", 1.5, 0, "automatic", 1, "", 1),
    ], 2):
        for c, val in enumerate(row_data, 1):
            if val != "":
                ws7.cell(row=r, column=c, value=val)
    _widths(ws7)

    # ── Sheet 8: solver_controls ──
    ws8 = wb.create_sheet("solver_controls")
    _headers(ws8, ["variable", "linear_relaxation", "error_compute_frequency"])
    for r, (var, lr, ecf) in enumerate([
        ("pressure", 0.3, 0),
        ("temperature", 0.7, 0),
    ], 2):
        ws8.cell(row=r, column=1, value=var)
        ws8.cell(row=r, column=2, value=lr)
        ws8.cell(row=r, column=3, value=ecf)
    _widths(ws8)

    # ── Sheet 9: transient_overall ──
    ws9 = wb.create_sheet("transient_overall")
    _headers(ws9, ["start_time", "end_time", "duration", "keypoint_tolerance"])
    for c, val in enumerate([0, 60, 60, 0.0001], 1):
        ws9.cell(row=2, column=c, value=val)
    _widths(ws9)

    # ── Sheet 10: transient_save_times ──
    ws10 = wb.create_sheet("transient_save_times")
    _headers(ws10, ["save_time"])
    for r, val in enumerate([0, 30, 60], 2):
        ws10.cell(row=r, column=1, value=val)
    _widths(ws10)

    # ── Sheet 11: time_patches ──
    ws11 = wb.create_sheet("time_patches")
    _headers(ws11, [
        "name", "start_time", "end_time",
        "step_control", "additional_steps", "minimum_number", "maximum_size",
        "step_distribution", "distribution_index",
    ])
    for r, row_data in enumerate([
        ("First", 0, 30, "minimum_number", None, 15, None, "increasing_power", 1.4),
        ("Second", 30, 60, "minimum_number", None, 12, None, "uniform", 1),
    ], 2):
        for c, val in enumerate(row_data, 1):
            if val is not None:
                ws11.cell(row=r, column=c, value=val)
    _widths(ws11)

    wb.save(output_path)
    print(f"[OK] Template created: {output_path}")


# ============================================================================
# XML 缩进
# ============================================================================

def indent_xml(elem: ET.Element, level: int = 0) -> None:
    indent = "\n" + ("    " * level)
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = indent + "    "
        for child in elem:
            indent_xml(child, level + 1)
        if not elem[-1].tail or not elem[-1].tail.strip():
            elem[-1].tail = indent
    if level and (not elem.tail or not elem.tail.strip()):
        elem.tail = indent


# ============================================================================
# 命令行入口
# ============================================================================

def main() -> int:
    parser = argparse.ArgumentParser(
        description="往 FloXML 里加求解/瞬态设置"
    )
    parser.add_argument("input", nargs="?", help="输入 FloXML 文件")
    parser.add_argument("--config", help="配置文件 (.json 或 .xlsx)")
    parser.add_argument("-o", "--output", help="输出 FloXML 文件")
    parser.add_argument("--create-template", metavar="PATH",
                        help="生成一个 Excel 模板")
    args = parser.parse_args()

    if args.create_template:
        create_template_excel(args.create_template)
        return 0

    if not args.input or not args.config:
        parser.error("需要指定 input 和 --config（或者用 --create-template）")

    input_path = Path(args.input)
    config_path = Path(args.config)
    output_path = (Path(args.output) if args.output
                   else input_path.with_name(f"{input_path.stem}_with_solve{input_path.suffix}"))

    tree = ET.parse(input_path)
    root = tree.getroot()
    config = load_config(config_path)
    root = apply_solve_settings(root, config)
    indent_xml(root)
    tree.write(output_path, encoding="utf-8", xml_declaration=True)

    # 打印写了哪些东西
    sections = []
    if "overall_control" in config:
        sections.append("overall_control")
    if "variable_controls" in config:
        sections.append(f"{len(config['variable_controls'])} variable_control(s)")
    if "solver_controls" in config:
        sections.append(f"{len(config['solver_controls'])} solver_control(s)")
    if "transient" in config:
        tr = config["transient"]
        parts = []
        if "overall_transient" in tr:
            parts.append("overall_transient")
        if "time_patches" in tr:
            parts.append(f"{len(tr['time_patches'])} time_patch(es)")
        if "save_times" in tr:
            parts.append(f"{len(tr['save_times'])} save_time(s)")
        if parts:
            sections.append("transient(" + ", ".join(parts) + ")")
    print(f"[OK] Applied solve settings ({', '.join(sections)}): {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
