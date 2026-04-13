#!/usr/bin/env python3
"""
Add / modify boundary conditions in an existing FloXML project using JSON config.

Supported boundary condition types:

1. **Ambient** (ambient_att)
   - Temperature, pressure, heat transfer coefficient, velocity, radiation

2. **Solution Domain Boundaries** (solution_domain)
   - x/y/z low/high boundary type: ambient | symmetry | wall | opening | recirculation
   - Per-face ambient assignment

3. **Surface Property** (surface_att)
   - Emissivity, roughness, rsurf_fluid, rsurf_solid, area_factor, solar_reflectivity

4. **Radiation** (radiation_att)
   - Surface type, min_area, subdivided_surface_tolerance

5. **Source** (source_att)
   - Total power / fixed temperature / volumetric heat source
   - Per-variable source options (temperature, x/y/z_velocity, etc.)

6. **Surface Exchange** (surface_exchange_att)
   - Volume / surface heat transfer method
   - Profile or constant heat transfer coefficient
   - Reference temperature

7. **Thermal Model** (thermal_att)
   - Conduction / convection model
   - Total power

Usage:
    python floxml_boundary_conditions.py input.xml --config config.json -o output.xml
    python floxml_boundary_conditions.py input.xml --config config.xlsx -o output.xml
    python floxml_boundary_conditions.py --create-template template.xlsx
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List, Optional
import xml.etree.ElementTree as ET


# ============================================================================
# XML helpers (mirrors floxml_add_volume_regions.py style)
# ============================================================================

def _append_text(parent: ET.Element, tag: str, text: str) -> ET.Element:
    elem = ET.SubElement(parent, tag)
    elem.text = text
    return elem


def _set_text(parent: ET.Element, tag: str, text: str) -> ET.Element:
    child = parent.find(tag)
    if child is None:
        child = ET.SubElement(parent, tag)
    child.text = text
    return child


def _set_vector(parent: ET.Element, tag: str, vec: List[float]) -> ET.Element:
    """Set <tag><x/><y/><z/></tag> from a 3-element list."""
    elem = parent.find(tag)
    if elem is None:
        elem = ET.SubElement(parent, tag)
    for axis, val in zip(("x", "y", "z"), vec):
        _set_text(elem, axis, f"{val:.6g}")
    return elem


def _set_optional_text(parent: ET.Element, tag: str, value: Any) -> None:
    """Set a text element only if value is not None."""
    if value is None:
        return
    _set_text(parent, tag, str(value))


def _set_optional_float(parent: ET.Element, tag: str, value: Any) -> None:
    if value is None:
        return
    _set_text(parent, tag, f"{float(value):.6g}")


# ============================================================================
# Attribute section helpers
# ============================================================================

def _find_or_create_section(root: ET.Element, section_tag: str) -> ET.Element:
    """Find or create a top-level attributes section (e.g. ambients, surfaces)."""
    attributes = root.find("attributes")
    if attributes is None:
        # Insert after <geometry> or at end
        insert_at = 0
        for idx, child in enumerate(list(root)):
            if child.tag in {"geometry", "solution_domain"}:
                insert_at = idx
            insert_at = idx + 1
        attributes = ET.Element("attributes")
        root.insert(insert_at, attributes)

    section = attributes.find(section_tag)
    if section is None:
        section = ET.SubElement(attributes, section_tag)
    return section


def _find_attribute_by_name(section: ET.Element, name: str, att_tag: str) -> Optional[ET.Element]:
    """Find an existing attribute element by name within a section."""
    for child in section.findall(att_tag):
        if (child.findtext("name") or "").strip() == name:
            return child
    return None


def _upsert_attribute(section: ET.Element, name: str, att_tag: str) -> ET.Element:
    """Find or create an attribute element by name."""
    existing = _find_attribute_by_name(section, name, att_tag)
    if existing is not None:
        return existing
    elem = ET.SubElement(section, att_tag)
    _append_text(elem, "name", name)
    return elem


# ============================================================================
# Boundary condition builders
# ============================================================================

def _apply_ambient(section: ET.Element, cfg: Dict) -> ET.Element:
    """Create or update an ambient_att."""
    name = cfg.get("name", "Ambient")
    elem = _upsert_attribute(section, name, "ambient_att")

    _set_optional_float(elem, "pressure", cfg.get("pressure"))
    _set_optional_float(elem, "temperature", cfg.get("temperature"))
    _set_optional_float(elem, "radiant_temperature", cfg.get("radiant_temperature"))
    _set_optional_float(elem, "heat_transfer_coeff", cfg.get("heat_transfer_coeff"))

    velocity = cfg.get("velocity")
    if velocity is not None and len(velocity) == 3:
        _set_vector(elem, "velocity", [float(v) for v in velocity])

    _set_optional_float(elem, "turbulent_kinetic_energy", cfg.get("turbulent_kinetic_energy"))
    _set_optional_float(elem, "turbulent_dissipation_rate", cfg.get("turbulent_dissipation_rate"))

    for i in range(1, 6):
        _set_optional_float(elem, f"concentration_{i}", cfg.get(f"concentration_{i}"))

    return elem


def _apply_solution_domain(root: ET.Element, cfg: Dict) -> ET.Element:
    """Modify solution_domain boundary conditions."""
    sd = root.find("solution_domain")
    if sd is None:
        raise ValueError("FloXML missing <solution_domain> section")

    # Boundary type for each face: ambient | symmetry | wall | opening | recirculation
    boundary_axes = [
        ("x_low", "x_high"),
        ("y_low", "y_high"),
        ("z_low", "z_high"),
    ]
    for axis_prefix, sides in boundary_axes:
        for side in sides:
            key = f"{axis_prefix}_{side}_boundary"
            value = cfg.get(key)
            if value is not None:
                _set_text(sd, key, str(value))

            # Per-face ambient assignment
            ambient_key = f"{axis_prefix}_{side}_ambient"
            ambient_value = cfg.get(ambient_key)
            if ambient_value is not None:
                _set_text(sd, ambient_key, str(ambient_value))

    return sd


def _apply_surface(section: ET.Element, cfg: Dict) -> ET.Element:
    """Create or update a surface_att (surface property / emissivity)."""
    name = cfg.get("name", "Default")
    elem = _upsert_attribute(section, name, "surface_att")

    _set_optional_float(elem, "emissivity", cfg.get("emissivity"))
    _set_optional_float(elem, "roughness", cfg.get("roughness"))
    _set_optional_float(elem, "rsurf_fluid", cfg.get("rsurf_fluid"))
    _set_optional_float(elem, "rsurf_solid", cfg.get("rsurf_solid"))
    _set_optional_float(elem, "area_factor", cfg.get("area_factor"))
    _set_optional_float(elem, "solar_reflectivity", cfg.get("solar_reflectivity"))

    # Display settings (optional)
    display = cfg.get("display_settings")
    if display:
        ds_elem = elem.find("display_settings")
        if ds_elem is None:
            ds_elem = ET.SubElement(elem, "display_settings")
        color = display.get("color")
        if color and len(color) == 3:
            color_elem = ds_elem.find("color")
            if color_elem is None:
                color_elem = ET.SubElement(ds_elem, "color")
            _set_text(color_elem, "red", f"{float(color[0]):.6g}")
            _set_text(color_elem, "green", f"{float(color[1]):.6g}")
            _set_text(color_elem, "blue", f"{float(color[2]):.6g}")
        _set_optional_float(ds_elem, "shininess", display.get("shininess"))
        _set_optional_float(ds_elem, "brightness", display.get("brightness"))

    return elem


def _apply_radiation(section: ET.Element, cfg: Dict) -> ET.Element:
    """Create or update a radiation_att."""
    name = cfg.get("name", "Default")
    elem = _upsert_attribute(section, name, "radiation_att")

    _set_text(elem, "surface", str(cfg.get("surface", "discrete_ordinates")))
    _set_optional_float(elem, "min_area", cfg.get("min_area"))
    _set_optional_float(elem, "subdivided_surface_tolerance", cfg.get("subdivided_surface_tolerance"))

    return elem


def _apply_source(section: ET.Element, cfg: Dict) -> ET.Element:
    """Create or update a source_att."""
    name = cfg.get("name", "Source")
    elem = _upsert_attribute(section, name, "source_att")

    # Remove existing source_options to rebuild
    existing_opts = elem.find("source_options")
    if existing_opts is not None:
        elem.remove(existing_opts)

    options = cfg.get("source_options", [])
    if options:
        opts_elem = ET.SubElement(elem, "source_options")
        for opt in options:
            opt_elem = ET.SubElement(opts_elem, "option")
            _set_text(opt_elem, "applies_to", str(opt.get("applies_to", "temperature")))
            _set_text(opt_elem, "type", str(opt.get("type", "total")))
            _set_optional_float(opt_elem, "value", opt.get("value"))
            _set_optional_float(opt_elem, "power", opt.get("power"))
            _set_optional_float(opt_elem, "linear_coefficient", opt.get("linear_coefficient"))
            if opt.get("transient"):
                _set_text(opt_elem, "transient", str(opt["transient"]))

    notes = cfg.get("notes")
    if notes:
        _set_text(elem, "notes", str(notes))

    return elem


def _apply_surface_exchange(section: ET.Element, cfg: Dict) -> ET.Element:
    """Create or update a surface_exchange_att."""
    name = cfg.get("name", "Surface Exchange")
    elem = _upsert_attribute(section, name, "surface_exchange_att")

    _set_text(elem, "heat_transfer_method", str(cfg.get("heat_transfer_method", "surface")))
    _set_optional_float(elem, "extent_of_heat_transfer", cfg.get("extent_of_heat_transfer"))
    _set_optional_float(elem, "wetted_area_volume_transfer", cfg.get("wetted_area_volume_transfer"))

    htc = cfg.get("heat_transfer_coefficient")
    if htc:
        _set_text(elem, "heat_transfer_coefficient", str(htc))

        if htc == "profile":
            profile = cfg.get("profile", [])
            # Remove existing profile
            existing_profile = elem.find("profile")
            if existing_profile is not None:
                elem.remove(existing_profile)

            if profile:
                profile_elem = ET.SubElement(elem, "profile")
                for pt in profile:
                    pt_elem = ET.SubElement(profile_elem, "heat_sink_curve_point")
                    _set_optional_float(pt_elem, "speed", pt.get("speed"))
                    _set_optional_float(pt_elem, "thermal_resistance", pt.get("thermal_resistance"))

        elif htc == "constant":
            _set_optional_float(elem, "specified_constant_value", cfg.get("specified_constant_value"))

    ref_temp = cfg.get("reference_temperature")
    if ref_temp:
        _set_text(elem, "reference_temperature", str(ref_temp))
        if ref_temp == "specified":
            _set_optional_float(elem, "reference_temperature_value", cfg.get("reference_temperature_value"))

    notes = cfg.get("notes")
    if notes:
        _set_text(elem, "notes", str(notes))

    return elem


def _apply_thermal(section: ET.Element, cfg: Dict) -> ET.Element:
    """Create or update a thermal_att."""
    name = cfg.get("name", "Thermal")
    elem = _upsert_attribute(section, name, "thermal_att")

    _set_text(elem, "thermal_model", str(cfg.get("thermal_model", "conduction")))
    _set_optional_float(elem, "power", cfg.get("power"))

    if cfg.get("transient"):
        _set_text(elem, "transient", str(cfg["transient"]))

    return elem


# ============================================================================
# Geometry face boundary conditions
# ============================================================================

def _iter_geometry(
    geometry: ET.Element,
    parent_global: tuple = (0.0, 0.0, 0.0),
    assembly_path: tuple = (),
):
    """Iterate all geometry items recursively."""
    from floxml_add_volume_regions import GeometryItem, _parse_position, _parse_size, _vector_add
    for child in list(geometry):
        tag = child.tag
        position = _parse_position(child)
        size = _parse_size(child)
        global_position = _vector_add(parent_global, position)
        name = (child.findtext("name") or "").strip()
        item = GeometryItem(
            name=name,
            tag=tag,
            position=position,
            size=size,
            global_position=global_position,
            global_size=size,
            element=child,
            parent_geometry=geometry,
            assembly_path=assembly_path,
        )
        yield item
        child_geometry = child.find("geometry")
        if child_geometry is not None:
            next_path = assembly_path
            if tag == "assembly" and name:
                next_path = assembly_path + (name,)
            yield from _iter_geometry(
                child_geometry,
                parent_global=global_position,
                assembly_path=next_path,
            )


def _apply_geometry_face_conditions(root: ET.Element, face_conditions: List[Dict]) -> int:
    """Apply boundary conditions to geometry faces (plates, cuboids, etc.)."""
    import fnmatch

    geometry = root.find("geometry")
    if geometry is None:
        raise ValueError("FloXML missing <geometry> section")

    all_items = list(_iter_geometry(geometry))
    applied = 0

    for fc in face_conditions:
        target_names = list(fc.get("target_names", []))
        target_patterns = list(fc.get("target_patterns", []))
        target_tags = list(fc.get("target_tags", []))

        if not target_names and not target_patterns and not target_tags:
            continue

        matched = []
        for item in all_items:
            if not item.name:
                continue
            if target_tags and item.tag not in target_tags:
                continue
            if item.name in target_names or any(
                fnmatch.fnmatch(item.name, p) for p in target_patterns
            ):
                matched.append(item)

        if not matched:
            print(
                f"Warning: face_conditions matched no geometry: "
                f"names={target_names}, patterns={target_patterns}, tags={target_tags}",
                file=__import__("sys").stderr,
            )
            continue

        for item in matched:
            elem = item.element

            # Surface property
            if fc.get("surface"):
                _set_text(elem, "surface", str(fc["surface"]))

            # Radiation model
            if fc.get("radiation"):
                _set_text(elem, "radiation", str(fc["radiation"]))

            # Thermal model
            if fc.get("thermal"):
                _set_text(elem, "thermal", str(fc["thermal"]))

            # Surface exchange
            if fc.get("surface_exchange"):
                _set_text(elem, "surface_exchange", str(fc["surface_exchange"]))

            # Heat transfer coefficient override
            if fc.get("heat_transfer_coeff") is not None:
                _set_text(elem, "heat_transfer_coeff", str(fc["heat_transfer_coeff"]))

            applied += 1

    return applied


# ============================================================================
# Main apply function
# ============================================================================

def apply_boundary_conditions(root: ET.Element, config: Dict) -> ET.Element:
    """Apply all boundary conditions from config to the FloXML root element."""
    counters = {}

    # 1. Ambients
    ambients = config.get("ambients", [])
    if ambients:
        section = _find_or_create_section(root, "ambients")
        for a in ambients:
            _apply_ambient(section, a)
        counters["ambients"] = len(ambients)

    # 2. Solution domain boundaries
    sd = config.get("solution_domain")
    if sd:
        _apply_solution_domain(root, sd)
        counters["solution_domain"] = 1

    # 3. Surface properties
    surfaces = config.get("surfaces", [])
    if surfaces:
        section = _find_or_create_section(root, "surfaces")
        for s in surfaces:
            _apply_surface(section, s)
        counters["surfaces"] = len(surfaces)

    # 4. Radiations
    radiations = config.get("radiations", [])
    if radiations:
        section = _find_or_create_section(root, "radiations")
        for r in radiations:
            _apply_radiation(section, r)
        counters["radiations"] = len(radiations)

    # 5. Sources
    sources = config.get("sources", [])
    if sources:
        section = _find_or_create_section(root, "sources")
        for s in sources:
            _apply_source(section, s)
        counters["sources"] = len(sources)

    # 6. Surface exchanges
    surface_exchanges = config.get("surface_exchanges", [])
    if surface_exchanges:
        section = _find_or_create_section(root, "surface_exchanges")
        for se in surface_exchanges:
            _apply_surface_exchange(section, se)
        counters["surface_exchanges"] = len(surface_exchanges)

    # 7. Thermals
    thermals = config.get("thermals", [])
    if thermals:
        section = _find_or_create_section(root, "thermals")
        for t in thermals:
            _apply_thermal(section, t)
        counters["thermals"] = len(thermals)

    # 8. Geometry face conditions
    face_conditions = config.get("face_conditions", [])
    if face_conditions:
        n = _apply_geometry_face_conditions(root, face_conditions)
        counters["face_conditions"] = n

    if not counters:
        raise ValueError(
            "Config must contain at least one of: ambients, solution_domain, surfaces, "
            "radiations, sources, surface_exchanges, thermals, face_conditions"
        )

    return root


# ============================================================================
# XML formatting
# ============================================================================

def indent_xml(elem: ET.Element, level: int = 0) -> None:
    indent = "\n" + ("    " * level)
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = indent + "    "
        for child in elem:
            indent_xml(child, level + 1)
        if not elem[-1].tail or not elem[-1].tail.strip():
            elem[-1].tail = indent
    if level and (not elem.tail or not elem.tail.strip()):
        elem.tail = indent


# ============================================================================
# Config loaders
# ============================================================================

def load_json(path: Path) -> Dict:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def load_excel(path: Path) -> Dict:
    """Load boundary condition config from Excel file."""
    try:
        from openpyxl import load_workbook as _load_wb
    except ImportError:
        raise ImportError("需要安装 openpyxl 来读取 Excel 配置: pip install openpyxl")

    wb = _load_wb(str(path), read_only=True, data_only=True)
    config: Dict = {}

    def _parse_number(value):
        if value is None:
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None

    def _split_list(value):
        if value is None:
            return []
        if isinstance(value, str):
            return [s.strip() for s in value.split(",") if s.strip()]
        if isinstance(value, (list, tuple)):
            return [str(v).strip() for v in value if str(v).strip()]
        return [str(value)]

    def _read_sheet(ws) -> List[Dict]:
        headers = [str(ws.cell(row=1, column=c).value or "").strip()
                    for c in range(1, ws.max_column + 1)]
        if not headers or not headers[0]:
            return []
        results = []
        for row in range(2, ws.max_row + 1):
            values = [ws.cell(row=row, column=c).value for c in range(1, len(headers) + 1)]
            if all(v is None for v in values):
                continue
            rd = dict(zip(headers, values))
            # Skip rows with no name
            if not str(rd.get("name", "")).strip():
                continue
            results.append(rd)
        return results

    # Ambients sheet
    if "ambients" in wb.sheetnames:
        rows = _read_sheet(wb["ambients"])
        if rows:
            config["ambients"] = []
            for rd in rows:
                entry: Dict = {"name": str(rd["name"]).strip()}
                for key in ("pressure", "temperature", "radiant_temperature",
                            "heat_transfer_coeff", "turbulent_kinetic_energy",
                            "turbulent_dissipation_rate"):
                    v = _parse_number(rd.get(key))
                    if v is not None:
                        entry[key] = v
                for i in range(1, 6):
                    v = _parse_number(rd.get(f"concentration_{i}"))
                    if v is not None:
                        entry[f"concentration_{i}"] = v
                vel = rd.get("velocity")
                if vel and isinstance(vel, str):
                    parts = [s.strip() for s in vel.split(",")]
                    if len(parts) == 3:
                        entry["velocity"] = [float(p) for p in parts]
                config["ambients"].append(entry)

    # Solution domain sheet
    if "solution_domain" in wb.sheetnames:
        ws = wb["solution_domain"]
        headers = [str(ws.cell(row=1, column=c).value or "").strip()
                    for c in range(1, ws.max_column + 1)]
        sd: Dict = {}
        for col_idx, header in enumerate(headers, 1):
            val = ws.cell(row=2, column=col_idx).value
            if val is not None and str(val).strip():
                sd[header] = str(val).strip()
        if sd:
            config["solution_domain"] = sd

    # Surfaces sheet
    if "surfaces" in wb.sheetnames:
        rows = _read_sheet(wb["surfaces"])
        if rows:
            config["surfaces"] = []
            for rd in rows:
                entry: Dict = {"name": str(rd["name"]).strip()}
                for key in ("emissivity", "roughness", "rsurf_fluid", "rsurf_solid",
                            "area_factor", "solar_reflectivity"):
                    v = _parse_number(rd.get(key))
                    if v is not None:
                        entry[key] = v
                config["surfaces"].append(entry)

    # Radiations sheet
    if "radiations" in wb.sheetnames:
        rows = _read_sheet(wb["radiations"])
        if rows:
            config["radiations"] = []
            for rd in rows:
                entry: Dict = {"name": str(rd["name"]).strip()}
                if rd.get("surface"):
                    entry["surface"] = str(rd["surface"]).strip()
                for key in ("min_area", "subdivided_surface_tolerance"):
                    v = _parse_number(rd.get(key))
                    if v is not None:
                        entry[key] = v
                config["radiations"].append(entry)

    # Sources sheet
    if "sources" in wb.sheetnames:
        rows = _read_sheet(wb["sources"])
        if rows:
            config["sources"] = []
            for rd in rows:
                entry: Dict = {"name": str(rd["name"]).strip()}
                notes = rd.get("notes")
                if notes and str(notes).strip():
                    entry["notes"] = str(notes).strip()
                # Source options: encoded as "applies_to|type|value|power" in a single column
                options_raw = rd.get("source_options")
                if options_raw and isinstance(options_raw, str):
                    opts = []
                    for line in options_raw.strip().split("\n"):
                        parts = [p.strip() for p in line.split("|")]
                        if len(parts) >= 2:
                            opt: Dict = {"applies_to": parts[0], "type": parts[1]}
                            if len(parts) > 2 and parts[2]:
                                opt["value"] = float(parts[2])
                            if len(parts) > 3 and parts[3]:
                                opt["power"] = float(parts[3])
                            if len(parts) > 4 and parts[4]:
                                opt["linear_coefficient"] = float(parts[4])
                            if len(parts) > 5 and parts[5]:
                                opt["transient"] = parts[5]
                            opts.append(opt)
                    if opts:
                        entry["source_options"] = opts
                config["sources"].append(entry)

    # Thermals sheet
    if "thermals" in wb.sheetnames:
        rows = _read_sheet(wb["thermals"])
        if rows:
            config["thermals"] = []
            for rd in rows:
                entry: Dict = {"name": str(rd["name"]).strip()}
                if rd.get("thermal_model"):
                    entry["thermal_model"] = str(rd["thermal_model"]).strip()
                v = _parse_number(rd.get("power"))
                if v is not None:
                    entry["power"] = v
                if rd.get("transient") and str(rd["transient"]).strip():
                    entry["transient"] = str(rd["transient"]).strip()
                config["thermals"].append(entry)

    # Surface exchanges sheet
    if "surface_exchanges" in wb.sheetnames:
        rows = _read_sheet(wb["surface_exchanges"])
        if rows:
            config["surface_exchanges"] = []
            for rd in rows:
                entry: Dict = {"name": str(rd["name"]).strip()}
                for key in ("heat_transfer_method", "extent_of_heat_transfer",
                            "wetted_area_volume_transfer", "heat_transfer_coefficient",
                            "specified_constant_value", "reference_temperature",
                            "reference_temperature_value"):
                    v = rd.get(key)
                    if v is not None and str(v).strip():
                        if key in ("extent_of_heat_transfer", "wetted_area_volume_transfer",
                                   "specified_constant_value", "reference_temperature_value"):
                            entry[key] = float(v)
                        else:
                            entry[key] = str(v).strip()
                notes = rd.get("notes")
                if notes and str(notes).strip():
                    entry["notes"] = str(notes).strip()
                config["surface_exchanges"].append(entry)

    # Face conditions sheet (apply to geometry objects)
    if "face_conditions" in wb.sheetnames:
        rows = _read_sheet(wb["face_conditions"])
        if rows:
            config["face_conditions"] = []
            for rd in rows:
                entry: Dict = {}
                names = _split_list(rd.get("target_names"))
                if names:
                    entry["target_names"] = names
                patterns = _split_list(rd.get("target_patterns"))
                if patterns:
                    entry["target_patterns"] = patterns
                tags = _split_list(rd.get("target_tags"))
                if tags:
                    entry["target_tags"] = tags
                for key in ("surface", "radiation", "thermal", "surface_exchange"):
                    v = rd.get(key)
                    if v and str(v).strip():
                        entry[key] = str(v).strip()
                v = _parse_number(rd.get("heat_transfer_coeff"))
                if v is not None:
                    entry["heat_transfer_coeff"] = v
                if entry:
                    config["face_conditions"].append(entry)

    wb.close()
    return config


def load_config(path: Path) -> Dict:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls", ".xlsm"):
        return load_excel(path)
    return load_json(path)


# ============================================================================
# Excel template
# ============================================================================

def create_template_excel(output_path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center')

    def _write_headers(ws, headers):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align

    def _set_col_widths(ws, width=22):
        for idx in range(1, ws.max_column + 1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    # ── Sheet 1: ambients ──
    ws1 = wb.active
    ws1.title = "ambients"
    _write_headers(ws1, [
        "name", "pressure", "temperature", "radiant_temperature",
        "heat_transfer_coeff", "velocity", "turbulent_kinetic_energy",
        "turbulent_dissipation_rate", "concentration_1",
    ])
    ws1.cell(row=2, column=1, value="Ambient")
    ws1.cell(row=2, column=3, value=293)
    ws1.cell(row=2, column=4, value=293)
    ws1.cell(row=2, column=5, value=12)
    ws1.cell(row=2, column=6, value="0,0,0")
    _set_col_widths(ws1)

    # ── Sheet 2: solution_domain ──
    ws2 = wb.create_sheet("solution_domain")
    _write_headers(ws2, [
        "x_low_boundary", "x_high_boundary", "y_low_boundary", "y_high_boundary",
        "z_low_boundary", "z_high_boundary",
        "x_low_ambient", "x_high_ambient", "y_low_ambient", "y_high_ambient",
        "z_low_ambient", "z_high_ambient",
    ])
    for col, val in enumerate(
        ["ambient", "ambient", "ambient", "ambient", "ambient", "ambient",
         "Ambient", "Ambient", "Ambient", "Ambient", "Ambient", "Ambient"],
        1,
    ):
        ws2.cell(row=2, column=col, value=val)
    _set_col_widths(ws2)

    # ── Sheet 3: surfaces ──
    ws3 = wb.create_sheet("surfaces")
    _write_headers(ws3, [
        "name", "emissivity", "roughness", "rsurf_fluid",
        "rsurf_solid", "area_factor", "solar_reflectivity",
    ])
    ws3.cell(row=2, column=1, value="Paint")
    ws3.cell(row=2, column=2, value=0.88)
    _set_col_widths(ws3)

    # ── Sheet 4: radiations ──
    ws4 = wb.create_sheet("radiations")
    _write_headers(ws4, [
        "name", "surface", "min_area", "subdivided_surface_tolerance",
    ])
    ws4.cell(row=2, column=1, value="Sub-Divided")
    ws4.cell(row=2, column=2, value="subdivided_radiating")
    ws4.cell(row=2, column=3, value=0)
    ws4.cell(row=2, column=4, value=0.01)
    _set_col_widths(ws4)

    # ── Sheet 5: sources ──
    ws5 = wb.create_sheet("sources")
    _write_headers(ws5, [
        "name", "source_options", "notes",
    ])
    ws5.cell(row=2, column=1, value="Heat Source")
    ws5.cell(row=2, column=2, value="temperature|total||23.3")
    ws5.cell(row=2, column=3, value="23.3W chip power")
    _set_col_widths(ws5, 30)

    # ── Sheet 6: surface_exchanges ──
    ws6 = wb.create_sheet("surface_exchanges")
    _write_headers(ws6, [
        "name", "heat_transfer_method", "extent_of_heat_transfer",
        "wetted_area_volume_transfer", "heat_transfer_coefficient",
        "specified_constant_value", "reference_temperature",
        "reference_temperature_value", "notes",
    ])
    ws6.cell(row=2, column=1, value="Surface")
    ws6.cell(row=2, column=2, value="surface")
    ws6.cell(row=2, column=5, value="calculated")
    ws6.cell(row=2, column=7, value="calculated")
    _set_col_widths(ws6)

    # ── Sheet 7: thermals ──
    ws7 = wb.create_sheet("thermals")
    _write_headers(ws7, [
        "name", "thermal_model", "power", "transient",
    ])
    ws7.cell(row=2, column=1, value="Conduction")
    ws7.cell(row=2, column=2, value="conduction")
    ws7.cell(row=2, column=3, value=12.5)
    _set_col_widths(ws7)

    # ── Sheet 8: face_conditions ──
    ws8 = wb.create_sheet("face_conditions")
    _write_headers(ws8, [
        "target_names", "target_patterns", "target_tags",
        "surface", "radiation", "thermal", "surface_exchange",
        "heat_transfer_coeff",
    ])
    ws8.cell(row=2, column=1, value="PCB,Heatsink")
    ws8.cell(row=2, column=4, value="Paint")
    ws8.cell(row=2, column=5, value="Sub-Divided")
    _set_col_widths(ws8)

    wb.save(output_path)
    print(f"[OK] 模板已创建: {output_path}")


# ============================================================================
# CLI
# ============================================================================

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Add / modify boundary conditions in a FloXML project file"
    )
    parser.add_argument("input", nargs="?", help="Input FloXML file")
    parser.add_argument("--config", help="Config file (.json or .xlsx)")
    parser.add_argument("-o", "--output", help="Output FloXML file")
    parser.add_argument(
        "--create-template", metavar="PATH",
        help="Create an example Excel template at PATH",
    )
    args = parser.parse_args()

    if args.create_template:
        create_template_excel(args.create_template)
        return 0

    if not args.input or not args.config:
        parser.error("input and --config are required (unless using --create-template)")

    input_path = Path(args.input)
    config_path = Path(args.config)
    output_path = (
        Path(args.output) if args.output
        else input_path.with_name(f"{input_path.stem}_with_bc{input_path.suffix}")
    )

    tree = ET.parse(input_path)
    root = tree.getroot()
    config = load_config(config_path)
    root = apply_boundary_conditions(root, config)
    indent_xml(root)
    tree.write(output_path, encoding="utf-8", xml_declaration=True)

    # Summary
    parts = []
    for key in ("ambients", "surfaces", "radiations", "sources",
                "surface_exchanges", "thermals", "face_conditions"):
        val = config.get(key)
        if val:
            label = key.replace("_", " ")
            parts.append(f"{len(val)} {label}")
    if "solution_domain" in config:
        parts.append("1 solution domain")

    print(f"[OK] Applied {', '.join(parts)}: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
