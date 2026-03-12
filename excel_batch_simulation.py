#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 多配置批量仿真工具

功能:
    1. 从 Excel 读取多个测试配置
    2. 根据每个配置修改 ECXML 模板文件
    3. 批量调用 FloTHERM 求解

使用方法:
    python excel_batch_simulation.py template.ecxml config.xlsx -o ./output
    python excel_batch_simulation.py template.ecxml config.xlsx -o ./output --no-solve
    python excel_batch_simulation.py template.ecxml config.xlsx -o ./output --sheet "配置1"

Excel 格式（长格式，每个参数一行）:
    | config_name | name     | attribute   | value |
    |-------------|----------|-------------|-------|
    | case1       | CPU      | powerDissipation | 10  |
    | case1       | Heatsink | Material.density  | 8900|
    | case1       | PCB      | Size@width       | 0.1 |
    | case2       | CPU      | powerDissipation | 15  |
    | case2       | Heatsink | Material.density  | 8500|

    - config_name: 配置名称（同一配置的多行会被合并处理）
    - name: 元素名称，用于在 materials.material 中查找 name={name} 的元素
    - attribute: 要修改的属性路径，支持点分隔（如 Material.density）和 @ 属性（如 Size@width）
    - value: 要设置的值

    路径组合逻辑: materials.material[name={name}].{attribute} = value
"""

import os
import sys
import argparse
import shutil
import time
import subprocess
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional

# 尝试导入 openpyxl 或 pandas
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

# 导入现有的 ECXML 编辑器
from ecxml_editor import ECXMLParser


class ExcelConfigReader:
    """Excel 配置读取器（长格式）"""

    def __init__(self, filepath: str, sheet_name: str = None):
        """
        初始化

        Args:
            filepath: Excel 文件路径
            sheet_name: Sheet 名称（可选，默认使用第一个 sheet）
        """
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.raw_rows = []  # 原始行数据
        self.configs = []   # 合并后的配置

        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Excel 文件不存在: {filepath}")

    def read(self) -> List[Dict[str, Any]]:
        """
        读取 Excel 配置并按 config_name 分组

        Returns:
            配置列表，每个配置是一个字典，包含多个参数修改项
        """
        if HAS_OPENPYXL:
            raw_rows = self._read_with_openpyxl()
        elif HAS_PANDAS:
            raw_rows = self._read_with_pandas()
        else:
            raise ImportError("需要安装 openpyxl 或 pandas 来读取 Excel 文件\n"
                            "安装命令: pip install openpyxl 或 pip install pandas")

        self.raw_rows = raw_rows
        self.configs = self._group_by_config_name(raw_rows)
        return self.configs

    def _read_with_openpyxl(self) -> List[Dict[str, Any]]:
        """使用 openpyxl 读取原始行"""
        wb = openpyxl.load_workbook(self.filepath)

        # 选择 sheet
        if self.sheet_name:
            if self.sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{self.sheet_name}' 不存在，可用: {wb.sheetnames}")
            ws = wb[self.sheet_name]
        else:
            ws = wb.active

        # 读取表头
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip().lower())
            else:
                headers.append(f"col_{cell.column}")

        # 验证必要列
        required_cols = ['config_name', 'name', 'attribute', 'value']
        for col in required_cols:
            if col not in headers:
                raise ValueError(f"Excel 缺少必要列: {col}，当前列: {headers}")

        # 读取数据行
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # 跳过空行
                continue

            row_data = {'_row': row_idx}
            for i, header in enumerate(headers):
                if i < len(row):
                    row_data[header] = row[i]

            rows.append(row_data)

        wb.close()
        return rows

    def _read_with_pandas(self) -> List[Dict[str, Any]]:
        """使用 pandas 读取原始行"""
        df = pd.read_excel(self.filepath, sheet_name=self.sheet_name or 0)

        # 统一列名为小写
        df.columns = [c.lower().strip() for c in df.columns]

        # 验证必要列
        required_cols = ['config_name', 'name', 'attribute', 'value']
        for col in required_cols:
            if col not in df.columns:
                raise ValueError(f"Excel 缺少必要列: {col}，当前列: {list(df.columns)}")

        # 转换为字典列表
        rows = []
        for idx, row in df.iterrows():
            row_data = {'_row': idx + 2}
            for col in df.columns:
                val = row[col]
                if pd.isna(val):
                    continue
                row_data[col] = val
            rows.append(row_data)

        return rows

    def _group_by_config_name(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        按 config_name 分组，将同一配置的多行合并

        Returns:
            配置列表，每个配置包含:
            - config_name: 配置名称
            - params: 参数列表，每个参数是 {name, attribute, value}
        """
        configs_dict = {}

        for row in rows:
            config_name = str(row.get('config_name', '')).strip()
            if not config_name:
                continue

            if config_name not in configs_dict:
                configs_dict[config_name] = {
                    'config_name': config_name,
                    'params': []
                }

            # 添加参数
            param = {
                'name': row.get('name', ''),
                'attribute': row.get('attribute', ''),
                'value': row.get('value'),
                '_row': row.get('_row')
            }

            # 只添加有效参数
            if param['name'] and param['attribute'] and param['value'] is not None:
                configs_dict[config_name]['params'].append(param)

        # 转换为列表，保持原始顺序
        configs = []
        seen = set()
        for row in rows:
            config_name = str(row.get('config_name', '')).strip()
            if config_name and config_name not in seen and config_name in configs_dict:
                configs.append(configs_dict[config_name])
                seen.add(config_name)

        return configs

    def get_config_names(self) -> List[str]:
        """获取所有配置名称"""
        return [c['config_name'] for c in self.configs]


def find_flotherm_executable() -> Optional[str]:
    """查找 FloTHERM 可执行文件"""
    possible_paths = [
        # Siemens 版本路径
        r"C:\Program Files\Siemens\SimcenterFlotherm\2020.2\bin\flotherm.exe",
        r"C:\Program Files\Siemens\SimcenterFlotherm\2021.1\bin\flotherm.exe",
        r"C:\Program Files\Siemens\SimcenterFlotherm\2021.2\bin\flotherm.exe",
        r"C:\Program Files\Siemens\SimcenterFlotherm\2022.1\bin\flotherm.exe",
        r"C:\Program Files\Siemens\SimcenterFlotherm\2023.1\bin\flotherm.exe",
        r"C:\Program Files\Siemens\SimcenterFlotherm\2024.1\bin\flotherm.exe",
        r"C:\Program Files (x86)\Siemens\SimcenterFlotherm\2020.2\bin\flotherm.exe",
        # Mentor 版本路径
        r"C:\Program Files\MentorMA\flosuite_v13\flotherm\bin\flotherm.exe",
        r"C:\Program Files (x86)\MentorMA\flosuite_v13\flotherm\bin\flotherm.exe",
        # 尝试 PATH 中的 flotherm
        "flotherm",
        "flotherm.exe",
    ]

    for path in possible_paths:
        if os.path.exists(path):
            return path

    return None


def apply_config_to_ecxml(template_path: str, config: Dict[str, Any], output_path: str,
                           base_path: str = "materials.material") -> bool:
    """
    将配置应用到 ECXML 模板

    新格式配置:
        config = {
            'config_name': 'case1',
            'params': [
                {'name': 'CPU', 'attribute': 'powerDissipation', 'value': 10},
                {'name': 'Heatsink', 'attribute': 'Material.density', 'value': 8900},
            ]
        }

    路径组合: {base_path}[name={name}].{attribute} = value

    Args:
        template_path: 模板 ECXML 文件路径
        config: 配置字典，包含 config_name 和 params
        output_path: 输出文件路径
        base_path: 基础路径，默认为 "materials.material"

    Returns:
        是否成功
    """
    # 复制模板
    shutil.copy(template_path, output_path)

    # 创建解析器并修改
    parser = ECXMLParser(output_path)

    modified_count = 0
    params = config.get('params', [])

    for param in params:
        name = param.get('name', '')
        attribute = param.get('attribute', '')
        value = param.get('value')

        if not name or not attribute or value is None:
            continue

        # 尝试转换为数值
        try:
            if isinstance(value, str):
                # 尝试转换为数值
                try:
                    value = float(value)
                    if value == int(value):
                        value = int(value)
                except ValueError:
                    pass  # 保留字符串
            elif isinstance(value, (int, float)):
                pass  # 保持原样
            else:
                value = str(value)
        except (ValueError, TypeError):
            value = str(value)

        # 构建完整路径: materials.material[name={name}].{attribute}
        full_path = f"{base_path}[name={name}].{attribute}"

        # 使用路径定位设置值
        if parser.set_value_by_path(full_path, value):
            modified_count += 1
        else:
            print(f"    ⚠ 未找到参数: {full_path}")

    # 保存修改
    parser.save(output_path)

    return modified_count > 0


def solve_ecxml(flotherm_exe: str, ecxml_path: str, output_pack_path: str,
                output_html_path: str, index: int, total: int) -> tuple:
    """
    求解单个 ECXML 文件

    Args:
        flotherm_exe: FloTHERM 可执行文件路径
        ecxml_path: 输入的 ECXML 文件路径
        output_pack_path: 输出的 PACK 文件路径
        output_html_path: 输出的 HTML 报告路径
        index: 当前索引（从1开始）
        total: 总文件数

    Returns:
        (success, elapsed_time, message)
    """
    print(f"\n{'='*60}")
    print(f"  [{index}/{total}] 正在求解: {Path(ecxml_path).name}")
    print(f"{'='*60}")
    print(f"  输入: {ecxml_path}")
    print(f"  输出: {output_pack_path}")
    print(f"  开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  (命令行会等待求解完成，请勿关闭)\n")

    start_time = time.time()

    # 构建命令
    cmd = [
        flotherm_exe,
        "-b",
        str(ecxml_path),
        "-z",
        str(output_pack_path),
        "-r",
        str(output_html_path)
    ]

    try:
        # 运行 FloTHERM
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
        )

        elapsed_time = time.time() - start_time

        if result.returncode == 0:
            # 检查输出文件是否存在
            if os.path.exists(output_pack_path):
                file_size = os.path.getsize(output_pack_path) / (1024 * 1024)  # MB
                print(f"\n  ✅ 求解完成!")
                print(f"     耗时: {elapsed_time:.1f} 秒")
                print(f"     文件大小: {file_size:.2f} MB")
                return (True, elapsed_time, "成功")
            else:
                print(f"\n  ⚠️ 命令执行成功但未找到输出文件")
                return (False, elapsed_time, "输出文件不存在")
        else:
            print(f"  ❌ 求解失败!")
            print(f"     返回码: {result.returncode}")
            if result.stderr:
                print(f"     错误信息: {result.stderr[:500]}")
            return (False, elapsed_time, f"返回码: {result.returncode}")

    except subprocess.TimeoutExpired:
        elapsed_time = time.time() - start_time
        print(f"\n  ❌ 超时!")
        return (False, elapsed_time, "超时")
    except Exception as e:
        elapsed_time = time.time() - start_time
        print(f"\n  ❌ 异常: {e}")
        return (False, elapsed_time, str(e))


def generate_summary_report(output_folder: Path, configs: List[Dict],
                           results: List[Dict], template_name: str):
    """
    生成汇总报告

    Args:
        output_folder: 输出文件夹
        configs: 配置列表
        results: 结果列表
        template_name: 模板文件名
    """
    # 文本报告
    report_path = output_folder / "batch_report.txt"
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("Excel 多配置批量仿真报告\n")
        f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"模板文件: {template_name}\n")
        f.write(f"{'='*60}\n\n")
        f.write(f"总配置数: {len(results)}\n")
        success_count = sum(1 for r in results if r["success"])
        f.write(f"成功: {success_count}\n")
        f.write(f"失败: {len(results) - success_count}\n\n")

        f.write("详细结果:\n")
        f.write("-" * 60 + "\n")

        for r in results:
            status = "成功" if r["success"] else "失败"
            f.write(f"\n[{status}] {r['config_name']}\n")
            f.write(f"  ECXML: {r['ecxml_path']}\n")
            f.write(f"  PACK: {r['pack_path']}\n")
            if r.get('html_path'):
                f.write(f"  报告: {r['html_path']}\n")
            f.write(f"  耗时: {r['elapsed']:.1f}s\n")
            f.write(f"  信息: {r['message']}\n")

            # 写入配置参数
            config = r.get('config', {})
            params = config.get('params', [])
            if params:
                f.write(f"  配置参数:\n")
                for p in params:
                    f.write(f"    {p['name']}.{p['attribute']}: {p['value']}\n")

    print(f"\n📄 报告已保存: {report_path}")

    # 如果有 pandas，生成 Excel 汇总
    if HAS_PANDAS:
        summary_path = output_folder / "summary.xlsx"
        summary_data = []
        for r in results:
            row = {
                'config_name': r['config_name'],
                'status': '成功' if r['success'] else '失败',
                'elapsed_s': r['elapsed'],
                'ecxml': Path(r['ecxml_path']).name if r.get('ecxml_path') else '',
                'pack': Path(r['pack_path']).name if r.get('pack_path') else '',
                'message': r['message']
            }
            # 添加配置参数（将参数展开为列）
            config = r.get('config', {})
            params = config.get('params', [])
            for p in params:
                col_name = f"{p['name']}.{p['attribute']}"
                row[col_name] = p['value']
            summary_data.append(row)

        df = pd.DataFrame(summary_data)
        df.to_excel(summary_path, index=False)
        print(f"📊 Excel 汇总: {summary_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Excel 多配置批量仿真工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Excel 格式示例（长格式，每个参数一行）:
    | config_name | name     | attribute         | value |
    |-------------|----------|-------------------|-------|
    | case1       | CPU      | powerDissipation  | 10    |
    | case1       | Heatsink | Material.density  | 8900  |
    | case2       | CPU      | powerDissipation  | 15    |
    | case2       | Heatsink | Material.density  | 8500  |

示例命令:
    python excel_batch_simulation.py template.ecxml config.xlsx -o ./output
    python excel_batch_simulation.py template.ecxml config.xlsx -o ./output --no-solve
    python excel_batch_simulation.py template.ecxml config.xlsx -o ./output --sheet "Sheet2"
        """
    )

    parser.add_argument(
        "template",
        help="ECXML 模板文件路径"
    )

    parser.add_argument(
        "excel",
        help="Excel 配置文件路径"
    )

    parser.add_argument(
        "-o", "--output",
        required=True,
        help="输出文件夹路径"
    )

    parser.add_argument(
        "--sheet",
        help="Excel Sheet 名称（默认使用第一个 sheet）"
    )

    parser.add_argument(
        "--flotherm",
        help="FloTHERM 可执行文件路径（可选，自动检测）"
    )

    parser.add_argument(
        "--no-solve",
        action="store_true",
        help="仅生成 ECXML 文件，不调用 FloTHERM 求解"
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="仅显示将要处理的配置，不实际执行"
    )

    args = parser.parse_args()

    # 打印标题
    print("""
╔════════════════════════════════════════════════════════════╗
║          Excel 多配置批量仿真工具 v2.0                      ║
║                                                            ║
║  流程: Excel长格式配置 → 修改ECXML → 批量求解              ║
║  格式: config_name | name | attribute | value              ║
╚════════════════════════════════════════════════════════════╝
""")

    # 检查模板文件
    if not os.path.exists(args.template):
        print(f"❌ 错误: 模板文件不存在: {args.template}")
        sys.exit(1)

    # 转换为绝对路径
    args.template = str(Path(args.template).resolve())
    print(f"✅ 模板文件: {args.template}")

    # 检查 Excel 文件
    if not os.path.exists(args.excel):
        print(f"❌ 错误: Excel 文件不存在: {args.excel}")
        sys.exit(1)

    print(f"✅ Excel 文件: {args.excel}")

    # 读取 Excel 配置
    print(f"\n📖 读取 Excel 配置...")
    try:
        reader = ExcelConfigReader(args.excel, args.sheet)
        configs = reader.read()
    except Exception as e:
        print(f"❌ 读取 Excel 失败: {e}")
        sys.exit(1)

    if not configs:
        print("❌ 错误: Excel 中没有找到配置数据")
        sys.exit(1)

    print(f"✅ 找到 {len(configs)} 个配置")

    # 显示配置预览
    print(f"\n📋 配置预览:")
    for i, config in enumerate(configs[:5], 1):  # 最多显示5个
        config_name = config.get('config_name', f"config_{i}")
        params = config.get('params', [])
        param_str = ', '.join([f"{p['name']}.{p['attribute']}={p['value']}" for p in params[:3]])
        if len(params) > 3:
            param_str += f" ... ({len(params)} 个参数)"
        print(f"   {i}. {config_name}: {param_str}")
    if len(configs) > 5:
        print(f"   ... 还有 {len(configs) - 5} 个配置")

    # 如果是 dry-run 模式，到此结束
    if args.dry_run:
        print("\n🔍 Dry-run 模式，不执行操作")
        sys.exit(0)

    # 创建输出文件夹（带时间戳的子文件夹）
    base_output = Path(args.output).resolve()  # 转换为绝对路径
    batch_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_folder = base_output / f"batch_{batch_timestamp}"
    output_folder.mkdir(parents=True, exist_ok=True)
    print(f"\n✅ 输出文件夹: {output_folder}")

    # 查找 FloTHERM（如果需要求解）
    flotherm_exe = None
    if not args.no_solve:
        if args.flotherm:
            flotherm_exe = args.flotherm
            if not os.path.exists(flotherm_exe):
                print(f"❌ 错误: 指定的 FloTHERM 不存在: {flotherm_exe}")
                sys.exit(1)
        else:
            flotherm_exe = find_flotherm_executable()
            if not flotherm_exe:
                print("❌ 错误: 未找到 FloTHERM，请使用 --flotherm 参数指定路径")
                sys.exit(1)

        print(f"✅ FloTHERM: {flotherm_exe}")

    # 开始处理
    print(f"\n{'='*60}")
    print(f"  开始生成 ECXML 文件")
    print(f"{'='*60}")

    results = []
    total_start_time = time.time()

    for i, config in enumerate(configs, 1):
        # 获取配置名称
        config_name = config.get('config_name', f"config_{i}")

        # 生成输出文件名
        ecxml_path = output_folder / f"{config_name}.ecxml"
        pack_path = output_folder / f"{config_name}.pack"
        html_path = output_folder / f"{config_name}_report.html"

        print(f"\n[{i}/{len(configs)}] 生成: {config_name}")

        # 应用配置到模板
        try:
            success = apply_config_to_ecxml(args.template, config, str(ecxml_path))
            if not success:
                print(f"    ⚠️ 没有修改任何参数")
        except Exception as e:
            print(f"    ❌ 修改失败: {e}")
            results.append({
                'config_name': config_name,
                'config': config,
                'ecxml_path': str(ecxml_path),
                'pack_path': '',
                'html_path': '',
                'success': False,
                'elapsed': 0,
                'message': f'修改失败: {e}'
            })
            continue

        # 求解（如果需要）
        if not args.no_solve and flotherm_exe:
            solve_success, elapsed, message = solve_ecxml(
                flotherm_exe,
                str(ecxml_path),
                str(pack_path),
                str(html_path),
                i,
                len(configs)
            )

            results.append({
                'config_name': config_name,
                'config': config,
                'ecxml_path': str(ecxml_path),
                'pack_path': str(pack_path) if solve_success else '',
                'html_path': str(html_path) if solve_success else '',
                'success': solve_success,
                'elapsed': elapsed,
                'message': message
            })
        else:
            # 仅生成模式
            results.append({
                'config_name': config_name,
                'config': config,
                'ecxml_path': str(ecxml_path),
                'pack_path': '',
                'html_path': '',
                'success': True,
                'elapsed': 0,
                'message': '仅生成 ECXML'
            })

    # 打印总结
    total_elapsed = time.time() - total_start_time
    success_count = sum(1 for r in results if r["success"])

    print(f"\n{'='*60}")
    print(f"  处理完成")
    print(f"{'='*60}")
    print(f"  总配置数: {len(results)}")
    print(f"  成功: {success_count}")
    print(f"  失败: {len(results) - success_count}")
    print(f"  总耗时: {total_elapsed:.1f} 秒")

    # 生成汇总报告
    generate_summary_report(output_folder, configs, results, Path(args.template).name)

    # 返回码
    sys.exit(0 if success_count == len(results) else 1)


if __name__ == "__main__":
    main()
